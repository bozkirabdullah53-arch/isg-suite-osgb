from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.risk_nace_roadmap import build_risk_nace_roadmap
from app.services.risk_reports import build_risk_excel, build_risk_pdf


def _company(nace_code=None):
    return SimpleNamespace(id=46, name="ALÇAT ÇELİK", nace_code=nace_code)


def test_exact_nace_468306_returns_controlled_risk_scope_without_guessing():
    roadmap = build_risk_nace_roadmap(
        _company("46.83.06"),
        coverage={"risk_records": 2, "departments": 1, "open_dofs": 1, "completed_dofs": 0},
    )

    assert roadmap["status"] == "verified"
    assert roadmap["exact_catalog_match"] is True
    assert roadmap["identity"]["code"] == "46.83.06"
    assert "Metalden prefabrik" in roadmap["identity"]["description"]
    labels = {row["label"] for row in roadmap["technical_risk_tags"]}
    assert "Kaldırma ve taşıma faaliyetleri" in labels
    assert "Yük sabitleme ve güvenli yükleme" in labels
    assert "Araç-yaya trafiği ve çarpışma" in labels
    special = {row["label"] for row in roadmap["special_risks"]}
    assert "Yük düşmesi" in special
    assert len(roadmap["report_checklist"]) >= 10
    assert len(roadmap["roadmap"]) == 10
    assert roadmap["coverage"]["risk_records"] == 2


def test_missing_or_unknown_nace_fails_closed_and_keeps_general_checklist():
    for code, expected in ((None, "missing"), ("99.99.99", "invalid"), ("46.83", "invalid")):
        roadmap = build_risk_nace_roadmap(_company(code))
        assert roadmap["status"] == expected
        assert roadmap["exact_catalog_match"] is False
        assert roadmap["identity"] is None
        assert roadmap["technical_risk_tags"] == []
        assert roadmap["special_risks"] == []
        assert roadmap["report_checklist"]
        assert roadmap["roadmap"]
        assert roadmap["warnings"]
        assert code is None or code in (roadmap["entered_nace_code"] or "")


def test_nace_roadmap_is_exported_as_additive_excel_sheet():
    roadmap = build_risk_nace_roadmap(_company("46.83.06"))
    company = SimpleNamespace(
        id=46,
        name="ALÇAT ÇELİK",
        authorized_person="İşveren",
        phone="555",
        hazard_class="Tehlikeli",
        sgk_registry_no="123",
        nace_code="46.83.06",
        risk_document_no="RD-46",
        risk_revision_no="00",
    )
    risk = SimpleNamespace(
        risk_code="RSK-0001",
        department_name="Depo",
        activity="Metal yapı elemanlarının yüklenmesi",
        hazard_id=1,
        risk_definition="Yük düşmesi",
        affected_people="Çalışanlar",
        probability=3,
        severity=4,
        risk_score=12,
        risk_level="Orta",
        term_date=None,
        existing_measures="Yük sabitleme",
        additional_measures="İstif kontrolü",
        status="Açık",
        dofs=[],
        media_files=[],
    )
    workbook = load_workbook(BytesIO(build_risk_excel(
        company=company,
        risks=[risk],
        hazard_map={1: SimpleNamespace(id=1, code="MEK-001", name="Yük düşmesi")},
        nace_roadmap=roadmap,
    )))
    assert "Risk Değerlendirme" in workbook.sheetnames
    assert "NACE Yol Haritası" in workbook.sheetnames
    ws = workbook["NACE Yol Haritası"]
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    joined = " ".join(str(value) for value in values)
    assert "46.83.06" in joined
    assert "Kaldırma ve taşıma faaliyetleri" in joined
    assert "Yol Haritası" in joined


def test_nace_roadmap_is_rendered_in_pdf_without_changing_existing_contract():
    roadmap = build_risk_nace_roadmap(_company("46.83.06"))
    company = SimpleNamespace(
        id=46,
        name="ALÇAT ÇELİK",
        authorized_person="İşveren",
        phone="555",
        address="Adres",
        hazard_class="Tehlikeli",
        sgk_registry_no="123",
        tax_number=None,
        nace_code="46.83.06",
        risk_method="5x5_l",
        risk_document_no="RD-46",
        risk_revision_no="00",
        risk_revision_reason=None,
        risk_scope_note=None,
    )
    risk = SimpleNamespace(
        risk_code="RSK-0001",
        department_name="Depo",
        activity="Metal yapı elemanlarının yüklenmesi",
        hazard_id=1,
        risk_definition="Yük düşmesi",
        affected_people="Çalışanlar",
        affected_group=None,
        probability=3,
        severity=4,
        risk_score=12,
        risk_level="Orta",
        term_date=None,
        existing_measures="Yük sabitleme",
        additional_measures="İstif kontrolü",
        status="Açık",
        revision_no=0,
        dofs=[],
    )
    pdf = build_risk_pdf(
        company=company,
        risks=[risk],
        hazard_map={1: SimpleNamespace(id=1, code="MEK-001", name="Yük düşmesi", regulations='[]')},
        prepared_by="Uzman",
        nace_code="46.83.06",
        nace_roadmap=roadmap,
    )
    assert pdf.startswith(b"%PDF")
    assert len(pdf) > 3500
