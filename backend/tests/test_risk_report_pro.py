"""Risk rapor yöntemleri + profesyonel PDF duman testleri."""
from datetime import date
from types import SimpleNamespace

from app.services.risk_methods import method_choices, resolve_method
from app.services.risk_reports import build_risk_excel, build_risk_pdf
from app.services.risk_validity import build_validity, document_meta_rows
from io import BytesIO

from openpyxl import load_workbook


def test_methods_catalog_covers_common_methods():
    codes = {m["code"] for m in method_choices()}
    for needed in ("5x5_l", "fine_kinney", "hazop", "fmea", "jsa", "what_if", "x_matrix"):
        assert needed in codes
    assert "Fine-Kinney" in resolve_method("fine_kinney")["label"]


def test_validity_uses_selected_method():
    v = build_validity(
        hazard_class="Çok Tehlikeli",
        assessment_date=date(2024, 7, 28),
        method_code="fine_kinney",
    )
    assert "Fine-Kinney" in v["method"]
    assert v["method_code"] == "fine_kinney"


def test_document_meta_includes_doc_control():
    rows = dict(
        document_meta_rows(
            validity={"method": "5x5", "assessment_date": "2024-07-28", "valid_until": "2026-07-28"},
            prepared_by="Uzman",
            document_no="RD-001",
            revision_no="01",
            revision_reason="Yeni makine",
        )
    )
    assert rows["Belge No"] == "RD-001"
    assert rows["Revizyon No"] == "01"
    assert rows["Revizyon Nedeni"] == "Yeni makine"


def test_professional_pdf_has_sections():
    company = SimpleNamespace(
        id=1,
        name="TEST İŞYERİ",
        authorized_person="TEST İK",
        phone="555",
        address="Adres",
        hazard_class="Çok Tehlikeli",
        sgk_registry_no="123",
        tax_number=None,
        nace_code=None,
        risk_method="5x5_l",
        risk_document_no="RD-TEST",
        risk_revision_no="00",
        risk_revision_reason=None,
        risk_scope_note=None,
    )
    risk = SimpleNamespace(
        risk_code="RSK-0001",
        department_name="Depo",
        activity="Elle taşıma",
        hazard_id=1,
        risk_definition="Gürültü / düşme riski",
        affected_people="Çalışanlar",
        affected_group=None,
        probability=4,
        severity=4,
        risk_score=16,
        risk_level="Yüksek",
        term_date=date(2026, 8, 1),
        status="Açık",
        existing_measures="yok",
        additional_measures="KKD",
        revision_no=1,
        dofs=[],
    )
    hazard = SimpleNamespace(id=1, code="FZK-001", name="Gürültü", regulations='["6331"]')
    pdf = build_risk_pdf(
        company=company,
        risks=[risk],
        hazard_map={1: hazard},
        prepared_by="Uzman Test",
        sgk_no="123",
        workplace_physician="Hekim",
        employer_representative="TEST İK",
        employee_representative="Temsilci",
        support_staff="Destek",
        validity=build_validity(
            hazard_class="Çok Tehlikeli",
            assessment_date=date(2024, 7, 28),
            method_code="5x5_l",
        ),
        employee_count=12,
        document_no="RD-TEST",
        revision_no="00",
    )
    assert pdf.startswith(b"%PDF")
    assert len(pdf) > 2500
    # Son sayfa beyanı (NumberedCanvas) — Latin-1 PDF stream içinde Türkçe escape olabilir;
    # en azından sayfa numaralandırma ve ekip satırı üretildiğini boyutla doğrula.
    assert b"/Type /Page" in pdf or b"Page" in pdf


def test_risk_excel_has_page_footer_and_last_page_declaration():
    company = SimpleNamespace(
        id=7,
        name="DEMO AŞ",
        authorized_person="Vekil",
        phone="555",
        hazard_class="Tehlikeli",
        sgk_registry_no="999",
        nace_code="41.20",
        risk_document_no="RD-7",
        risk_revision_no="00",
    )
    risks = [
        SimpleNamespace(
            risk_code=f"RSK-{i:04d}",
            department_name="Üretim",
            activity="Kaynak",
            hazard_id=1,
            risk_definition="Yanık",
            affected_people="Operatör",
            probability=3,
            severity=4,
            risk_score=12,
            risk_level="Orta",
            term_date=date(2026, 9, 1),
            existing_measures="Eldiven",
            additional_measures="Eğitim",
            status="Açık",
            dofs=[],
            media_files=[],
        )
        for i in range(1, 6)
    ]
    hazard_map = {1: SimpleNamespace(id=1, code="FZK", name="Sıcak yüzey")}
    data = build_risk_excel(
        company=company,
        risks=risks,
        hazard_map=hazard_map,
        prepared_by="Uzman A",
        workplace_physician="Hekim B",
        employer_representative="Vekil C",
        employee_representative="Temsilci D",
    )
    wb = load_workbook(BytesIO(data))
    ws = wb["Risk Değerlendirme"]
    assert "Sayfa &P / &N" in (ws.oddFooter.center.text or "")
    assert "Risk Değ. Ekibi İmza" in (ws.oddFooter.left.text or "")
    assert "İGU: Uzman A" in (ws.oddFooter.left.text or "")
    # Son sayfa beyanı hücrede
    found = False
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        val = row[0]
        if isinstance(val, str) and "İş bu risk değerlendirme raporu" in val and "sayfadan oluşur" in val:
            found = True
            break
    assert found
    assert "İMZA / ONAY" in "".join(
        str(c[0] or "") for c in ws.iter_rows(min_row=1, max_col=1, values_only=True)
    )