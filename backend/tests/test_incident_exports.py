"""Olay Excel / CAPA board builder smoke tests."""
from datetime import date
from types import SimpleNamespace

from openpyxl import load_workbook
from io import BytesIO

from app.services.incident_reports import build_capa_board_excel, build_incident_excel


def test_build_incident_excel_headers_and_row():
    row = SimpleNamespace(
        form_no="ISG-RK-2026-0001",
        company_id=1,
        event_type="ramak_kala",
        status="Aktif",
        event_date=date(2026, 7, 1),
        event_time="10:00",
        location="Atölye",
        department="Üretim",
        area="A",
        classification="Tehlikeli durum",
        short_summary="Kayma riski gözlendi",
        probability=3,
        severity=4,
        risk_score=12,
        risk_level="Orta",
        root_cause=SimpleNamespace(root_cause="Zemin ıslak"),
        dofs=[SimpleNamespace()],
        sgk_reported=False,
        police_reported=False,
        safety_specialist="Uzman A",
        workplace_physician="Hekim B",
        employer_representative="Vekil C",
        recorded_by_name="Kayıtçı",
    )
    data = build_incident_excel(rows=[row], company_names={1: "Demo AŞ"})
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert ws["A1"].value == "Form No"
    assert ws["A2"].value == "ISG-RK-2026-0001"
    assert ws["B2"].value == "Demo AŞ"
    assert ws["P2"].value == "Zemin ıslak"
    assert ws["Q2"].value == 1


def test_build_capa_board_excel():
    data = build_capa_board_excel(
        rows=[
            {
                "source": "Olay",
                "code": "ISG-DOF-1",
                "parent": "ISG-RK-1",
                "parentSummary": "Özet",
                "title": "Tespit",
                "action": "Düzelt",
                "responsible": "Ali",
                "term": "2026-08-01",
                "status": "Açık",
                "priority": "Yüksek",
            }
        ]
    )
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert ws["A1"].value == "Kaynak"
    assert ws["B2"].value == "ISG-DOF-1"
