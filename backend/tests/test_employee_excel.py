"""Personel Excel şablon / import smoke."""
from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.services.employee_excel import (
    TEMPLATE_DATA_START,
    TEMPLATE_HEADERS,
    build_import_template_xlsx,
    map_header,
    parse_employees_workbook,
)


def test_map_adi_soyadi_variants():
    assert map_header("Adı Soyadı") == "full_name"
    assert map_header("Ad Soyad") == "full_name"
    assert map_header("ADI SOYADI") == "full_name"
    assert map_header("TC") == "national_id_masked"
    assert map_header("TC Kimlik No") == "national_id_masked"
    assert map_header("Görevi") == "job_title"
    assert map_header("İşe Giriş Tarihi") == "start_date"
    assert map_header("Engelli/Hükümlü") == "special_status"
    assert map_header("Engelli/Hükümlü Durumu") == "special_status"


def test_blank_template_does_not_import_placeholder_rows():
    rows = parse_employees_workbook(build_import_template_xlsx())
    assert rows == []


def test_filled_template_roundtrip():
    raw = build_import_template_xlsx()
    wb = load_workbook(BytesIO(raw))
    ws = wb["Personel"]
    assert list(ws.tables) == ["PersonelListesi"]
    assert [ws.cell(TEMPLATE_DATA_START - 1, i).value for i in range(1, 6)] == TEMPLATE_HEADERS
    ws.cell(TEMPLATE_DATA_START, 1, "Ali Veli")
    ws.cell(TEMPLATE_DATA_START, 2, "12345678901")
    ws.cell(TEMPLATE_DATA_START, 3, "Kaynakçı")
    ws.cell(TEMPLATE_DATA_START, 4, date(2024, 1, 15))
    ws.cell(TEMPLATE_DATA_START, 5, "Yok")
    ws.cell(TEMPLATE_DATA_START + 1, 1, "Ayşe Yılmaz")
    ws.cell(TEMPLATE_DATA_START + 1, 3, "Operatör")
    ws.cell(TEMPLATE_DATA_START + 1, 5, "Engelli")
    buf = BytesIO()
    wb.save(buf)
    wb.close()

    rows = parse_employees_workbook(buf.getvalue())
    assert [row["full_name"] for row in rows] == ["Ali Veli", "Ayşe Yılmaz"]
    assert rows[0]["national_id_masked"] == "12345678901"
    assert rows[0]["job_title"] == "Kaynakçı"
    assert rows[0]["start_date"] == date(2024, 1, 15)
    assert rows[0]["special_status"] is None
    assert rows[1]["special_status"] == "Engelli"


def test_legacy_simple_workbook_still_imports():
    wb = Workbook()
    ws = wb.active
    ws.append(["Adı Soyadı", "TC Kimlik", "Görevi", "İşe Giriş Tarihi", "Engelli/Hükümlü Durumu"])
    ws.append(["Ali Veli", "12345678901", "Kaynakçı", "2024-01-15", ""])
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    rows = parse_employees_workbook(buf.getvalue())
    assert rows[0]["full_name"] == "Ali Veli"
    assert rows[0]["job_title"] == "Kaynakçı"


def test_numeric_tc_suffix_is_removed_from_personnel_import():
    wb = Workbook()
    ws = wb.active
    ws.append(["Adı Soyadı", "TC Kimlik"])
    ws.append(["Ali Veli", 26230266894.0])
    buf = BytesIO()
    wb.save(buf)
    wb.close()

    rows = parse_employees_workbook(buf.getvalue())
    assert rows[0]["national_id_masked"] == "26230266894"
