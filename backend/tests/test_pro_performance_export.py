"""Professional performance CSV compatibility + formatted XLSX exports."""
from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


@pytest.fixture()
def client(tmp_path, monkeypatch):
    db_file = tmp_path / "pro_perf_export.db"
    url = f"sqlite:///{db_file.as_posix()}"
    monkeypatch.setenv("DATABASE_URL", url)
    monkeypatch.setenv("ENVIRONMENT", "development")
    monkeypatch.setenv("SECRET_KEY", "test-secret-key-at-least-32-chars-long!!")
    monkeypatch.setattr("app.api.auth.role_requires_mfa", lambda _role: False)

    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    import app.core.database as dbmod
    import app.models.entities as ent
    from app.core.config import settings

    settings.database_url = url
    settings.secret_key = "test-secret-key-at-least-32-chars-long!!"
    settings.environment = "development"

    engine = create_engine(url, connect_args={"check_same_thread": False})
    dbmod.engine = engine
    dbmod.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    ent.Base.metadata.create_all(bind=engine)

    from app.main import app

    return TestClient(app)


def _seed(client: TestClient) -> tuple[str, dict]:
    from app.core.database import SessionLocal
    from app.core.security import get_password_hash
    from app.models.entities import (
        AssignmentStatus,
        Company,
        IsgProfessional,
        OsgbOrganization,
        ProfessionalType,
        User,
        UserRole,
        WorkplaceAssignment,
    )

    with SessionLocal() as db:
        osgb = OsgbOrganization(
            name="Perf OSGB",
            authorization_number="YETKI-PERF-1",
            tax_number="9876543210",
            responsible_manager="Yonetici",
            email="perf-osgb@test.com",
            is_active=True,
        )
        db.add(osgb)
        db.flush()
        company = Company(name="Perf Firma", osgb_id=osgb.id, is_active=True, hazard_class="Tehlikeli")
        db.add(company)
        db.flush()
        pro = IsgProfessional(
            osgb_id=osgb.id,
            full_name="Zeynep Uzman",
            professional_type=ProfessionalType.SAFETY_SPECIALIST,
            certificate_class="A",
            certificate_number="UZM-PERF",
            is_active=True,
        )
        db.add(pro)
        db.flush()
        db.add(
            WorkplaceAssignment(
                osgb_id=osgb.id,
                company_id=company.id,
                professional_id=pro.id,
                professional_type=ProfessionalType.SAFETY_SPECIALIST,
                status=AssignmentStatus.ACTIVE,
                start_date=date.today() - timedelta(days=10),
                required_minutes_monthly=400,
                planned_minutes_monthly=400,
                actual_minutes_monthly=100,
            )
        )
        db.add(
            User(
                email="perf-admin@test.com",
                full_name="Perf Admin",
                hashed_password=get_password_hash("TestPass123!"),
                role=UserRole.COMPANY_ADMIN,
                osgb_id=osgb.id,
                is_active=True,
            )
        )
        db.commit()
        seed = {"osgb_id": osgb.id, "professional_id": pro.id}

    r = client.post(
        "/api/v1/auth/login",
        json={"email": "perf-admin@test.com", "password": "TestPass123!"},
    )
    assert r.status_code == 200, r.text
    return r.json()["access_token"], seed


def _sheet_values(ws) -> list[object]:
    return [cell.value for row in ws.iter_rows() for cell in row if cell.value not in (None, "")]


def test_health_flag_pro_performance_export(release_flags):
    body = release_flags
    assert body.get("version")
    assert body["pro_performance_export"] == "csv-v1"
    assert body["pro_performance_excel"] == "xlsx-v1"
    assert body["csgb_company_snapshot"] == "read-only-v1"


def test_performance_roster_csv_remains_backward_compatible(client):
    token, seed = _seed(client)
    headers = {"Authorization": f"Bearer {token}"}
    r = client.get(
        f"/api/v1/osgb/professionals/performance/export.csv?osgb_id={seed['osgb_id']}",
        headers=headers,
    )
    assert r.status_code == 200, r.text
    assert "text/csv" in (r.headers.get("content-type") or "")
    text = r.content.decode("utf-8-sig")
    assert "professional_id" in text
    assert "Zeynep Uzman" in text


def test_performance_detail_csv_remains_backward_compatible(client):
    token, seed = _seed(client)
    headers = {"Authorization": f"Bearer {token}"}
    pid = seed["professional_id"]
    r = client.get(f"/api/v1/osgb/professionals/{pid}/performance/export.csv", headers=headers)
    assert r.status_code == 200, r.text
    text = r.content.decode("utf-8-sig")
    assert "row_type" in text
    assert "Zeynep Uzman" in text


def test_performance_roster_xlsx_is_readable_and_multisheet(client):
    token, seed = _seed(client)
    headers = {"Authorization": f"Bearer {token}"}
    r = client.get(
        f"/api/v1/osgb/professionals/performance/export.xlsx?osgb_id={seed['osgb_id']}",
        headers=headers,
    )
    assert r.status_code == 200, r.text
    content_type = r.headers.get("content-type") or ""
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in content_type
    assert ".xlsx" in (r.headers.get("content-disposition") or "")
    assert r.content[:2] == b"PK"

    wb = load_workbook(BytesIO(r.content), data_only=True)
    try:
        assert wb.sheetnames == ["Performans Özeti", "Eksik Kontroller", "Kontrol Özeti", "Açıklamalar"]
        ws = wb["Performans Özeti"]
        assert ws["A1"].value == "ÇSGB Profesyonel Performans Raporu"
        assert ws["A9"].value == "Sıra"
        assert ws["B9"].value == "Ad Soyad"
        assert ws["C9"].value == "Unvan"
        assert ws.freeze_panes == "A10"
        values = _sheet_values(ws)
        assert "Zeynep Uzman" in values
        assert "İş Güvenliği Uzmanı" in values
        assert not any(isinstance(value, str) and "professional_id,full_name" in value for value in values)

        gap_values = _sheet_values(wb["Eksik Kontroller"])
        assert "Zeynep Uzman" in gap_values
        assert "Perf Firma" in gap_values
        assert "Açıklama" in gap_values
        assert "Mevzuat Dayanağı" in gap_values
    finally:
        wb.close()


def test_performance_detail_xlsx_contains_summary_and_checklists(client):
    token, seed = _seed(client)
    headers = {"Authorization": f"Bearer {token}"}
    pid = seed["professional_id"]
    r = client.get(f"/api/v1/osgb/professionals/{pid}/performance/export.xlsx", headers=headers)
    assert r.status_code == 200, r.text
    assert ".xlsx" in (r.headers.get("content-disposition") or "")

    wb = load_workbook(BytesIO(r.content), data_only=True)
    try:
        assert wb.sheetnames == ["Profesyonel Özeti", "Eksik Kontroller", "Tamamlanan", "Firma Checklist"]
        summary = wb["Profesyonel Özeti"]
        assert summary["A1"].value == "Profesyonel Performans Raporu — Zeynep Uzman"
        assert "Zeynep Uzman" in _sheet_values(summary)
        assert "İş Güvenliği Uzmanı" in _sheet_values(summary)

        gaps = wb["Eksik Kontroller"]
        gap_values = _sheet_values(gaps)
        assert "Perf Firma" in gap_values
        assert "Eksik" in gap_values

        checklist = wb["Firma Checklist"]
        checklist_values = _sheet_values(checklist)
        assert "Firma Bazlı Checklist" in checklist_values
        assert "Firma Puanı (%)" in checklist_values
        assert "Perf Firma" in checklist_values
    finally:
        wb.close()
