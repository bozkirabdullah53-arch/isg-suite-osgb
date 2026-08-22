"""Yetkili firma yönetimi: tenant, tarih, mahremiyet, skor ve çıktı kabul testleri."""
from __future__ import annotations

import json
import zipfile
from datetime import date, timedelta
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


@pytest.fixture()
def client(tmp_path, monkeypatch):
    db_file = tmp_path / "authorized-firms.db"
    url = f"sqlite:///{db_file.as_posix()}"
    monkeypatch.setenv("DATABASE_URL", url)
    monkeypatch.setenv("ENVIRONMENT", "development")
    monkeypatch.setenv("SECRET_KEY", "test-secret-key-at-least-32-chars-long!!")
    monkeypatch.setattr("app.api.auth.role_requires_mfa", lambda _role: False)

    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    import app.core.database as dbmod
    import app.models.authorized_firm  # noqa: F401
    import app.models.entities as ent
    from app.core.config import settings

    settings.database_url = url
    settings.secret_key = "test-secret-key-at-least-32-chars-long!!"
    settings.environment = "development"
    engine = create_engine(url, connect_args={"check_same_thread": False})
    dbmod.engine = engine
    dbmod.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    ent.Base.metadata.create_all(bind=engine)

    from app.main import app

    return TestClient(app)


def _seed() -> dict:
    from app.core.database import SessionLocal
    from app.core.security import get_password_hash
    from app.models.entities import (
        AssignmentStatus,
        Company,
        Employee,
        HealthFitnessStatus,
        HealthRecord,
        HealthRecordType,
        IsgProfessional,
        OsgbOrganization,
        ProfessionalType,
        ServiceContract,
        User,
        UserRole,
        WorkplaceAssignment,
    )

    password = "TestPass123!"
    today = date.today()
    with SessionLocal() as db:
        first = OsgbOrganization(name="Mavi OSGB", authorization_number="MAVI-OSGB", is_active=True)
        second = OsgbOrganization(name="Turuncu OSGB", authorization_number="TRNC-OSGB", is_active=True)
        db.add_all([first, second])
        db.flush()
        own = Company(
            name="Korunaklı Fabrika",
            osgb_id=first.id,
            is_active=True,
            hazard_class="Tehlikeli",
            address="Güvenli Mahalle No 10",
            authorized_person="Ayşe Yönetici",
        )
        legacy = Company(name="Eski Akış İşyeri", osgb_id=first.id, is_active=True, hazard_class="Az Tehlikeli")
        foreign = Company(name="Yabancı Fabrika", osgb_id=second.id, is_active=True, hazard_class="Tehlikeli")
        db.add_all([own, legacy, foreign])
        db.flush()
        admin = User(
            email="admin@mavi.example.com",
            full_name="Mavi OSGB Admin",
            hashed_password=get_password_hash(password),
            role=UserRole.COMPANY_ADMIN,
            osgb_id=first.id,
            is_active=True,
        )
        foreign_admin = User(
            email="admin@turuncu.example.com",
            full_name="Turuncu OSGB Admin",
            hashed_password=get_password_hash(password),
            role=UserRole.COMPANY_ADMIN,
            osgb_id=second.id,
            is_active=True,
        )
        workplace_admin = User(
            email="isyeri@mavi.example.com",
            full_name="İşyeri Yöneticisi",
            hashed_password=get_password_hash(password),
            role=UserRole.COMPANY_ADMIN,
            company_id=own.id,
            osgb_id=first.id,
            is_active=True,
        )
        global_admin = User(
            email="platform@example.com",
            full_name="Platform Yöneticisi",
            hashed_password=get_password_hash(password),
            role=UserRole.GLOBAL_ADMIN,
            is_active=True,
        )
        db.add_all([admin, foreign_admin, workplace_admin, global_admin])
        db.flush()
        professionals = []
        for index, name in enumerate(("Umut Uzman", "Selin Uzman", "Deniz Uzman"), start=1):
            professional = IsgProfessional(
                osgb_id=first.id,
                full_name=name,
                email=f"uzman{index}@mavi.example.com",
                professional_type=ProfessionalType.SAFETY_SPECIALIST,
                certificate_class="A",
                certificate_number=f"A-{index}00",
                certificate_date=today - timedelta(days=365),
                is_active=True,
            )
            db.add(professional)
            db.flush()
            professionals.append(professional)
        contract = ServiceContract(
            osgb_id=first.id,
            company_id=own.id,
            contract_number="HS-100",
            start_date=today - timedelta(days=30),
            end_date=today + timedelta(days=120),
            status="active",
        )
        db.add(contract)
        db.flush()
        db.add(
            WorkplaceAssignment(
                osgb_id=first.id,
                company_id=own.id,
                professional_id=professionals[0].id,
                professional_type=ProfessionalType.SAFETY_SPECIALIST,
                start_date=today,
                end_date=today + timedelta(days=90),
                required_minutes_monthly=600,
                planned_minutes_monthly=600,
                actual_minutes_monthly=600,
                isg_katip_contract_number="K-100",
                status=AssignmentStatus.ACTIVE,
            )
        )
        employee = Employee(company_id=own.id, full_name="Mahrem Çalışan", is_active=True)
        db.add(employee)
        db.flush()
        db.add(
            HealthRecord(
                company_id=own.id,
                employee_id=employee.id,
                record_type=HealthRecordType.PERIODIC_EXAM,
                examination_date=today - timedelta(days=300),
                next_examination_date=today + timedelta(days=30),
                fitness_status=HealthFitnessStatus.CONDITIONAL,
                summary="YETKILI_FIRMA_GIZLI_TANI",
                confidential_note="YETKILI_FIRMA_GIZLI_HEKIM_NOTU",
                restrictions="YETKILI_FIRMA_GIZLI_KISIT",
                created_by_id=admin.id,
            )
        )
        db.commit()
        return {
            "password": password,
            "osgb": first.id,
            "foreign_osgb": second.id,
            "company": own.id,
            "legacy_company": legacy.id,
            "foreign_company": foreign.id,
            "contract": contract.id,
            "professionals": [item.id for item in professionals],
            "admin": admin.email,
            "foreign_admin": foreign_admin.email,
            "workplace_admin": workplace_admin.email,
            "global_admin": global_admin.email,
        }


def _token(client: TestClient, email: str, password: str) -> str:
    response = client.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200, response.text
    return response.json()["access_token"]


def _headers(client: TestClient, seed: dict, key: str = "admin") -> dict:
    return {"Authorization": f"Bearer {_token(client, seed[key], seed['password'])}"}


def _create_profile(client: TestClient, seed: dict, *, headers: dict | None = None, name: str = "Korunaklı Yetkili Firma") -> dict:
    today = date.today()
    response = client.post(
        "/api/v1/authorized-firms",
        headers=headers or _headers(client, seed),
        json={
            "osgb_id": seed["osgb"],
            "company_id": seed["company"],
            "firm_name": name,
            "firm_type": "OSGB müşterisi",
            "province": "Ankara",
            "district": "Çankaya",
            "address": "Güvenli Mahalle No 10",
            "authorized_representative": "Ayşe Yönetici",
            "contact_email": "yetkili@example.com",
            "employee_count_declared": 1,
            "hazard_class": "Tehlikeli",
            "authorization_scope": "İş sağlığı ve güvenliği hizmet kapsamı",
            "authorization_number": "YF-100",
            "authorization_issue_date": (today - timedelta(days=30)).isoformat(),
            "authorization_start_date": today.isoformat(),
            "authorization_expiry_date": (today + timedelta(days=60)).isoformat(),
            "review_state": "internal_record",
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


def test_tenant_isolation_and_company_bound_admin_block(client):
    seed = _seed()
    own_headers = _headers(client, seed)
    profile = _create_profile(client, seed, headers=own_headers)

    own = client.get(f"/api/v1/authorized-firms/{profile['id']}", headers=own_headers)
    assert own.status_code == 200
    assert own.headers["cache-control"] == "no-store"
    assert own.json()["company_id"] == seed["company"]
    partial = client.patch(
        f"/api/v1/authorized-firms/{profile['id']}",
        headers=own_headers,
        json={"firm_name": "Korunaklı Firma Güncel"},
    )
    assert partial.status_code == 200, partial.text
    assert partial.json()["province"] == "Ankara"
    assert partial.json()["authorization_number"] == "YF-100"

    foreign_headers = _headers(client, seed, "foreign_admin")
    foreign = client.get(f"/api/v1/authorized-firms/{profile['id']}", headers=foreign_headers)
    assert foreign.status_code == 404
    foreign_by_company = client.get(
        f"/api/v1/authorized-firms/by-company/{seed['company']}", headers=foreign_headers
    )
    assert foreign_by_company.status_code == 404
    foreign_list = client.get("/api/v1/authorized-firms", headers=foreign_headers)
    assert foreign_list.status_code == 200
    assert foreign_list.headers["cache-control"] == "no-store"
    assert foreign_list.json()["total"] == 0

    workplace_headers = _headers(client, seed, "workplace_admin")
    blocked = client.get("/api/v1/authorized-firms", headers=workplace_headers)
    assert blocked.status_code == 403
    comparison = client.get("/api/v1/authorized-firms/comparison", headers=own_headers)
    assert comparison.status_code == 403
    global_comparison = client.get(
        "/api/v1/authorized-firms/comparison",
        headers=_headers(client, seed, "global_admin"),
    )
    assert global_comparison.status_code == 200, global_comparison.text
    assert global_comparison.headers["cache-control"] == "no-store"
    assert global_comparison.json()["privacy"]["sensitive_fields_exposed"] is False


def test_invalid_profile_and_contract_date_ranges_are_rejected(client):
    seed = _seed()
    headers = _headers(client, seed)
    today = date.today()
    invalid = client.post(
        "/api/v1/authorized-firms",
        headers=headers,
        json={
            "osgb_id": seed["osgb"],
            "company_id": seed["company"],
            "firm_name": "Tarih Kontrollü Firma",
            "authorization_start_date": (today + timedelta(days=2)).isoformat(),
            "authorization_expiry_date": today.isoformat(),
        },
    )
    assert invalid.status_code == 422

    contract = client.post(
        "/api/v1/osgb/contracts",
        headers=headers,
        json={
            "osgb_id": seed["osgb"],
            "company_id": seed["company"],
            "contract_number": "TERS-1",
            "start_date": (today + timedelta(days=10)).isoformat(),
            "end_date": today.isoformat(),
        },
    )
    assert contract.status_code == 422


def test_alerts_professional_status_visible_score_history_and_privacy(client):
    seed = _seed()
    headers = _headers(client, seed)
    profile = _create_profile(client, seed, headers=headers)
    today = date.today()

    document = client.post(
        f"/api/v1/authorized-firms/{profile['id']}/documents",
        headers=headers,
        json={
            "document_type": "yetki_belgesi",
            "title": "Yetki belgesi",
            "mandatory": True,
            "start_date": today.isoformat(),
            "expiry_date": (today + timedelta(days=30)).isoformat(),
        },
    )
    assert document.status_code == 200, document.text

    compliance = client.put(
        f"/api/v1/authorized-firms/{profile['id']}/professionals/{seed['professionals'][0]}/compliance",
        headers=headers,
        json={
            "certificate_issue_date": (today - timedelta(days=500)).isoformat(),
            "certificate_expiry_date": (today - timedelta(days=1)).isoformat(),
            "document_review_date": (today - timedelta(days=40)).isoformat(),
            "document_renewal_date": today.isoformat(),
            "required_documents_status": "complete",
        },
    )
    assert compliance.status_code == 200, compliance.text
    assert compliance.json()["status"] == "expired_documents"

    snapshot = client.post(
        f"/api/v1/authorized-firms/{profile['id']}/score-snapshots", headers=headers
    )
    assert snapshot.status_code == 200
    detail = client.get(f"/api/v1/authorized-firms/{profile['id']}", headers=headers)
    assert detail.status_code == 200
    payload = detail.json()
    serialized = detail.text
    assert "YETKILI_FIRMA_GIZLI_TANI" not in serialized
    assert "YETKILI_FIRMA_GIZLI_HEKIM_NOTU" not in serialized
    assert "YETKILI_FIRMA_GIZLI_KISIT" not in serialized
    assert payload["workplace_status"]["privacy"] == {
        "health_mode": "aggregate_only",
        "sensitive_fields_exposed": False,
    }
    assert payload["authorization_validity"]["code"] == "due_60"
    assert any(alert["status"] == "due_30" for alert in payload["alerts"])
    assert payload["professionals"][0]["status"] == "expired_documents"
    score = payload["compliance_score"]
    assert score["black_box"] is False
    assert score["calculation"]
    assert len(score["categories"]) == 10
    assert all({"score", "weight", "detail", "recommended_action"} <= set(item) for item in score["categories"])
    assert len(payload["onboarding"]["steps"]) == 11
    assert payload["automatic_task_checklist"]
    assert payload["score_history"]

    refresh = client.post(f"/api/v1/notifications/refresh?osgb_id={seed['osgb']}", headers=headers)
    assert refresh.status_code == 200, refresh.text
    notifications = client.get("/api/v1/notifications", headers=headers)
    assert notifications.status_code == 200
    assert any(item["entity_type"] in {"authorized_firm_profile", "authorized_firm_document", "professional_compliance"} for item in notifications.json())


def test_assignment_must_fit_active_contract_only_for_authorized_profile(client):
    seed = _seed()
    headers = _headers(client, seed)
    profile = _create_profile(client, seed, headers=headers)
    today = date.today()

    outside = client.post(
        "/api/v1/osgb/assignments",
        headers=headers,
        json={
            "osgb_id": seed["osgb"],
            "company_id": seed["company"],
            "professional_id": seed["professionals"][1],
            "professional_type": "safety_specialist",
            "start_date": today.isoformat(),
            "end_date": (today + timedelta(days=121)).isoformat(),
            "isg_katip_contract_number": "K-200",
        },
    )
    assert outside.status_code == 400
    assert "sözleşme" in outside.text.lower()

    inside = client.post(
        "/api/v1/osgb/assignments",
        headers=headers,
        json={
            "osgb_id": seed["osgb"],
            "company_id": seed["company"],
            "professional_id": seed["professionals"][1],
            "professional_type": "safety_specialist",
            "start_date": today.isoformat(),
            "end_date": (today + timedelta(days=60)).isoformat(),
            "isg_katip_contract_number": "K-201",
        },
    )
    assert inside.status_code == 200, inside.text

    indefinite = client.post(
        "/api/v1/osgb/assignments",
        headers=headers,
        json={
            "osgb_id": seed["osgb"],
            "company_id": seed["company"],
            "professional_id": seed["professionals"][2],
            "professional_type": "safety_specialist",
            "start_date": today.isoformat(),
            "isg_katip_contract_number": "K-202",
        },
    )
    assert indefinite.status_code == 400

    legacy = client.post(
        "/api/v1/osgb/assignments",
        headers=headers,
        json={
            "osgb_id": seed["osgb"],
            "company_id": seed["legacy_company"],
            "professional_id": seed["professionals"][2],
            "professional_type": "safety_specialist",
            "start_date": today.isoformat(),
            "isg_katip_contract_number": "K-LEGACY",
        },
    )
    assert legacy.status_code == 200, legacy.text

    shorten = client.patch(
        f"/api/v1/osgb/contracts/{seed['contract']}",
        headers=headers,
        json={"end_date": (today + timedelta(days=30)).isoformat()},
    )
    assert shorten.status_code == 400
    detail = client.get(f"/api/v1/authorized-firms/{profile['id']}", headers=headers)
    assert detail.status_code == 200
    assert any(item["contract_covered"] for item in detail.json()["assignments"])


def test_pdf_excel_status_and_inspection_exports_are_safe(client):
    seed = _seed()
    headers = _headers(client, seed)
    profile = _create_profile(client, seed, headers=headers, name="=HYPERLINK Formula Firma")
    today = date.today()
    document = client.post(
        f"/api/v1/authorized-firms/{profile['id']}/documents",
        headers=headers,
        json={
            "document_type": "yetki",
            "title": "+HYPERLINK Kötü Belge",
            "mandatory": True,
            "start_date": today.isoformat(),
            "expiry_date": (today + timedelta(days=90)).isoformat(),
            "notes": "<script>çalışmaz</script> & güvenli",
        },
    )
    assert document.status_code == 200, document.text

    pdf = client.get(f"/api/v1/authorized-firms/{profile['id']}/export.pdf", headers=headers)
    assert pdf.status_code == 200, pdf.text
    assert pdf.content.startswith(b"%PDF")
    assert pdf.headers["cache-control"] == "no-store"

    excel = client.get(f"/api/v1/authorized-firms/{profile['id']}/export.xlsx", headers=headers)
    assert excel.status_code == 200, excel.text
    workbook = load_workbook(BytesIO(excel.content), data_only=False)
    assert workbook["Firma Kartı"]["B1"].value.startswith("'=")
    assert workbook["Belgeler"]["A2"].value.startswith("'+")

    status = client.get("/api/v1/authorized-firms/status-report.xlsx", headers=headers)
    assert status.status_code == 200
    status_workbook = load_workbook(BytesIO(status.content), data_only=False)
    assert status_workbook["Yetkili Firma Durumu"]["A2"].value.startswith("'=")

    package = client.get(
        f"/api/v1/authorized-firms/{profile['id']}/inspection-package.zip", headers=headers
    )
    assert package.status_code == 200
    with zipfile.ZipFile(BytesIO(package.content)) as archive:
        names = archive.namelist()
        assert any(name.endswith(".pdf") for name in names)
        assert any(name.endswith(".xlsx") for name in names)
        manifest = json.loads(archive.read("eksik-kontrol-listesi.json"))
        assert manifest["external_submission_performed"] is False
        assert manifest["privacy"]["sensitive_fields_exposed"] is False


def test_onboarding_requires_all_eleven_steps_before_completion(client):
    seed = _seed()
    headers = _headers(client, seed)
    profile = _create_profile(client, seed, headers=headers)

    invalid = client.patch(
        f"/api/v1/authorized-firms/{profile['id']}/onboarding",
        headers=headers,
        json={"current_step": 11, "completed_steps": [1, 2, 3], "status": "completed"},
    )
    assert invalid.status_code == 422

    completed = client.patch(
        f"/api/v1/authorized-firms/{profile['id']}/onboarding",
        headers=headers,
        json={"current_step": 11, "completed_steps": list(range(1, 12)), "status": "completed"},
    )
    assert completed.status_code == 200, completed.text
    assert completed.json()["onboarding"]["status"] == "completed"
    assert completed.json()["onboarding"]["completed_steps"] == list(range(1, 12))
