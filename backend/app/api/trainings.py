import hashlib
import logging
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import or_, select
from sqlalchemy.orm import Session, selectinload

from app.api.company_access import (
    accessible_company_ids_or_empty,
    effective_company_id,
    ensure_company_access,
)
from app.api.deps import get_current_user
from app.api.files import safe_upload_root
from app.core.config import settings
from app.core.database import get_db
from app.models.entities import (
    Branch,
    Company,
    Employee,
    TrainingParticipant,
    TrainingSession,
    TrainingStatus,
    User,
    UserRole,
)
from app.models.training_presentation_approval import TrainingPresentationApproval
from app.schemas.training import (
    TrainingArchiveRequest,
    TrainingCreate,
    TrainingResponse,
    TrainingUpdate,
    TrainingVerifyResponse,
)
from app.services.assigned_team import assigned_team, training_defaults
from app.services.training_employee_import import resolve_or_create_employees
from app.services.training_excel import parse_employee_upload
from app.services.training_pdfs import build_attendance_pdf, build_certificates_pdf
from app.services.training_exam_pdf import build_exam_pdf
from app.services.upload_gateway import delete_relative, persist_relative
from app.services.upload_security import assert_safe_upload
from app.services.training_topics import (
    meta_payload,
    sektor_kodu_cozumle,
    sektor_tehlike_sinifi,
    sectors_list_for_api,
)
from app.services import training_validity
from app.services.special_training_profiles import (
    resolve_special_duration_hours,
    resolve_special_profile_key,
    special_meta_for_api,
)
from app.schemas.training import resolve_training_hours


router = APIRouter(prefix="/trainings", tags=["Eğitim Yönetimi"])
logger = logging.getLogger(__name__)
PACKAGE_MANAGER_ROLES = (UserRole.GLOBAL_ADMIN, UserRole.COMPANY_ADMIN)
OPERATIONAL_ROLES = (UserRole.GLOBAL_ADMIN, UserRole.COMPANY_ADMIN, UserRole.SAFETY_SPECIALIST)
# Cevap anahtarı gibi eğitici çıktılar için geriye dönük rol kümesi.
# Paket yaşam döngüsü için PACKAGE_MANAGER_ROLES + OSGB kapsamı ayrıca doğrulanır.
EDIT_ROLES = OPERATIONAL_ROLES
OPERATIONAL_UPDATE_FIELDS = frozenset({
    "status",
    "attendance_verified",
    "success_verified",
    "participant_ids",
})
RULES = {"Az Tehlikeli": (8, 3), "Tehlikeli": (12, 2), "Çok Tehlikeli": (16, 1)}


def is_training_package_manager(user: User) -> bool:
    """Klasik eğitim paketini yalnız OSGB merkezi yöneticisi yönetebilir."""
    return user.role == UserRole.GLOBAL_ADMIN or (
        user.role == UserRole.COMPANY_ADMIN
        and bool(user.osgb_id)
        and user.company_id is None
    )


def require_training_package_manager(user: User = Depends(get_current_user)) -> User:
    if not is_training_package_manager(user):
        raise HTTPException(
            status_code=403,
            detail=(
                "Yüz yüze eğitim paketini yalnızca OSGB yöneticisi "
                "veya global yönetici yönetebilir."
            ),
        )
    return user


def assert_training_creation_scope(user: User) -> None:
    """Uzman, erişebildiği işyerinde klasik eğitim kaydı hazırlayabilir.

    ``create_training`` önce ``ensure_access`` ile şirket kapsamını doğrular.
    Bu kontrol yalnızca yetkili rolün tanım alanlarıyla yeni kayıt açmasını
    sağlar; OSGB geneli arşivleme, silme ve logo yönetimi aşağıdaki ayrı
    package-manager bağımlılıklarında korunur.
    """
    if is_training_package_manager(user) or user.role == UserRole.SAFETY_SPECIALIST:
        return
    raise HTTPException(
        status_code=403,
        detail=(
            "Klasik eğitim kaydını yalnızca OSGB yöneticisi, global yönetici "
            "veya görevlendirilmiş İSG uzmanı oluşturabilir."
        ),
    )


def require_training_operator(user: User = Depends(get_current_user)) -> User:
    if user.role not in OPERATIONAL_ROLES:
        raise HTTPException(status_code=403, detail="Bu eğitim işlemi için yetkiniz yok.")
    return user


def assert_training_update_scope(user: User, fields) -> None:
    """Uzman ve işyeri hesaplarının PATCH işlemini operasyonla sınırlar."""
    if is_training_package_manager(user):
        return
    if set(fields) - OPERATIONAL_UPDATE_FIELDS:
        raise HTTPException(
            status_code=403,
            detail=(
                "Eğitim paketinin tanım alanlarını yalnızca OSGB yöneticisi "
                "değiştirebilir; uzmanlar katılımcı ve gerçekleşme durumunu yönetebilir."
            ),
        )
LOGO_EXT = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
LOGO_MIME = {"image/png", "image/jpeg", "image/jpg", "image/webp", "image/gif"}
EXCEL_EXT = (".xlsx", ".xlsm", ".csv")


def ensure_access(db: Session, user: User, company_id: int):
    ensure_company_access(db, user, company_id)


def _ensure_hygiene_instructor_is_authorized(
    db: Session,
    *,
    company_id: int,
    payload: TrainingCreate,
) -> None:
    """Hijyen eğitiminde yalnız izin verilen sağlık personeli unvanlarını kabul et."""
    profile_key = resolve_special_profile_key(payload)
    if profile_key not in {"hijyen_sanitasyon", "gida_su_hijyeni"}:
        return
    qualification = (
        str(payload.instructor_qualification or "")
        .strip()
        .casefold()
        .replace("i\u0307", "i")
    )
    allowed_titles = (
        "işyeri hekimi",
        "hemşire",
        "işyeri hemşiresi",
        "diğer sağlık personeli",
    )
    if not any(title in qualification for title in allowed_titles):
        raise HTTPException(
            422,
            "Hijyen eğitimini yalnız işyeri hekimi, hemşire veya diğer sağlık personeli verebilir.",
        )

def add_years(d: date, years: int) -> date:
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        return d.replace(month=2, day=28, year=d.year + years)


def _load_training(db: Session, training_id: int) -> TrainingSession:
    row = db.scalar(
        select(TrainingSession)
        .options(selectinload(TrainingSession.participants))
        .where(TrainingSession.id == training_id)
    )
    if not row:
        raise HTTPException(404, "Eğitim kaydı bulunamadı.")
    return row


def _employees_map(db: Session, training: TrainingSession) -> dict:
    ids = [p.employee_id for p in training.participants]
    if not ids:
        return {}
    return {e.id: e for e in db.scalars(select(Employee).where(Employee.id.in_(ids))).all()}


def _attendance_workplace_context(db: Session, training: TrainingSession, company: Company | None) -> dict:
    """Resolve display-only workplace identity from existing records.

    The attendance PDF must reflect the selected workplace/branch. Nothing is
    written here: registry, NACE, and assigned physician values are passed to
    the renderer only for this document.
    """
    registry = getattr(company, "sgk_registry_no", None) if company else None
    branch_id = getattr(training, "branch_id", None)
    if branch_id:
        branch = db.get(Branch, branch_id)
        if branch and branch.sgk_registry_no:
            registry = branch.sgk_registry_no

    physician = str(getattr(training, "workplace_physician", None) or "").strip()
    if not physician:
        try:
            defaults = training_defaults(db, training.company_id).get("defaults") or {}
            physician = str(defaults.get("workplace_physician") or "").strip()
        except Exception:
            # A document export should still work for legacy records whose
            # assignment history is incomplete; the persisted field remains
            # the first and authoritative source.
            physician = ""

    return {
        "workplace_sgk_registry_no": str(registry or "").strip() or None,
        "nace_code": getattr(company, "nace_code", None) if company else None,
        "physician_name": physician or None,
    }


def _err_detail(data) -> str:
    detail = data if not isinstance(data, dict) else data.get("detail", data)
    if isinstance(detail, list):
        parts = []
        for item in detail:
            if isinstance(item, dict):
                parts.append(str(item.get("msg") or item))
            else:
                parts.append(str(item))
        return "; ".join(parts) or "İşlem tamamlanamadı."
    return str(detail or "İşlem tamamlanamadı.")


@router.get("/sectors")
def list_sectors():
    """Canlı uyumlu sektör listesi (auth zorunlu değil)."""
    return sectors_list_for_api(include_legacy_nace_aliases=True)


@router.get("/layout-info")
def training_layout_info():
    """Canlı deploy doğrulama — eski API’de bu uç yoktur."""
    return {
        "pdf_layout": "pro-2026",
        "attendance_title": "KATILIMCI İMZA FORMU",
        "form_no": "İSG-EĞT-KF-01",
        "certificate_title": "TEMEL İŞ SAĞLIĞI VE GÜVENLİĞİ EĞİTİMİ KATILIM BELGESİ",
        "certificate_no_format": "ISG-GGAAYYYY-001",
        "endpoints": ["attendance.pdf", "certificates.pdf", "verify/{code}", "special-profiles"],
    }



@router.get("/special-profiles")
def list_special_profiles(user: User = Depends(get_current_user)):
    return special_meta_for_api()

@router.get("/meta")
def training_meta(user: User = Depends(get_current_user)):
    return meta_payload()


@router.get("/verify/{code}", response_model=TrainingVerifyResponse, response_model_exclude_none=True)
def verify_training(code: str, db: Session = Depends(get_db)):
    """Kamuya açık belge doğrulama — bakanlık / işveren kontrolü için."""
    clean = (code or "").strip().upper()
    if not clean or len(clean) < 8:
        return TrainingVerifyResponse(
            valid=False, verification_code=clean or "", message="Geçersiz doğrulama kodu."
        )
    row = db.scalar(
        select(TrainingSession)
        .options(selectinload(TrainingSession.participants))
        .where(TrainingSession.verification_code == clean)
    )
    if not row:
        return TrainingVerifyResponse(
            valid=False, verification_code=clean, message="Bu kodla eşleşen eğitim belgesi bulunamadı."
        )
    company = db.get(Company, row.company_id)
    emp_map = _employees_map(db, row)
    participants = []
    for p in row.participants:
        e = emp_map.get(p.employee_id)
        participants.append(
            {
                "full_name": e.full_name if e else f"#{p.employee_id}",
                "certificate_number": p.certificate_number,
            }
        )
    return TrainingVerifyResponse(
        valid=True,
        verification_code=clean,
        title=row.title,
        company_name=company.name if company else None,
        start_date=row.start_date,
        end_date=row.end_date,
        hazard_class=row.hazard_class,
        duration_hours=row.duration_hours,
        instructor_name=row.instructor_name,
        workplace_physician=row.workplace_physician,
        employer_representative=row.employer_representative,
        participant_count=len(participants),
        participants=participants,
        message="Belge doğrulandı.",
    )


def _is_basic_training(row: TrainingSession) -> bool:
    """Özel eğitim profiline (yüksekte çalışma, hijyen vb.) uymayan = temel İSG."""
    return not resolve_special_duration_hours(
        SimpleNamespace(
            training_type=row.training_type or "",
            title=row.title or "",
            notes=row.notes or "",
        )
    )


@router.get("/assigned-team")
def training_assigned_team(
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """İşyerine görevlendirilmiş uzman/hekim — form alanlarını kendiliğinden doldurur.

    Eğitici ve hekim adının elle yazılması yazım farkı üretiyordu; ad artık
    görevlendirme kaydından gelir.
    """
    effective = effective_company_id(db, user, company_id)
    ensure_access(db, user, effective)
    return training_defaults(db, effective)


@router.get("/employee-status")
def employee_training_status(
    company_id: int | None = None,
    status: str | None = Query(default=None, pattern="^(never|expired|due_soon|ok)$"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Çalışan bazlı temel eğitim geçerliliği — kimin eğitimi dolmuş?

    Kayıt bazlı `next_training_date` çalışana indirgenir; hiç eğitim almamış
    personel de listelenir (yönetmelik: işe başlamadan önce eğitim).
    """
    effective = effective_company_id(db, user, company_id)
    ensure_access(db, user, effective)

    employees = list(
        db.scalars(
            select(Employee)
            .where(Employee.company_id == effective, Employee.is_active.is_(True))
            .order_by(Employee.full_name)
        ).all()
    )
    if not employees:
        return {
            "company_id": effective,
            "summary": training_validity.summarize([]),
            "rows": [],
        }

    emp_ids = [e.id for e in employees]
    pairs = db.execute(
        select(TrainingParticipant.employee_id, TrainingSession)
        .join(TrainingSession, TrainingSession.id == TrainingParticipant.training_id)
        .where(
            TrainingSession.company_id == effective,
            TrainingSession.status != TrainingStatus.CANCELLED,
            TrainingParticipant.employee_id.in_(emp_ids),
        )
    ).all()

    latest: dict[int, dict] = {}
    for employee_id, session in pairs:
        if not _is_basic_training(session):
            continue
        finished = session.end_date or session.start_date
        current = latest.get(employee_id)
        if current is None or (finished and finished > current["end"]):
            latest[employee_id] = {
                "end": finished,
                "due": session.next_training_date,
                "title": session.title,
                "training_id": session.id,
            }

    rows: list[dict] = []
    for employee in employees:
        last = latest.get(employee.id)
        state = training_validity.evaluate_employee(
            hire_date=employee.start_date,
            last_training_end=last["end"] if last else None,
            next_due=last["due"] if last else None,
        )
        rows.append(
            {
                "employee_id": employee.id,
                "full_name": employee.full_name,
                "department": employee.department,
                "job_title": employee.job_title,
                "hire_date": employee.start_date.isoformat() if employee.start_date else None,
                "last_training_title": last["title"] if last else None,
                "last_training_id": last["training_id"] if last else None,
                **state,
            }
        )

    summary = training_validity.summarize(rows)
    rows.sort(key=training_validity.sort_key)
    if status:
        rows = [r for r in rows if r["status"] == status]
    return {"company_id": effective, "summary": summary, "rows": rows}


@router.post("/parse-excel")
async def parse_excel(
    company_id: int = Query(...),
    create_missing: bool = Query(False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_training_operator),
):
    """Excel/CSV çalışan listesini okur; isteğe bağlı eksik personeli oluşturur (Pro parity)."""
    ensure_access(db, user, company_id)
    company = db.get(Company, company_id)
    if not company:
        raise HTTPException(404, "Firma bulunamadı.")
    name = (file.filename or "").lower()
    if not name.endswith(EXCEL_EXT):
        raise HTTPException(
            422,
            "Geçersiz dosya! Lütfen .xlsx, .xlsm veya .csv uzantılı bir çalışan listesi yükleyin.",
        )
    if name.endswith(".xls") and not name.endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            422,
            "Eski .xls formatı desteklenmez. Lütfen dosyayı Excel'de .xlsx olarak kaydedip tekrar yükleyin.",
        )
    content = await file.read()
    suffix = Path(name).suffix.lower() or ".xlsx"
    if suffix != ".csv":
        assert_safe_upload(content, suffix, file.filename or "")
    try:
        rows, excel_meta, logo_bytes = parse_employee_upload(content, file.filename or "liste.xlsx")
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    if not rows:
        raise HTTPException(422, "Excel dosyasında katılımcı bulunamadı. Ad Soyad sütunu gerekli.")

    result, created = resolve_or_create_employees(db, company_id, rows, create_missing=create_missing)
    if create_missing:
        db.commit()
    ids = [r["employee_id"] for r in result if r["employee_id"]]
    if create_missing and not ids:
        raise HTTPException(
            422,
            "Excel okundu ama personel oluşturulamadı. Ad Soyad sütununu kontrol edin "
            "(Ad Soyad / Adı Soyadı).",
        )
    return {
        "count": len(result),
        "created": created,
        "matched": sum(1 for r in result if r["matched"]),
        "participants": result,
        "participant_ids": ids,
        "excel_meta": excel_meta or {},
        "has_embedded_logo": bool(logo_bytes),
    }


@router.get("", response_model=list[TrainingResponse])
def list_trainings(
    q: str | None = Query(None, max_length=100),
    company_id: int | None = None,
    include_archived: bool = Query(False),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    query = (
        select(TrainingSession)
        .options(selectinload(TrainingSession.participants))
        .order_by(TrainingSession.start_date.desc())
    )
    if not include_archived:
        query = query.where(TrainingSession.archived_at.is_(None))
    if user.role == UserRole.GLOBAL_ADMIN:
        if company_id:
            query = query.where(TrainingSession.company_id == company_id)
    else:
        allowed = accessible_company_ids_or_empty(db, user)
        if not allowed:
            return []
        if company_id:
            ensure_access(db, user, company_id)
            query = query.where(TrainingSession.company_id == company_id)
        else:
            query = query.where(TrainingSession.company_id.in_(allowed))
    if q:
        p = f"%{q.strip()}%"
        query = query.where(
            or_(
                TrainingSession.title.ilike(p),
                TrainingSession.instructor_name.ilike(p),
                TrainingSession.sector.ilike(p),
            )
        )
    return list(db.scalars(query).unique().all())


@router.get("/export.xlsx")
def export_trainings_xlsx(
    q: str | None = Query(None, max_length=100),
    company_id: int | None = None,
    include_archived: bool = Query(False),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Eğitim kayıt listesi Excel."""
    query = (
        select(TrainingSession)
        .options(selectinload(TrainingSession.participants))
        .order_by(TrainingSession.start_date.desc())
    )
    rows: list = []
    if not include_archived:
        query = query.where(TrainingSession.archived_at.is_(None))
    if user.role == UserRole.GLOBAL_ADMIN:
        if company_id:
            query = query.where(TrainingSession.company_id == company_id)
    else:
        allowed = accessible_company_ids_or_empty(db, user)
        if not allowed:
            query = None
        elif company_id:
            ensure_access(db, user, company_id)
            query = query.where(TrainingSession.company_id == company_id)
        else:
            query = query.where(TrainingSession.company_id.in_(allowed))
    if query is not None:
        if q:
            p = f"%{q.strip()}%"
            query = query.where(
                or_(
                    TrainingSession.title.ilike(p),
                    TrainingSession.instructor_name.ilike(p),
                    TrainingSession.sector.ilike(p),
                )
            )
        rows = list(db.scalars(query.limit(2000)).unique().all())
    companies = {
        c.id: c.name
        for c in db.scalars(
            select(Company).where(Company.id.in_({r.company_id for r in rows} or {-1}))
        ).all()
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Eğitimler"
    headers = [
        "Firma",
        "Başlık",
        "Tür",
        "Yöntem",
        "Başlangıç",
        "Bitiş",
        "Süre (saat)",
        "Tehlike Sınıfı",
        "Sektör",
        "Eğitmen",
        "Katılımcı",
        "Durum",
        "Doğrulama Kodu",
    ]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="0D6EFD")
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
    for r in rows:
        st = r.status.value if hasattr(r.status, "value") else str(r.status)
        ws.append(
            [
                companies.get(r.company_id, str(r.company_id)),
                r.title,
                r.training_type,
                r.delivery_method,
                r.start_date.isoformat() if r.start_date else "",
                r.end_date.isoformat() if r.end_date else "",
                r.duration_hours,
                r.hazard_class,
                r.sector or "",
                r.instructor_name,
                len(r.participants or []),
                st,
                r.verification_code or "",
            ]
        )
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    stamp = datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="egitim-listesi-{stamp}.xlsx"'},
    )


@router.post("", response_model=TrainingResponse)
def create_training(
    payload: TrainingCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_training_operator),
):
    ensure_access(db, user, payload.company_id)
    assert_training_creation_scope(user)
    company = db.get(Company, payload.company_id)
    if not company:
        raise HTTPException(404, "Firma bulunamadı.")
    _ensure_hygiene_instructor_is_authorized(
        db,
        company_id=payload.company_id,
        payload=payload,
    )
    kod = sektor_kodu_cozumle(payload.sector)
    if not kod:
        raise HTTPException(422, "Sektör / iş kolu resmî NACE listesinden seçilmelidir.")
    resmi_tehlike_sinifi = sektor_tehlike_sinifi(kod)
    etkin_tehlike_sinifi = resmi_tehlike_sinifi or payload.hazard_class
    if etkin_tehlike_sinifi not in RULES:
        raise HTTPException(422, "Seçilen NACE faaliyeti için geçerli tehlike sınıfı bulunamadı.")
    if payload.participant_ids:
        employees = list(
            db.scalars(
                select(Employee).where(
                    Employee.id.in_(payload.participant_ids),
                    Employee.company_id == payload.company_id,
                    Employee.is_active.is_(True),
                )
            ).all()
        )
        if len(employees) != len(set(payload.participant_ids)):
            raise HTTPException(422, "Katılımcılardan biri firmaya ait değil veya pasif.")
    _, years = RULES[etkin_tehlike_sinifi]
    hours = resolve_training_hours(
        training_type=payload.training_type,
        title=payload.title,
        notes=payload.notes,
        hazard_class=etkin_tehlike_sinifi,
    )
    calendar_days = (payload.end_date - payload.start_date).days + 1
    min_days = max(1, (hours + 7) // 8)
    if calendar_days < min_days:
        raise HTTPException(
            422,
            f"{hours} saatlik eğitim en az {min_days} güne yayılmalıdır "
            "(günde en fazla 8 ders saati).",
        )
    # Deterministik hash çakışmasın diye uuid ekle (aynı başlık/tarih tekrarında 500 önleme)
    code = None
    for _ in range(12):
        raw = (
            f"{payload.company_id}|{payload.title}|{payload.start_date.isoformat()}|"
            f"{payload.end_date.isoformat()}|{user.id}|{uuid4().hex}"
        )
        candidate = hashlib.sha256(raw.encode()).hexdigest()[:16].upper()
        exists = db.scalar(
            select(TrainingSession.id).where(TrainingSession.verification_code == candidate)
        )
        if not exists:
            code = candidate
            break
    if not code:
        raise HTTPException(500, "Doğrulama kodu üretilemedi; tekrar deneyin.")
    values = payload.model_dump(exclude={"participant_ids"})
    values["sector"] = kod
    values["hazard_class"] = etkin_tehlike_sinifi
    if not (values.get("stamp_text") or "").strip():
        values["stamp_text"] = (
            "6331 sayılı İş Sağlığı ve Güvenliği Kanunu ve "
            "Çalışanların İş Sağlığı ve Güvenliği Eğitimlerinin Usul ve Esasları Hakkında Yönetmelik "
            "kapsamında düzenlenmiştir."
        )
    row = TrainingSession(
        **values,
        duration_hours=hours,
        renewal_years=years,
        next_training_date=add_years(payload.end_date, years),
        verification_code=code,
        created_by_id=user.id,
    )
    db.add(row)
    db.flush()
    for eid in sorted(set(payload.participant_ids)):
        db.add(
            TrainingParticipant(
                training_id=row.id,
                employee_id=eid,
                certificate_number=f"EGT-{row.id:06d}-{eid:06d}",
            )
        )
    db.commit()
    return db.scalar(
        select(TrainingSession)
        .options(selectinload(TrainingSession.participants))
        .where(TrainingSession.id == row.id)
    )


@router.post("/{training_id}/upload-participants")
async def upload_participants(
    training_id: int,
    file: UploadFile = File(...),
    create_missing: bool = Query(True),
    db: Session = Depends(get_db),
    user: User = Depends(require_training_operator),
):
    """Canlı API uyumu: eğitim kaydına Excel ile katılımcı ekler."""
    row = _load_training(db, training_id)
    ensure_access(db, user, row.company_id)
    if row.archived_at:
        raise HTTPException(409, "Arşivlenmiş eğitim kaydı değiştirilemez.")
    name = (file.filename or "").lower()
    if not name.endswith(EXCEL_EXT):
        raise HTTPException(
            422,
            "Geçersiz dosya! Lütfen .xlsx, .xlsm veya .csv uzantılı bir çalışan listesi yükleyin.",
        )
    content = await file.read()
    suffix = Path(name).suffix.lower() or ".xlsx"
    if suffix != ".csv":
        assert_safe_upload(content, suffix, file.filename or "")
    try:
        parsed, _meta, _logo = parse_employee_upload(content, file.filename or "liste.xlsx")
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    if not parsed:
        raise HTTPException(422, "Excel dosyasında katılımcı bulunamadı.")

    result, created = resolve_or_create_employees(
        db, row.company_id, parsed, create_missing=create_missing
    )
    existing_ids = {p.employee_id for p in row.participants}
    added = 0
    for item in result:
        emp_id = item.get("employee_id")
        if not emp_id or emp_id in existing_ids:
            continue
        db.add(
            TrainingParticipant(
                training_id=row.id,
                employee_id=emp_id,
                certificate_number=f"EGT-{row.id:06d}-{emp_id:06d}",
            )
        )
        existing_ids.add(emp_id)
        added += 1
    db.commit()
    refreshed = _load_training(db, training_id)
    return {
        "added": added,
        "created_employees": created,
        "participant_count": len(refreshed.participants),
        "training": TrainingResponse.model_validate(refreshed),
    }


@router.post("/{training_id}/logo", response_model=TrainingResponse)
async def upload_training_logo(
    training_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_training_package_manager),
):
    """Firma / eğitim logosu — PDF başlığına basılır."""
    row = _load_training(db, training_id)
    ensure_access(db, user, row.company_id)
    if row.archived_at:
        raise HTTPException(409, "Arşivlenmiş eğitim kaydı değiştirilemez.")
    original = Path(file.filename or "logo.png")
    ext = original.suffix.lower()
    if ext not in LOGO_EXT or (file.content_type and file.content_type not in LOGO_MIME):
        raise HTTPException(400, "Logo için PNG veya JPG yükleyin.")
    content = await file.read(2 * 1024 * 1024 + 1)
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(413, "Logo en fazla 2 MB olabilir.")
    stored = f"{training_id}_{uuid4().hex[:10]}{ext}"
    rel = f"{row.company_id}/training-logos/{stored}"
    if settings.upload_gateway_enabled:
        persist_relative(content, relative_path=rel, original_name=original.name, max_bytes=2 * 1024 * 1024)
    else:
        assert_safe_upload(content, ext, original.name)
        company_dir = safe_upload_root() / str(row.company_id) / "training-logos"
        company_dir.mkdir(parents=True, exist_ok=True)
        target = (company_dir / stored).resolve()
        if safe_upload_root() not in target.parents:
            raise HTTPException(400, "Geçersiz dosya yolu.")
        target.write_bytes(content)
    row.logo_path = rel.replace("\\", "/")
    db.commit()
    return _load_training(db, training_id)


@router.patch("/{training_id}", response_model=TrainingResponse)
def update_training(
    training_id: int,
    payload: TrainingUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_training_operator),
):
    row = _load_training(db, training_id)
    ensure_access(db, user, row.company_id)
    if row.archived_at:
        raise HTTPException(409, "Arşivlenmiş eğitim kaydı değiştirilemez.")
    values = payload.model_dump(exclude_unset=True)
    assert_training_update_scope(user, values.keys())
    new_ids = values.pop("participant_ids", None)
    for k, v in values.items():
        setattr(row, k, v)
    if new_ids is not None:
        _replace_participants(db, row, new_ids)
    db.commit()
    return _load_training(db, training_id)


@router.post("/{training_id}/archive", response_model=TrainingResponse)
def archive_training(
    training_id: int,
    payload: TrainingArchiveRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_training_package_manager),
):
    """Tamamlanmış eğitim kaydını katılımcıları silmeden arşivler."""
    row = _load_training(db, training_id)
    ensure_access(db, user, row.company_id)
    if row.archived_at:
        return row
    row.archived_at = datetime.utcnow()
    row.archived_by_id = user.id
    row.archive_reason = payload.reason.strip()
    db.commit()
    return _load_training(db, training_id)


@router.delete("/{training_id}")
def delete_training(
    training_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_training_package_manager),
):
    """OSGB yöneticisinin erişebildiği klasik eğitim kaydını kaldırır.

    Katılımcı, sınav ve diğer eğitim alt kayıtları mevcut FK/ORM cascade
    kurallarıyla birlikte silinir. Çalışan, işyeri, atama ve kullanıcı
    kayıtlarına dokunulmaz. Onaylanmış sunumların tarihsel denetim izi de
    yanlışlıkla silinmemesi için bu kayıtlar ayrıca korunur.
    """
    row = _load_training(db, training_id)
    ensure_access(db, user, row.company_id)

    if row.archived_at:
        raise HTTPException(409, "Arşivlenmiş eğitim kaydı silinemez.")
    if row.status != TrainingStatus.PLANNED:
        raise HTTPException(
            409,
            "Tamamlanmış veya iptal edilmiş eğitim kaydı silinemez; "
            "katılımcı ve belge geçmişini korumak için Arşivle işlemini kullanın.",
        )

    approved_audit_id = db.scalar(
        select(TrainingPresentationApproval.id)
        .where(TrainingPresentationApproval.training_id == training_id)
        .limit(1)
    )
    if approved_audit_id is not None:
        raise HTTPException(
            409,
            "Onaylı eğitim sunumu bulunan kayıt silinemez; tarihsel denetim izi korunmalıdır.",
        )

    company_id = row.company_id
    logo_path = row.logo_path
    participant_count = len(row.participants)
    db.delete(row)
    db.commit()

    # Dosya temizliği veritabanı silinmesini başarısız kılmamalı: eski kurulumlar
    # yerel disk, yeni kurulumlar ise upload gateway kullanabilir.
    if logo_path:
        try:
            delete_relative(logo_path)
        except Exception:
            logger.warning(
                "training logo cleanup failed after record deletion: training_id=%s path=%s",
                training_id,
                logo_path,
                exc_info=True,
            )

    return {
        "ok": True,
        "deleted": True,
        "id": training_id,
        "company_id": company_id,
        "participant_count": participant_count,
        "message": "Eğitim kaydı ve bağlı katılımcı kayıtları silindi.",
    }


def _replace_participants(db: Session, row: TrainingSession, employee_ids: list[int]) -> None:
    """Katılımcı listesini verilen kümeye eşitler; kalanların puan/katılım verisi korunur."""
    wanted = {int(i) for i in employee_ids}
    if not wanted:
        raise HTTPException(422, "En az bir katılımcı gerekli; liste boşaltılamaz.")
    valid = set(
        db.scalars(
            select(Employee.id).where(
                Employee.id.in_(wanted),
                Employee.company_id == row.company_id,
                Employee.is_active.is_(True),
            )
        ).all()
    )
    if valid != wanted:
        raise HTTPException(422, "Katılımcılardan biri firmaya ait değil veya pasif.")
    current = {p.employee_id: p for p in row.participants}
    for eid, participant in current.items():
        if eid not in wanted:
            db.delete(participant)
    for eid in sorted(wanted - current.keys()):
        db.add(
            TrainingParticipant(
                training_id=row.id,
                employee_id=eid,
                certificate_number=f"EGT-{row.id:06d}-{eid:06d}",
            )
        )


@router.get("/{training_id}/attendance.pdf")
def attendance_pdf(
    training_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    row = _load_training(db, training_id)
    ensure_access(db, user, row.company_id)
    company = db.get(Company, row.company_id)
    employees = _employees_map(db, row)
    workplace_context = _attendance_workplace_context(db, row, company)
    try:
        pdf_bytes = build_attendance_pdf(
            company_name=company.name if company else str(row.company_id),
            training=row,
            employees=employees,
            **workplace_context,
        )
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(500, str(exc)) from exc
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="egitim-{training_id}-katilimci-imza-formu-PRO-v2.pdf"'},
    )


@router.get("/{training_id}/certificates.pdf")
def certificates_pdf(
    training_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    row = _load_training(db, training_id)
    ensure_access(db, user, row.company_id)
    if not row.participants:
        raise HTTPException(422, "Katılım belgesi için en az bir katılımcı gerekli.")
    if resolve_special_profile_key(row):
        if not row.attendance_verified:
            raise HTTPException(422, "Özel eğitim katılımı doğrulanmadan belge üretilemez. Katılım ve imza listesini doğrulayın.")
        if not row.success_verified:
            raise HTTPException(422, "Özel eğitim değerlendirmesi doğrulanmadan belge üretilemez.")
    company = db.get(Company, row.company_id)
    employees = _employees_map(db, row)
    try:
        pdf_bytes = build_certificates_pdf(
            company_name=company.name if company else str(row.company_id),
            training=row,
            employees=employees,
        )
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(500, str(exc)) from exc
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="egitim-{training_id}-katilim-belgeleri-v2.pdf"'},
    )



@router.get("/{training_id}/exam-answer-key.pdf")
def training_exam_answer_key_pdf(
    training_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Cevap anahtarını çalışan sınavından ayrı ve yalnız yetkili eğiticiye üret."""
    if user.role not in EDIT_ROLES:
        raise HTTPException(403, "Cevap anahtarını yalnız yetkili eğitici veya eğitim yöneticisi indirebilir.")
    row = _load_training(db, training_id)
    ensure_access(db, user, row.company_id)
    company = db.get(Company, row.company_id)
    try:
        pdf_bytes = build_exam_pdf(
            company_name=company.name if company else str(row.company_id),
            training=row,
            db=db,
            created_by_id=user.id,
            include_answer_key=True,
            answer_key_only=True,
        )
        db.commit()
    except ValueError as exc:
        db.rollback()
        raise HTTPException(422, str(exc)) from exc
    except RuntimeError as exc:
        db.rollback()
        raise HTTPException(500, str(exc)) from exc
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="egitim-{training_id}-egitici-cevap-anahtari.pdf"'},
    )


@router.get("/{training_id}/exam.pdf")
def training_exam_pdf(
    training_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """5 sabit temel + mevcut 5 ortak + 5 teknik + 5 sektör soruluk sınav üretir."""
    row = _load_training(db, training_id)
    ensure_access(db, user, row.company_id)
    company = db.get(Company, row.company_id)
    try:
        pdf_bytes = build_exam_pdf(
            company_name=company.name if company else str(row.company_id),
            training=row,
            db=db,
            created_by_id=user.id,
        )
        db.commit()
    except ValueError as exc:
        db.rollback()
        raise HTTPException(422, str(exc)) from exc
    except RuntimeError as exc:
        db.rollback()
        raise HTTPException(500, str(exc)) from exc
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="egitim-{training_id}-isg-sinavi.pdf"'},
    )
