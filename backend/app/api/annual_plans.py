"""Yıllık plan API — İSG PRO 2026 Planlama Merkezi parity."""
from __future__ import annotations

import logging
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import PlainTextResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.company_access import company_ids_for_query, effective_company_id, ensure_company_access
from app.api.deps import get_current_user, require_roles
from app.core.database import get_db
from app.models.entities import AnnualPlanItem, AnnualPlanStatus, Company, User, UserRole
from app.schemas.annual_plan import (
    AnnualPlanCreate,
    AnnualPlanGenerate,
    AnnualPlanResponse,
    AnnualPlanUpdate,
)
from app.services.annual_plan_pdf import build_annual_plan_pdf
from app.services.annual_plan_template import template_for_hazard
from app.services.assigned_team import team_names
from app.services.tr_calendar import is_non_working_day, plan_target_date

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/annual-plans", tags=["Yıllık Planlar"])

EDIT_ROLES = (
    UserRole.GLOBAL_ADMIN,
    UserRole.COMPANY_ADMIN,
    UserRole.SAFETY_SPECIALIST,
    UserRole.WORKPLACE_PHYSICIAN,
    UserRole.OTHER_HEALTH_PERSONNEL,
)

CATEGORIES = {
    "yillik_calisma": "Yıllık Çalışma Planı",
    "egitim": "Eğitim",
    "saglik": "Sağlık",
    "periyodik": "Periyodik Kontrol",
    "tatbikat": "Tatbikat / Acil Durum",
    "kkd": "KKD",
    "diger": "Diğer",
}


def ensure_access(db: Session, user: User, company_id: int) -> None:
    ensure_company_access(db, user, company_id)


def _active_stmt():
    return select(AnnualPlanItem).where(AnnualPlanItem.deleted_at.is_(None))


def _refresh_delayed(db: Session, items: list[AnnualPlanItem]) -> None:
    today = date.today()
    changed = False
    try:
        for item in items:
            if item.status in (AnnualPlanStatus.COMPLETED, AnnualPlanStatus.CANCELLED):
                continue
            if item.target_date and item.target_date < today and item.status != AnnualPlanStatus.DELAYED:
                item.status = AnnualPlanStatus.DELAYED
                changed = True
        if changed:
            db.commit()
            for item in items:
                db.refresh(item)
    except Exception:
        logger.warning("annual plan delayed-status refresh failed", exc_info=True)
        try:
            db.rollback()
        except Exception:
            logger.warning("annual plan delayed-status rollback failed", exc_info=True)


@router.get("/meta")
def annual_plan_meta():
    return {
        "categories": [{"code": k, "label": v} for k, v in CATEGORIES.items()],
        "statuses": [
            {"code": s.value, "label": {
                "planned": "Planlandı",
                "in_progress": "Devam Ediyor",
                "completed": "Tamamlandı",
                "delayed": "Gecikti",
                "cancelled": "İptal",
            }.get(s.value, s.value)}
            for s in AnnualPlanStatus
        ],
    }


@router.get("/summary")
def annual_plan_summary(
    year: int | None = None,
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    y = year or date.today().year
    effective = effective_company_id(db, user, company_id)
    items = list(
        db.scalars(
            _active_stmt().where(
                AnnualPlanItem.company_id == effective,
                AnnualPlanItem.year == y,
            )
        ).all()
    )
    _refresh_delayed(db, items)
    by_month = {m: 0 for m in range(1, 13)}
    for it in items:
        by_month[it.month] = by_month.get(it.month, 0) + 1
    return {
        "year": y,
        "company_id": effective,
        "total": len(items),
        "completed": sum(1 for i in items if i.status == AnnualPlanStatus.COMPLETED),
        "waiting": sum(
            1
            for i in items
            if i.status in (AnnualPlanStatus.PLANNED, AnnualPlanStatus.IN_PROGRESS)
        ),
        "delayed": sum(1 for i in items if i.status == AnnualPlanStatus.DELAYED),
        "cancelled": sum(1 for i in items if i.status == AnnualPlanStatus.CANCELLED),
        "by_month": by_month,
    }


@router.get("", response_model=list[AnnualPlanResponse])
def list_plan_items(
    year: int | None = None,
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    query = _active_stmt().order_by(AnnualPlanItem.year.desc(), AnnualPlanItem.month, AnnualPlanItem.id)
    company_ids = company_ids_for_query(db, user, company_id)
    if company_ids == []:
        return []
    if company_ids is not None:
        query = query.where(AnnualPlanItem.company_id.in_(company_ids))
    if year:
        query = query.where(AnnualPlanItem.year == year)
    items = list(db.scalars(query).all())
    _refresh_delayed(db, items)
    return items


@router.post("", response_model=AnnualPlanResponse)
def create_plan_item(
    payload: AnnualPlanCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT_ROLES)),
):
    ensure_access(db, user, payload.company_id)
    if not db.get(Company, payload.company_id):
        raise HTTPException(404, "Firma bulunamadı.")
    data = payload.model_dump()
    if not data.get("target_date"):
        data["target_date"] = plan_target_date(payload.year, payload.month, 15)
    item = AnnualPlanItem(**data, created_by_id=user.id)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.post("/generate")
def generate_template(
    payload: AnnualPlanGenerate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT_ROLES)),
):
    """Aynı yıl için yoksa mevzuat dayanaklı şablon maddelerini ekler (tehlike sınıfına göre)."""
    try:
        ensure_access(db, user, payload.company_id)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(403, f"Firma erişimi doğrulanamadı: {exc}") from exc
    company = db.get(Company, payload.company_id)
    if not company:
        raise HTTPException(404, "Firma bulunamadı.")
    if payload.year < 2020 or payload.year > 2100:
        raise HTTPException(422, "Geçersiz yıl.")
    template = template_for_hazard(company.hazard_class)
    existing = list(
        db.scalars(
            _active_stmt().where(
                AnnualPlanItem.company_id == payload.company_id,
                AnnualPlanItem.year == payload.year,
            )
        ).all()
    )
    existing_keys = {(i.month, (i.activity or "").strip().casefold()) for i in existing}
    created = 0
    targets: list[str] = []
    try:
        for month, category, activity, description, responsible, notes, legal in template:
            key = (month, activity.strip().casefold())
            if key in existing_keys:
                continue
            target = plan_target_date(payload.year, month, 15)
            # Son güvenlik: tatil/hafta sonuna asla yazma
            if is_non_working_day(target):
                from app.services.tr_calendar import next_workday
                import calendar as cal

                target = next_workday(target)
                if target.month != month:
                    last = cal.monthrange(payload.year, month)[1]
                    target = plan_target_date(payload.year, month, last)
            note_extra = notes or ""
            preferred = date(payload.year, month, min(15, cal_last_day(payload.year, month)))
            if target != preferred:
                shift_note = (
                    f"Hedef {target.strftime('%d.%m.%Y')} "
                    "(hafta sonu/resmi tatil nedeniyle iş gününe kaydırıldı)."
                )
                note_extra = f"{note_extra} {shift_note}".strip() if note_extra else shift_note
            if is_non_working_day(target):
                raise HTTPException(
                    500,
                    f"Hedef tarih tatil/hafta sonuna düştü ({target}) — plan üretimi iptal.",
                )
            db.add(
                AnnualPlanItem(
                    company_id=payload.company_id,
                    year=payload.year,
                    month=month,
                    category=category,
                    activity=activity,
                    description=description,
                    responsible_name=responsible,
                    target_date=target,
                    status=AnnualPlanStatus.PLANNED,
                    notes=(note_extra[:1500] if note_extra else None),
                    legal_basis=(legal[:240] if legal else None),
                    created_by_id=user.id,
                )
            )
            targets.append(target.isoformat())
            created += 1
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            500,
            f"Otomatik plan üretilemedi: {exc}. "
            "Firma görevlendirmenizi ve API sürümünü kontrol edin.",
        ) from exc
    return {
        "company_id": payload.company_id,
        "year": payload.year,
        "hazard_class": company.hazard_class,
        "created": created,
        "skipped_existing": len(template) - created,
        "template_size": len(template),
        "workday_adjusted": True,
        "holiday_safe": True,
        "target_dates": targets,
        "message": (
            f"{created} madde eklendi (tehlike sınıfı: {company.hazard_class or '—'}; "
            "tatil/hafta sonu hedefleri iş gününe kaydırıldı)."
            if created
            else "Tüm şablon maddeleri zaten mevcut — yeni kayıt eklenmedi."
        ),
    }


def cal_last_day(year: int, month: int) -> int:
    import calendar as cal

    return cal.monthrange(year, month)[1]


@router.get("/export.xlsx")
def export_plan_xlsx(
    year: int | None = None,
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Yıllık çalışma planı — Excel (.xlsx)."""
    y = year or date.today().year
    effective = effective_company_id(db, user, company_id)
    company = db.get(Company, effective)
    items = list(
        db.scalars(
            _active_stmt()
            .where(AnnualPlanItem.company_id == effective, AnnualPlanItem.year == y)
            .order_by(AnnualPlanItem.month, AnnualPlanItem.id)
        ).all()
    )
    status_labels = {
        "planned": "Planlandı",
        "in_progress": "Devam Ediyor",
        "completed": "Tamamlandı",
        "delayed": "Gecikti",
        "cancelled": "İptal",
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Yıllık Plan"
    ws.append(
        [
            "Firma",
            "Yıl",
            "Ay",
            "Ay Adı",
            "Kategori",
            "Faaliyet",
            "Açıklama",
            "Mevzuat Dayanağı",
            "Sorumlu",
            "Hedef Tarih",
            "Durum",
            "Tamamlanma",
            "Notlar",
        ]
    )
    month_names = [
        "",
        "Ocak",
        "Şubat",
        "Mart",
        "Nisan",
        "Mayıs",
        "Haziran",
        "Temmuz",
        "Ağustos",
        "Eylül",
        "Ekim",
        "Kasım",
        "Aralık",
    ]
    firm = company.name if company else str(effective)
    for it in items:
        st = it.status.value if hasattr(it.status, "value") else str(it.status or "")
        ws.append(
            [
                firm,
                it.year,
                it.month,
                month_names[it.month] if 1 <= int(it.month or 0) <= 12 else it.month,
                CATEGORIES.get(it.category or "", it.category or ""),
                it.activity or "",
                it.description or "",
                it.legal_basis or "",
                it.responsible_name or "",
                it.target_date.isoformat() if it.target_date else "",
                status_labels.get(st, st),
                it.completion_date.isoformat() if it.completion_date else "",
                it.notes or "",
            ]
        )
    widths = [28, 8, 6, 12, 18, 40, 36, 36, 22, 14, 14, 14, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    fname = f"yillik-plan-{y}-{effective}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/export.pdf")
def export_plan_pdf(
    year: int | None = None,
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Yıllık çalışma planı — imza kutulu PDF."""
    y = year or date.today().year
    effective = effective_company_id(db, user, company_id)
    company = db.get(Company, effective)
    if not company:
        raise HTTPException(404, "Firma bulunamadı.")
    items = list(
        db.scalars(
            _active_stmt()
            .where(AnnualPlanItem.company_id == effective, AnnualPlanItem.year == y)
            .order_by(AnnualPlanItem.month, AnnualPlanItem.id)
        ).all()
    )
    if not items:
        raise HTTPException(404, "Bu yıl için plan maddesi yok. Önce otomatik plan üretin.")
    names = team_names(db, effective)
    pdf = build_annual_plan_pdf(
        company_name=company.name,
        year=y,
        items=items,
        hazard_class=company.hazard_class,
        specialist_name=names.get("safety_specialist"),
        physician_name=names.get("workplace_physician"),
        employer_name=company.authorized_person,
    )
    fname = f"yillik-plan-{y}-{effective}.pdf"
    return StreamingResponse(
        BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/export.txt")
def export_plan_txt(
    year: int | None = None,
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    y = year or date.today().year
    effective = effective_company_id(db, user, company_id)
    company = db.get(Company, effective)
    items = list(
        db.scalars(
            _active_stmt()
            .where(AnnualPlanItem.company_id == effective, AnnualPlanItem.year == y)
            .order_by(AnnualPlanItem.month, AnnualPlanItem.id)
        ).all()
    )
    lines = [
        "İSG Suite OSGB — Yıllık Çalışma Planı",
        f"Firma: {company.name if company else effective}",
        f"Yıl: {y}",
        f"Olusturma: {datetime.utcnow().strftime('%d.%m.%Y %H:%M')}",
        "-" * 72,
    ]
    for it in items:
        cat = CATEGORIES.get(it.category or "", it.category or "—")
        lines.append(
            f"{it.month:02d}. ay | {cat} | {it.activity} | "
            f"Sorumlu: {it.responsible_name or '—'} | "
            f"Hedef: {it.target_date or '—'} | Durum: {it.status.value}"
        )
        if it.legal_basis:
            lines.append(f"   Mevzuat: {it.legal_basis}")
        if it.description:
            lines.append(f"   Aciklama: {it.description}")
    body = "\n".join(lines) + "\n"
    return PlainTextResponse(
        body,
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="yillik-plan-{y}.txt"'},
    )


@router.patch("/{item_id}", response_model=AnnualPlanResponse)
def update_plan_item(
    item_id: int,
    payload: AnnualPlanUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT_ROLES)),
):
    item = db.get(AnnualPlanItem, item_id)
    if not item or item.deleted_at:
        raise HTTPException(404, "Plan maddesi bulunamadı.")
    ensure_access(db, user, item.company_id)
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/{item_id}")
def delete_plan_item(
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT_ROLES)),
):
    item = db.get(AnnualPlanItem, item_id)
    if not item or item.deleted_at:
        raise HTTPException(404, "Plan maddesi bulunamadı.")
    ensure_access(db, user, item.company_id)
    item.deleted_at = datetime.utcnow()
    db.commit()
    return {"ok": True, "id": item_id}
