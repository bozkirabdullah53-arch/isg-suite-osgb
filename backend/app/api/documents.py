import logging
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.api.company_access import company_ids_for_query, ensure_company_access
from app.api.deps import get_current_user, is_workplace_manager_account, require_roles
from app.core.database import get_db
from app.models.entities import Company, DocumentRecord, User, UserRole
from app.schemas.document import DocumentCreate, DocumentResponse

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/documents", tags=["Dokümanlar"])

EDIT_ROLES = (
    UserRole.GLOBAL_ADMIN,
    UserRole.COMPANY_ADMIN,
    UserRole.SAFETY_SPECIALIST,
    UserRole.WORKPLACE_PHYSICIAN,
)


def ensure_access(db: Session, user: User, company_id: int) -> None:
    ensure_company_access(db, user, company_id)


def require_document_editor(
    user: User = Depends(require_roles(*EDIT_ROLES)),
) -> User:
    """OSGB/profesyonel editörleri kabul eder, işyeri hesabını salt-okunur tutar."""
    if is_workplace_manager_account(user):
        raise HTTPException(
            status_code=403,
            detail="Dokümanlar işyeri hesabında sadece görüntülenebilir.",
        )
    return user


@router.get("/export.xlsx")
def export_documents_xlsx(
    company_id: int | None = None,
    q: str | None = Query(default=None, max_length=100),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Doküman kayıt defteri Excel."""
    query = select(DocumentRecord).order_by(DocumentRecord.created_at.desc())
    company_ids = company_ids_for_query(db, user, company_id)
    if company_ids == []:
        rows = []
    else:
        if company_ids is not None:
            query = query.where(DocumentRecord.company_id.in_(company_ids))
        if q:
            pattern = f"%{q.strip()}%"
            query = query.where(
                or_(DocumentRecord.title.ilike(pattern), DocumentRecord.description.ilike(pattern))
            )
        rows = list(db.scalars(query.limit(2000)).all())

    companies = {
        c.id: c.name
        for c in db.scalars(
            select(Company).where(Company.id.in_({r.company_id for r in rows} or {-1}))
        ).all()
    }
    cat_labels = {
        "general": "Genel",
        "risk": "Risk",
        "training": "Eğitim",
        "health": "Sağlık",
        "emergency": "Acil Durum",
        "legal": "Mevzuat",
        "annual_plan": "Yıllık Plan",
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Dokümanlar"
    headers = [
        "Firma",
        "Başlık",
        "Kategori",
        "Dosya Adı",
        "Versiyon",
        "Başlangıç",
        "Geçerlilik Sonu",
        "Durum",
        "Açıklama",
    ]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="0D6EFD")
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
    for r in rows:
        cat = r.category.value if hasattr(r.category, "value") else str(r.category)
        ws.append(
            [
                companies.get(r.company_id, str(r.company_id)),
                r.title,
                cat_labels.get(cat, cat),
                r.file_name or "",
                r.version or "",
                r.valid_from.isoformat() if r.valid_from else "",
                r.valid_until.isoformat() if r.valid_until else "",
                "Aktif" if r.is_active else "Pasif",
                r.description or "",
            ]
        )
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    stamp = datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="dokuman-kayitlari-{stamp}.xlsx"'},
    )


@router.get("", response_model=list[DocumentResponse])
def list_documents(
    company_id: int | None = None,
    q: str | None = Query(default=None, max_length=100),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    query = select(DocumentRecord).order_by(DocumentRecord.created_at.desc())
    company_ids = company_ids_for_query(db, user, company_id)
    if company_ids == []:
        return []
    if company_ids is not None:
        query = query.where(DocumentRecord.company_id.in_(company_ids))
    if q:
        pattern = f"%{q.strip()}%"
        query = query.where(or_(DocumentRecord.title.ilike(pattern), DocumentRecord.description.ilike(pattern)))
    return list(db.scalars(query).all())


@router.post("", response_model=DocumentResponse)
def create_document(
    payload: DocumentCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_document_editor),
):
    ensure_access(db, user, payload.company_id)
    record = DocumentRecord(**payload.model_dump(), created_by_id=user.id)
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.patch("/{document_id}/deactivate", response_model=DocumentResponse)
def deactivate_document(
    document_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_document_editor),
):
    """Dokümanı pasife alır; bağlı dosya merkezi arşive kopyalanır (EİSA erişimi)."""
    from pathlib import Path

    from app.core.config import settings
    from app.services.archive_store import archive_file_before_delete

    record = db.get(DocumentRecord, document_id)
    if not record:
        raise HTTPException(404, "Doküman bulunamadı.")
    ensure_access(db, user, record.company_id)

    marker = "[stored:"
    desc = record.description or ""
    if marker in desc:
        stored_name = desc.split(marker, 1)[1].split("]", 1)[0]
        path = (Path(settings.upload_dir).resolve() / str(record.company_id) / stored_name)
        try:
            archive_file_before_delete(
                db,
                source=path,
                user=user,
                company_id=record.company_id,
                entity_type="document",
                entity_id=str(record.id),
                original_name=record.file_name,
                notes="Doküman pasife alınmadan önce arşivlendi",
            )
        except Exception:
            logger.warning(
                "document: archive-before-deactivate failed id=%s",
                document_id,
                exc_info=True,
            )
    record.is_active = False
    db.commit()
    db.refresh(record)
    return record
