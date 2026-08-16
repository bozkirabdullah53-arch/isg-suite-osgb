"""6331 uyum sicilleri API — periyodik kontrol, acil plan, ortam ölçüm, İSG kurulu, belge onay."""
from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.api.company_access import company_ids_for_query, ensure_company_access
from app.api.deps import get_current_user, require_roles, require_roles_or_workplace_manager
from app.core.config import settings
from app.core.database import get_db
from app.models.entities import (
    DocumentApproval,
    EmergencyPlan,
    EmergencyPlanFloor,
    EmergencyTeam,
    EmergencyTeamAssignment,
    OhsCommitteeMeeting,
    OhsCommitteeMember,
    PeriodicControl,
    User,
    UserRole,
    WorkplaceMeasurement,
)
from app.schemas.compliance import (
    CommitteeMeetingCreate,
    CommitteeMeetingResponse,
    CommitteeMemberCreate,
    CommitteeMemberResponse,
    DocumentApprovalCreate,
    DocumentApprovalResponse,
    EmergencyFloorCreate,
    EmergencyFloorResponse,
    EmergencyFloorUpdate,
    EmergencyPlanCreate,
    EmergencyPlanResponse,
    EmergencyPlanUpdate,
    MeasurementCreate,
    MeasurementResponse,
    MeasurementUpdate,
    PeriodicControlCreate,
    PeriodicControlResponse,
    PeriodicControlUpdate,
)
from app.services.upload_gateway import persist_relative
from app.services.upload_security import assert_safe_upload

EMPTY_SCENE = '{"version":1,"objects":[],"paths":[]}'

EDIT = (UserRole.GLOBAL_ADMIN, UserRole.SAFETY_SPECIALIST)
VIEW = (UserRole.GLOBAL_ADMIN, UserRole.SAFETY_SPECIALIST, UserRole.WORKPLACE_PHYSICIAN)
# Ortam ölçümü ortak işyeri kaydıdır: uzman yazar, hekim (ve mevcut sağlık
# rolü görünümü) yalnızca atanmış işyerlerinin kayıtlarını okur/dışa aktarır.
PHYS_VIEW = VIEW + (UserRole.OTHER_HEALTH_PERSONNEL,)

PERIODIC_CATEGORIES = [
    {"code": "elektrik", "label": "Elektrik / topraklama"},
    {"code": "kaldirma", "label": "Kaldırma / iletme (vinç, forklift)"},
    {"code": "basincli_kap", "label": "Basınçlı kap"},
    {"code": "asansor", "label": "Asansör"},
    {"code": "yangin_tup", "label": "Yangın tüpü"},
    {"code": "yangin_hidrant", "label": "Yangın hidrant / dolap"},
    {"code": "yangin_algilama", "label": "Yangın algılama / sprinkler"},
    {"code": "diger", "label": "Diğer ekipman"},
]

MEASUREMENT_TYPES = [
    {"code": "gurultu", "label": "Gürültü"},
    {"code": "toz", "label": "Toz / partikül"},
    {"code": "kimyasal", "label": "Kimyasal maruziyet"},
    {"code": "aydinlatma", "label": "Aydınlatma"},
    {"code": "termal", "label": "Termal konfor"},
    {"code": "titresim", "label": "Titreşim"},
    {"code": "diger", "label": "Diğer"},
]

COMMITTEE_ROLES = [
    {"code": "isveren_vekili", "label": "İşveren / vekili"},
    {"code": "igu", "label": "İş güvenliği uzmanı"},
    {"code": "hekim", "label": "İşyeri hekimi"},
    {"code": "calisan_temsilcisi", "label": "Çalışan temsilcisi"},
    {"code": "destek", "label": "Destek elemanı"},
    {"code": "diger", "label": "Diğer üye"},
]


def _due_status(d: date | None) -> str:
    if not d:
        return "unset"
    today = date.today()
    if d < today:
        return "overdue"
    if d <= today + timedelta(days=30):
        return "due_soon"
    return "ok"


def _xlsx(rows: list[list], sheet: str, filename: str) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    fill = PatternFill("solid", fgColor="0D6EFD")
    for r_i, row in enumerate(rows, 1):
        ws.append(row)
        if r_i == 1:
            for col, _ in enumerate(row, 1):
                cell = ws.cell(1, col)
                cell.fill = fill
                cell.font = Font(bold=True, color="FFFFFF")
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Periyodik kontrol ─────────────────────────────────────────────
pc_router = APIRouter(prefix="/periodic-controls", tags=["Periyodik Kontrol"])


@pc_router.get("/meta")
def pc_meta(user: User = Depends(get_current_user)):
    return {"categories": PERIODIC_CATEGORIES, "engine": "periodic-control-v1"}


@pc_router.get("", response_model=list[PeriodicControlResponse])
def list_pc(
    company_id: int | None = None,
    q: str | None = Query(None, max_length=100),
    category: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles_or_workplace_manager(*VIEW)),
):
    stmt = select(PeriodicControl).where(PeriodicControl.is_active.is_(True)).order_by(PeriodicControl.id.desc())
    ids = company_ids_for_query(db, user, company_id)
    if ids == []:
        return []
    if ids is not None:
        stmt = stmt.where(PeriodicControl.company_id.in_(ids))
    if category:
        stmt = stmt.where(PeriodicControl.category == category)
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                PeriodicControl.equipment_name.ilike(like),
                PeriodicControl.location.ilike(like),
                PeriodicControl.serial_no.ilike(like),
            )
        )
    out = []
    for r in db.scalars(stmt.limit(2000)).all():
        item = PeriodicControlResponse.model_validate(r)
        item.review_status = _due_status(r.next_due_date)
        out.append(item)
    return out


@pc_router.get("/export.xlsx")
def export_pc(
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles_or_workplace_manager(*VIEW)),
):
    rows = list_pc(company_id=company_id, q=None, category=None, db=db, user=user)
    cat = {c["code"]: c["label"] for c in PERIODIC_CATEGORIES}
    data = [["Kategori", "Ekipman", "Yer", "Seri No", "Son Kontrol", "Sonraki Termin", "Firma", "Rapor", "Sonuç", "Durum", "Not"]]
    for r in rows:
        data.append([
            cat.get(r.category, r.category),
            r.equipment_name,
            r.location or "",
            r.serial_no or "",
            r.last_control_date.isoformat() if r.last_control_date else "",
            r.next_due_date.isoformat() if r.next_due_date else "",
            r.control_firm or "",
            r.report_ref or "",
            r.result or "",
            r.review_status,
            r.notes or "",
        ])
    stamp = datetime.now().strftime("%Y%m%d")
    return _xlsx(data, "Periyodik Kontrol", f"periyodik-kontrol-{stamp}.xlsx")


@pc_router.post("", response_model=PeriodicControlResponse)
def create_pc(
    payload: PeriodicControlCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles_or_workplace_manager(*EDIT)),
):
    ensure_company_access(db, user, payload.company_id)
    if payload.category not in {c["code"] for c in PERIODIC_CATEGORIES}:
        raise HTTPException(422, "Geçersiz kategori.")
    row = PeriodicControl(**payload.model_dump(), created_by_id=user.id)
    db.add(row)
    db.commit()
    db.refresh(row)
    resp = PeriodicControlResponse.model_validate(row)
    resp.review_status = _due_status(row.next_due_date)
    return resp


@pc_router.patch("/{item_id}", response_model=PeriodicControlResponse)
def update_pc(
    item_id: int,
    payload: PeriodicControlUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles_or_workplace_manager(*EDIT)),
):
    row = db.get(PeriodicControl, item_id)
    if not row:
        raise HTTPException(404, "Kayıt bulunamadı.")
    ensure_company_access(db, user, row.company_id)
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(row, k, v)
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    resp = PeriodicControlResponse.model_validate(row)
    resp.review_status = _due_status(row.next_due_date)
    return resp


# ─── Acil durum planı ──────────────────────────────────────────────
ep_router = APIRouter(prefix="/emergency-plans", tags=["Acil Durum Planı"])


def _upload_root() -> Path:
    return Path(settings.upload_dir).resolve()


def _safe_upload_path(rel: str) -> Path:
    root = _upload_root()
    path = (root / rel.replace("\\", "/")).resolve()
    if not str(path).startswith(str(root)):
        raise HTTPException(400, "Geçersiz dosya yolu.")
    return path


def _ep_enrich(db: Session, row: EmergencyPlan) -> EmergencyPlanResponse:
    item = EmergencyPlanResponse.model_validate(row)
    item.review_status = _due_status(row.next_review_date)
    floors = list(
        db.scalars(
            select(EmergencyPlanFloor).where(EmergencyPlanFloor.plan_id == row.id)
        ).all()
    )
    item.floor_count = len(floors)
    item.has_scene = any(
        (f.scene_json and f.scene_json not in ("", EMPTY_SCENE)) or f.background_storage_path
        for f in floors
    ) or bool(row.kroki_storage_path)
    return item


def _get_plan(db: Session, item_id: int, user: User) -> EmergencyPlan:
    row = db.get(EmergencyPlan, item_id)
    if not row or not row.is_active:
        raise HTTPException(404, "Plan bulunamadı.")
    ensure_company_access(db, user, row.company_id)
    return row


def _get_floor(db: Session, plan_id: int, floor_id: int, user: User) -> tuple[EmergencyPlan, EmergencyPlanFloor]:
    plan = _get_plan(db, plan_id, user)
    fl = db.get(EmergencyPlanFloor, floor_id)
    if not fl or fl.plan_id != plan.id:
        raise HTTPException(404, "Kat bulunamadı.")
    return plan, fl


def _parse_scene(raw: str | None) -> dict:
    if not raw:
        return {"version": 1, "objects": [], "paths": []}
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(400, f"Geçersiz scene_json: {exc}") from exc
    if not isinstance(data, dict):
        raise HTTPException(400, "scene_json nesne olmalı.")
    objs = data.get("objects")
    paths = data.get("paths")
    if objs is not None and not isinstance(objs, list):
        raise HTTPException(400, "objects liste olmalı.")
    if paths is not None and not isinstance(paths, list):
        raise HTTPException(400, "paths liste olmalı.")
    if objs is not None and len(objs) > 2000:
        raise HTTPException(400, "En fazla 2000 nesne.")
    data.setdefault("version", 1)
    data.setdefault("objects", objs or [])
    data.setdefault("paths", paths or [])
    return data


@ep_router.get("/meta")
def ep_meta(user: User = Depends(get_current_user)):
    return {
        "engine": "emergency-kroki-v2.2",
        "note": "Kat bazlı kroki + akıllı tahliye asistanı; ekipler ve tatbikat ile birlikte.",
        "symbols": [
            "exit", "door_exit", "stairs", "assembly", "extinguisher", "hose", "alarm",
            "firstaid", "aed", "electric", "youarehere", "route", "wall", "room",
            "door", "measure", "text", "north",
        ],
    }


@ep_router.get("", response_model=list[EmergencyPlanResponse])
def list_ep(
    company_id: int | None = None,
    q: str | None = Query(None, max_length=100),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*VIEW)),
):
    stmt = select(EmergencyPlan).where(EmergencyPlan.is_active.is_(True)).order_by(EmergencyPlan.id.desc())
    ids = company_ids_for_query(db, user, company_id)
    if ids == []:
        return []
    if ids is not None:
        stmt = stmt.where(EmergencyPlan.company_id.in_(ids))
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(or_(EmergencyPlan.title.ilike(like), EmergencyPlan.assembly_areas.ilike(like)))
    return [_ep_enrich(db, r) for r in db.scalars(stmt.limit(500)).all()]


@ep_router.post("", response_model=EmergencyPlanResponse)
def create_ep(
    payload: EmergencyPlanCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    ensure_company_access(db, user, payload.company_id)
    row = EmergencyPlan(**payload.model_dump(), created_by_id=user.id)
    db.add(row)
    db.flush()
    db.add(
        EmergencyPlanFloor(
            plan_id=row.id,
            company_id=row.company_id,
            name="Zemin",
            sort_order=0,
            scene_json=EMPTY_SCENE,
            width=1600,
            height=1000,
        )
    )
    db.commit()
    db.refresh(row)
    return _ep_enrich(db, row)


@ep_router.get("/export.xlsx")
def export_ep(
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*VIEW)),
):
    rows = list_ep(company_id=company_id, q=None, db=db, user=user)
    data = [["Başlık", "Rev", "Plan Tarihi", "Gözden Geçirme", "Toplanma Alanları", "Kroki", "Kat", "Durum", "Özet"]]
    for r in rows:
        data.append([
            r.title,
            r.revision_no,
            r.plan_date.isoformat() if r.plan_date else "",
            r.next_review_date.isoformat() if r.next_review_date else "",
            r.assembly_areas or "",
            r.kroki_file_name or "",
            r.floor_count,
            r.status,
            (r.scenario_summary or "")[:200],
        ])
    stamp = datetime.now().strftime("%Y%m%d")
    return _xlsx(data, "Acil Plan", f"acil-durum-plani-{stamp}.xlsx")


@ep_router.patch("/{item_id}", response_model=EmergencyPlanResponse)
def update_ep(
    item_id: int,
    payload: EmergencyPlanUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    row = _get_plan(db, item_id, user)
    if row.locked_at:
        raise HTTPException(409, "Plan kilitli; kroki düzenlenemez.")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(row, k, v)
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _ep_enrich(db, row)


@ep_router.delete("/{item_id}")
def delete_ep(
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    """Plan kaydını soft-delete (listeden düşer; kat/kroki dosyaları yerinde kalır)."""
    row = _get_plan(db, item_id, user)
    row.is_active = False
    row.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


@ep_router.post("/{item_id}/lock", response_model=EmergencyPlanResponse)
def lock_ep(
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    row = _get_plan(db, item_id, user)
    row.locked_at = datetime.utcnow()
    row.updated_at = row.locked_at
    db.commit()
    db.refresh(row)
    return _ep_enrich(db, row)


@ep_router.post("/{item_id}/unlock", response_model=EmergencyPlanResponse)
def unlock_ep(
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    row = _get_plan(db, item_id, user)
    row.locked_at = None
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _ep_enrich(db, row)


@ep_router.post("/{item_id}/kroki", response_model=EmergencyPlanResponse)
async def upload_kroki(
    item_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    row = _get_plan(db, item_id, user)
    name = file.filename or "kroki.png"
    data = await file.read()
    if len(data) > 8_000_000:
        raise HTTPException(413, "Kroki dosyası 8 MB sınırını aşıyor.")
    ext = Path(name).suffix.lower() or ".png"
    assert_safe_upload(data, ext, name)
    rel = f"{row.company_id}/emergency-plans/{row.id}_{uuid4().hex[:10]}{ext}"
    if settings.upload_gateway_enabled:
        persist_relative(data, relative_path=rel, original_name=name)
    else:
        target = Path(settings.upload_dir) / rel
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(data)
    row.kroki_file_name = name
    row.kroki_storage_path = rel.replace("\\", "/")
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _ep_enrich(db, row)


@ep_router.get("/{item_id}/floors", response_model=list[EmergencyFloorResponse])
def list_floors(
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*VIEW)),
):
    plan = _get_plan(db, item_id, user)
    rows = list(
        db.scalars(
            select(EmergencyPlanFloor)
            .where(EmergencyPlanFloor.plan_id == plan.id)
            .order_by(EmergencyPlanFloor.sort_order, EmergencyPlanFloor.id)
        ).all()
    )
    return [EmergencyFloorResponse.model_validate(r) for r in rows]


@ep_router.post("/{item_id}/floors", response_model=EmergencyFloorResponse)
def create_floor(
    item_id: int,
    payload: EmergencyFloorCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    plan = _get_plan(db, item_id, user)
    if plan.locked_at:
        raise HTTPException(409, "Plan kilitli.")
    max_ord = db.scalar(
        select(func.max(EmergencyPlanFloor.sort_order)).where(EmergencyPlanFloor.plan_id == plan.id)
    )
    sort_order = payload.sort_order if payload.sort_order is not None else (max_ord or 0) + 1
    fl = EmergencyPlanFloor(
        plan_id=plan.id,
        company_id=plan.company_id,
        name=payload.name.strip() or "Kat",
        sort_order=sort_order,
        scene_json=EMPTY_SCENE,
        width=payload.width,
        height=payload.height,
    )
    db.add(fl)
    plan.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(fl)
    return EmergencyFloorResponse.model_validate(fl)


@ep_router.patch("/{item_id}/floors/{floor_id}", response_model=EmergencyFloorResponse)
def update_floor(
    item_id: int,
    floor_id: int,
    payload: EmergencyFloorUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    plan, fl = _get_floor(db, item_id, floor_id, user)
    if plan.locked_at:
        raise HTTPException(409, "Plan kilitli.")
    data = payload.model_dump(exclude_unset=True)
    if "scene_json" in data and data["scene_json"] is not None:
        parsed = _parse_scene(data["scene_json"])
        data["scene_json"] = json.dumps(parsed, ensure_ascii=False, separators=(",", ":"))
    for k, v in data.items():
        setattr(fl, k, v)
    fl.updated_at = datetime.utcnow()
    plan.updated_at = fl.updated_at
    db.commit()
    db.refresh(fl)
    return EmergencyFloorResponse.model_validate(fl)


@ep_router.delete("/{item_id}/floors/{floor_id}")
def delete_floor(
    item_id: int,
    floor_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    plan, fl = _get_floor(db, item_id, floor_id, user)
    if plan.locked_at:
        raise HTTPException(409, "Plan kilitli.")
    n = db.scalar(
        select(func.count()).select_from(EmergencyPlanFloor).where(EmergencyPlanFloor.plan_id == plan.id)
    ) or 0
    if n <= 1:
        raise HTTPException(409, "Son kat silinemez.")
    db.delete(fl)
    plan.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


@ep_router.post("/{item_id}/floors/{floor_id}/background", response_model=EmergencyFloorResponse)
async def upload_floor_background(
    item_id: int,
    floor_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    plan, fl = _get_floor(db, item_id, floor_id, user)
    if plan.locked_at:
        raise HTTPException(409, "Plan kilitli.")
    name = file.filename or "plan.png"
    data = await file.read()
    if len(data) > 8_000_000:
        raise HTTPException(413, "Görsel 8 MB sınırını aşıyor.")
    ext = Path(name).suffix.lower() or ".png"
    if ext not in (".png", ".jpg", ".jpeg", ".webp"):
        raise HTTPException(400, "Yalnızca PNG/JPG/WEBP.")
    assert_safe_upload(data, ext, name)
    rel = f"{plan.company_id}/emergency-plans/{plan.id}/floor_{fl.id}_{uuid4().hex[:8]}{ext}"
    if settings.upload_gateway_enabled:
        persist_relative(data, relative_path=rel, original_name=name)
    else:
        target = Path(settings.upload_dir) / rel
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(data)
    fl.background_file_name = name
    fl.background_storage_path = rel.replace("\\", "/")
    fl.updated_at = datetime.utcnow()
    plan.updated_at = fl.updated_at
    db.commit()
    db.refresh(fl)
    return EmergencyFloorResponse.model_validate(fl)


@ep_router.get("/{item_id}/floors/{floor_id}/background")
def get_floor_background(
    item_id: int,
    floor_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*VIEW)),
):
    _plan, fl = _get_floor(db, item_id, floor_id, user)
    if not fl.background_storage_path:
        raise HTTPException(404, "Arka plan yok.")
    path = _safe_upload_path(fl.background_storage_path)
    if not path.is_file():
        raise HTTPException(404, "Dosya bulunamadı.")
    return FileResponse(
        path,
        filename=fl.background_file_name or path.name,
        media_type="image/png",
    )


@ep_router.post("/{item_id}/export-poster", response_model=EmergencyPlanResponse)
async def export_poster(
    item_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    """Editörden üretilen PNG/PDF posteri — kroki_* alanına yazar."""
    return await upload_kroki(item_id=item_id, file=file, db=db, user=user)


@ep_router.get("/{item_id}/legend")
def plan_legend(
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*VIEW)),
):
    """Sembol sayıları + acil ekipler özeti (lejant)."""
    plan = _get_plan(db, item_id, user)
    floors = list(
        db.scalars(
            select(EmergencyPlanFloor)
            .where(EmergencyPlanFloor.plan_id == plan.id)
            .order_by(EmergencyPlanFloor.sort_order)
        ).all()
    )
    symbol_counts: dict[str, int] = {}
    checks = {"exit": 0, "assembly": 0, "extinguisher": 0}
    for fl in floors:
        scene = _parse_scene(fl.scene_json)
        for obj in scene.get("objects") or []:
            t = str((obj or {}).get("type") or "")
            if not t:
                continue
            symbol_counts[t] = symbol_counts.get(t, 0) + 1
            if t in checks:
                checks[t] += 1
            elif t in ("door_exit",):
                checks["exit"] += 1
            elif t.startswith("extinguisher"):
                checks["extinguisher"] += 1
    teams = list(
        db.scalars(
            select(EmergencyTeam)
            .where(EmergencyTeam.company_id == plan.company_id, EmergencyTeam.is_active.is_(True))
            .order_by(EmergencyTeam.id)
            .limit(40)
        ).all()
    )
    team_out = []
    for t in teams:
        assigns = list(
            db.scalars(
                select(EmergencyTeamAssignment)
                .where(
                    EmergencyTeamAssignment.team_id == t.id,
                    EmergencyTeamAssignment.is_active.is_(True),
                )
                .limit(20)
            ).all()
        )
        members = []
        for a in assigns:
            emp = getattr(a, "employee", None)
            members.append(
                {
                    "name": (emp.full_name if emp else None) or "—",
                    "role": a.role_title or ("Lider" if a.is_leader else a.membership),
                    "phone": a.phone or None,
                }
            )
        team_out.append({"id": t.id, "name": t.name, "members": members})
    missing = []
    if checks["exit"] < 1:
        missing.append("En az bir acil çıkış gerekli")
    if checks["assembly"] < 1:
        missing.append("Toplanma alanı gerekli")
    if checks["extinguisher"] < 1:
        missing.append("Yangın söndürücü önerilir")
    return {
        "plan_id": plan.id,
        "title": plan.title,
        "revision_no": plan.revision_no,
        "locked": bool(plan.locked_at),
        "eyas_source_key": f"emergency:{plan.id}",
        "symbol_counts": symbol_counts,
        "checks": checks,
        "missing": missing,
        "teams": team_out,
        "floors": [{"id": f.id, "name": f.name, "sort_order": f.sort_order} for f in floors],
    }


# ─── Ortam ölçüm ───────────────────────────────────────────────────
wm_router = APIRouter(prefix="/workplace-measurements", tags=["Ortam Ölçüm"])


@wm_router.get("/meta")
def wm_meta(user: User = Depends(get_current_user)):
    return {"types": MEASUREMENT_TYPES, "engine": "workplace-measurement-v1"}


@wm_router.get("", response_model=list[MeasurementResponse])
def list_wm(
    company_id: int | None = None,
    q: str | None = Query(None, max_length=100),
    measurement_type: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles_or_workplace_manager(*PHYS_VIEW)),
):
    stmt = (
        select(WorkplaceMeasurement)
        .where(WorkplaceMeasurement.is_active.is_(True))
        .order_by(WorkplaceMeasurement.measured_at.desc())
    )
    ids = company_ids_for_query(db, user, company_id)
    if ids == []:
        return []
    if ids is not None:
        stmt = stmt.where(WorkplaceMeasurement.company_id.in_(ids))
    if measurement_type:
        stmt = stmt.where(WorkplaceMeasurement.measurement_type == measurement_type)
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                WorkplaceMeasurement.location.ilike(like),
                WorkplaceMeasurement.lab_name.ilike(like),
                WorkplaceMeasurement.report_ref.ilike(like),
            )
        )
    out = []
    for r in db.scalars(stmt.limit(2000)).all():
        item = MeasurementResponse.model_validate(r)
        item.review_status = _due_status(r.next_due_date)
        out.append(item)
    return out


@wm_router.post("", response_model=MeasurementResponse)
def create_wm(
    payload: MeasurementCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles_or_workplace_manager(*EDIT)),
):
    ensure_company_access(db, user, payload.company_id)
    if payload.measurement_type not in {t["code"] for t in MEASUREMENT_TYPES}:
        raise HTTPException(422, "Geçersiz ölçüm türü.")
    row = WorkplaceMeasurement(**payload.model_dump(), created_by_id=user.id)
    db.add(row)
    db.commit()
    db.refresh(row)
    resp = MeasurementResponse.model_validate(row)
    resp.review_status = _due_status(row.next_due_date)
    return resp


@wm_router.patch("/{item_id}", response_model=MeasurementResponse)
def update_wm(
    item_id: int,
    payload: MeasurementUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles_or_workplace_manager(*EDIT)),
):
    row = db.get(WorkplaceMeasurement, item_id)
    if not row:
        raise HTTPException(404, "Ölçüm bulunamadı.")
    ensure_company_access(db, user, row.company_id)
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(row, k, v)
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    resp = MeasurementResponse.model_validate(row)
    resp.review_status = _due_status(row.next_due_date)
    return resp


@wm_router.get("/export.xlsx")
def export_wm(
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles_or_workplace_manager(*PHYS_VIEW)),
):
    rows = list_wm(company_id=company_id, q=None, measurement_type=None, db=db, user=user)
    types = {t["code"]: t["label"] for t in MEASUREMENT_TYPES}
    data = [["Tür", "Yer", "Tarih", "Değer", "Birim", "Limit", "Laboratuvar", "Rapor", "Sonraki", "Durum"]]
    for r in rows:
        data.append([
            types.get(r.measurement_type, r.measurement_type),
            r.location or "",
            r.measured_at.isoformat(),
            r.value or "",
            r.unit or "",
            r.limit_value or "",
            r.lab_name or "",
            r.report_ref or "",
            r.next_due_date.isoformat() if r.next_due_date else "",
            r.review_status,
        ])
    stamp = datetime.now().strftime("%Y%m%d")
    return _xlsx(data, "Ortam Olcum", f"ortam-olcum-{stamp}.xlsx")


# ─── İSG Kurulu ────────────────────────────────────────────────────
oc_router = APIRouter(prefix="/ohs-committee", tags=["İSG Kurulu"])


@oc_router.get("/meta")
def oc_meta(user: User = Depends(get_current_user)):
    return {"roles": COMMITTEE_ROLES, "engine": "ohs-committee-v1"}


@oc_router.get("/members", response_model=list[CommitteeMemberResponse])
def list_members(
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*VIEW)),
):
    stmt = select(OhsCommitteeMember).where(OhsCommitteeMember.is_active.is_(True)).order_by(OhsCommitteeMember.id.desc())
    ids = company_ids_for_query(db, user, company_id)
    if ids == []:
        return []
    if ids is not None:
        stmt = stmt.where(OhsCommitteeMember.company_id.in_(ids))
    return list(db.scalars(stmt.limit(500)).all())


@oc_router.post("/members", response_model=CommitteeMemberResponse)
def create_member(
    payload: CommitteeMemberCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    ensure_company_access(db, user, payload.company_id)
    if payload.role_code not in {r["code"] for r in COMMITTEE_ROLES}:
        raise HTTPException(422, "Geçersiz kurul rolü.")
    row = OhsCommitteeMember(**payload.model_dump(), created_by_id=user.id)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@oc_router.delete("/members/{item_id}")
def deactivate_member(
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    row = db.get(OhsCommitteeMember, item_id)
    if not row:
        raise HTTPException(404, "Üye bulunamadı.")
    ensure_company_access(db, user, row.company_id)
    row.is_active = False
    db.commit()
    return {"ok": True}


@oc_router.get("/meetings", response_model=list[CommitteeMeetingResponse])
def list_meetings(
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*VIEW)),
):
    stmt = (
        select(OhsCommitteeMeeting)
        .where(OhsCommitteeMeeting.is_active.is_(True))
        .order_by(OhsCommitteeMeeting.meeting_date.desc())
    )
    ids = company_ids_for_query(db, user, company_id)
    if ids == []:
        return []
    if ids is not None:
        stmt = stmt.where(OhsCommitteeMeeting.company_id.in_(ids))
    return list(db.scalars(stmt.limit(500)).all())


@oc_router.post("/meetings", response_model=CommitteeMeetingResponse)
def create_meeting(
    payload: CommitteeMeetingCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    ensure_company_access(db, user, payload.company_id)
    row = OhsCommitteeMeeting(**payload.model_dump(), created_by_id=user.id)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@oc_router.get("/export.xlsx")
def export_committee(
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*VIEW)),
):
    members = list_members(company_id=company_id, db=db, user=user)
    meetings = list_meetings(company_id=company_id, db=db, user=user)
    roles = {r["code"]: r["label"] for r in COMMITTEE_ROLES}
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Uyeler"
    ws1.append(["Rol", "Ad Soyad", "Başlangıç", "Bitiş", "Not"])
    for m in members:
        ws1.append([
            roles.get(m.role_code, m.role_code),
            m.full_name,
            m.start_date.isoformat() if m.start_date else "",
            m.end_date.isoformat() if m.end_date else "",
            m.notes or "",
        ])
    ws2 = wb.create_sheet("Toplantilar")
    ws2.append(["Tarih", "Gündem", "Kararlar", "Katılımcılar", "Sonraki"])
    for m in meetings:
        ws2.append([
            m.meeting_date.isoformat(),
            m.agenda or "",
            m.decisions or "",
            m.attendees or "",
            m.next_meeting_date.isoformat() if m.next_meeting_date else "",
        ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    stamp = datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="isg-kurulu-{stamp}.xlsx"'},
    )


# ─── Belge onay (e-imza hazırlık) ──────────────────────────────────
da_router = APIRouter(prefix="/document-approvals", tags=["Belge Onay"])


@da_router.get("", response_model=list[DocumentApprovalResponse])
def list_approvals(
    company_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*VIEW)),
):
    stmt = select(DocumentApproval).where(DocumentApproval.is_active.is_(True)).order_by(DocumentApproval.id.desc())
    ids = company_ids_for_query(db, user, company_id)
    if ids == []:
        return []
    if ids is not None:
        stmt = stmt.where(DocumentApproval.company_id.in_(ids))
    return list(db.scalars(stmt.limit(500)).all())


@da_router.post("", response_model=DocumentApprovalResponse)
def create_approval(
    payload: DocumentApprovalCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    ensure_company_access(db, user, payload.company_id)
    row = DocumentApproval(**payload.model_dump(), created_by_id=user.id)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@da_router.post("/{item_id}/approve", response_model=DocumentApprovalResponse)
def mark_approved(
    item_id: int,
    signature_note: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    row = db.get(DocumentApproval, item_id)
    if not row:
        raise HTTPException(404, "Onay kaydı bulunamadı.")
    ensure_company_access(db, user, row.company_id)
    row.status = "Onaylandı"
    row.approved_at = date.today()
    if signature_note:
        row.signature_note = signature_note
    elif not row.signature_note:
        row.signature_note = f"Uygulama içi onay — {user.full_name} ({datetime.utcnow():%d.%m.%Y})"
    db.commit()
    db.refresh(row)
    return row


@da_router.post("/{item_id}/record-local-sign", response_model=DocumentApprovalResponse)
def record_local_sign(
    item_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    """Yerel ISG Suite Signer sonucunu onay kaydına işler — mevcut onay akışını bozmaz."""
    row = db.get(DocumentApproval, item_id)
    if not row:
        raise HTTPException(404, "Onay kaydı bulunamadı.")
    ensure_company_access(db, user, row.company_id)
    sha = str(payload.get("sha256") or "").strip()[:64]
    signer_cn = str((payload.get("signer") or {}).get("common_name") or payload.get("signer_cn") or "").strip()[:160]
    mode = str(payload.get("mode") or "").strip()[:40]
    sig_id = str(payload.get("signature_id") or "").strip()[:32]
    note = (
        f"Yerel e-imza köprüsü — {signer_cn or 'imzalayan'}; "
        f"SHA256={sha or '—'}; mode={mode or '—'}; id={sig_id or '—'}; "
        f"kullanıcı={user.full_name} ({datetime.utcnow():%d.%m.%Y %H:%M} UTC)"
    )
    row.signature_note = note[:1000]
    if payload.get("mark_approved", True):
        row.status = "Onaylandı"
        row.approved_at = date.today()
    db.commit()
    db.refresh(row)
    return row


@da_router.delete("/{item_id}")
def deactivate_approval(
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*EDIT)),
):
    """Belge onay kaydını soft-delete (listeden düşer; e-imza artifact’lara dokunmaz)."""
    row = db.get(DocumentApproval, item_id)
    if not row or not row.is_active:
        raise HTTPException(404, "Onay kaydı bulunamadı.")
    ensure_company_access(db, user, row.company_id)
    row.is_active = False
    db.commit()
    return {"ok": True}
