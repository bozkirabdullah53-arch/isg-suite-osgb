"""Yetkili firma kartı, durum raporu ve denetim hazırlık çıktıları."""
from __future__ import annotations

import json
import re
import zipfile
from datetime import datetime
from html import escape
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


_ASSETS = Path(__file__).resolve().parent.parent / "assets" / "fonts"
PDF_FONT = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"


def safe_filename(value: Any, *, fallback: str = "yetkili-firma") -> str:
    text = str(value or "").strip().lower()
    translate = str.maketrans("çğıöşü", "cgiosu")
    text = text.translate(translate)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return (text or fallback)[:80]


def _excel_safe(value: Any) -> Any:
    """Kullanıcı metninin Excel formülü olarak çalışmasını engeller."""
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    return "'" + value if value.lstrip().startswith(("=", "+", "-", "@")) else value


def _pdf_text(value: Any) -> str:
    return escape(str(value if value not in (None, "") else "—"))


def _register_fonts() -> None:
    global PDF_FONT, PDF_FONT_BOLD
    candidates = (
        (_ASSETS / "DejaVuSans.ttf", _ASSETS / "DejaVuSans-Bold.ttf"),
        (Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"), Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")),
    )
    for regular, bold in candidates:
        if not regular.exists():
            continue
        pdfmetrics.registerFont(TTFont("AuthorizedFirmDejaVu", str(regular)))
        pdfmetrics.registerFont(TTFont("AuthorizedFirmDejaVu-Bold", str(bold if bold.exists() else regular)))
        PDF_FONT = "AuthorizedFirmDejaVu"
        PDF_FONT_BOLD = "AuthorizedFirmDejaVu-Bold"
        return


def _style_header(row) -> None:
    fill = PatternFill("solid", fgColor="123B5D")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _finish_sheet(ws, widths: tuple[int, ...]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_authorized_firm_excel(detail: dict[str, Any]) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Firma Kartı"
    summary.append(["YETKİLİ FİRMA DOSYASI", _excel_safe(detail.get("firm_name"))])
    summary_rows = (
        ("İşyeri", detail.get("company_name")),
        ("OSGB", detail.get("osgb_name")),
        ("İl / İlçe", f"{detail.get('province') or '—'} / {detail.get('district') or '—'}"),
        ("Yetkili temsilci", detail.get("authorized_representative")),
        ("İletişim", detail.get("contact_email") or detail.get("contact_phone")),
        ("Çalışan sayısı", detail.get("employee_count")),
        ("Tehlike sınıfı", detail.get("hazard_class")),
        ("Yetki numarası", detail.get("authorization_number")),
        ("Yetki kapsamı", detail.get("authorization_scope")),
        ("Yetki başlangıç", detail.get("authorization_start_date")),
        ("Yetki bitiş", detail.get("authorization_expiry_date")),
        ("Kayıt durumu", detail.get("record_notice")),
        ("Uygunluk skoru", (detail.get("compliance_score") or {}).get("overall_score")),
        ("Kalite skoru", (detail.get("compliance_score") or {}).get("quality_score")),
        ("Üretim zamanı (UTC)", datetime.utcnow().isoformat(timespec="seconds")),
    )
    for label, value in summary_rows:
        summary.append([label, _excel_safe(value)])
    _style_header(summary[1])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 85
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    documents = wb.create_sheet("Belgeler")
    documents.append(["Belge", "Tür", "Zorunlu", "Başlangıç", "Bitiş", "Geçerlilik", "Kalan gün", "Not"])
    _style_header(documents[1])
    for item in detail.get("documents") or []:
        validity = item.get("validity") or {}
        documents.append([
            _excel_safe(item.get("title")), _excel_safe(item.get("document_type")),
            "Evet" if item.get("mandatory") else "Hayır", item.get("start_date") or "",
            item.get("expiry_date") or "", _excel_safe(validity.get("label")),
            validity.get("days_left"), _excel_safe(item.get("notes")),
        ])
    _finish_sheet(documents, (38, 20, 12, 14, 14, 24, 12, 60))

    professionals = wb.create_sheet("Profesyoneller")
    professionals.append(["Profesyonel", "Rol", "Belge no", "Belge bitiş", "Durum", "Puan", "Zorunlu dk", "Planlanan dk", "Gerçekleşen dk"])
    _style_header(professionals[1])
    for item in detail.get("professionals") or []:
        professionals.append([
            _excel_safe(item.get("full_name")), _excel_safe(item.get("professional_type")),
            _excel_safe(item.get("certificate_number")), item.get("certificate_expiry_date") or "",
            _excel_safe(item.get("status_label")), item.get("score"),
            item.get("required_minutes_monthly"), item.get("planned_minutes_monthly"),
            item.get("actual_minutes_monthly"),
        ])
    _finish_sheet(professionals, (30, 24, 22, 15, 24, 10, 14, 16, 17))

    score = detail.get("compliance_score") or {}
    categories = wb.create_sheet("Şeffaf Skor")
    categories.append(["Kategori", "Puan", "Ağırlık", "Geçti", "Detay", "Önerilen aksiyon"])
    _style_header(categories[1])
    for item in score.get("categories") or []:
        categories.append([
            _excel_safe(item.get("label")), item.get("score"), item.get("weight"),
            "Evet" if item.get("passed") else "Hayır", _excel_safe(item.get("detail")),
            _excel_safe(item.get("recommended_action")),
        ])
    _finish_sheet(categories, (32, 10, 10, 10, 70, 70))

    checklist = wb.create_sheet("Eksik Kontrol Listesi")
    checklist.append(["Öncelik", "Başlık", "Açıklama", "Önerilen aksiyon"])
    _style_header(checklist[1])
    blockers = score.get("critical_blockers") or []
    for item in blockers:
        checklist.append([
            "Kritik", _excel_safe(item.get("title")), _excel_safe(item.get("detail")),
            _excel_safe(item.get("recommended_action")),
        ])
    for item in score.get("failed_checks") or []:
        if any(blocker.get("code") == item.get("code") for blocker in blockers):
            continue
        checklist.append([
            "Gelişim", _excel_safe(item.get("label")), _excel_safe(item.get("detail")),
            _excel_safe(item.get("recommended_action")),
        ])
    _finish_sheet(checklist, (14, 36, 80, 80))

    history = wb.create_sheet("Skor Geçmişi")
    history.append(["Tarih", "Uygunluk skoru", "Kalite skoru", "Durum"])
    _style_header(history[1])
    for item in detail.get("score_history") or []:
        history.append([
            item.get("created_at") or "", item.get("overall_score"), item.get("quality_score"),
            _excel_safe(item.get("status")),
        ])
    _finish_sheet(history, (25, 18, 16, 20))

    privacy = wb.create_sheet("Gizlilik")
    privacy.append(["Kural", "Değer"])
    _style_header(privacy[1])
    privacy.append(["Sağlık verisi", "Yalnız anonim toplam; kişi veya klinik ayrıntı içermez."])
    privacy.append(["Skor yöntemi", _excel_safe(score.get("calculation"))])
    privacy.append(["Kara kutu", "Hayır"])
    privacy.column_dimensions["A"].width = 25
    privacy.column_dimensions["B"].width = 90

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _table_style() -> TableStyle:
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123B5D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), PDF_FONT),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])


def build_authorized_firm_pdf(detail: dict[str, Any]) -> bytes:
    _register_fonts()
    stream = BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Yetkili Firma Dosyası - {detail.get('firm_name') or ''}",
        author="İSG Suite",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle("AuthorizedFirmTitle", parent=styles["Title"], fontName=PDF_FONT_BOLD, fontSize=17, textColor=colors.HexColor("#123B5D"), spaceAfter=5)
    heading = ParagraphStyle("AuthorizedFirmHeading", parent=styles["Heading2"], fontName=PDF_FONT_BOLD, fontSize=11, textColor=colors.HexColor("#123B5D"), spaceBefore=6, spaceAfter=4)
    body = ParagraphStyle("AuthorizedFirmBody", parent=styles["BodyText"], fontName=PDF_FONT, fontSize=8, leading=10)
    small = ParagraphStyle("AuthorizedFirmSmall", parent=body, fontSize=7, leading=9)
    score = detail.get("compliance_score") or {}
    story: list[Any] = [
        Paragraph("YETKİLİ FİRMA UYGUNLUK DOSYASI", title),
        Paragraph(
            f"<b>{_pdf_text(detail.get('firm_name'))}</b> · İşyeri: {_pdf_text(detail.get('company_name'))} · "
            f"OSGB: {_pdf_text(detail.get('osgb_name'))} · Uygunluk: <b>{int(score.get('overall_score') or 0)}/100</b> · "
            f"Kalite: <b>{int(score.get('quality_score') or 0)}/100</b>",
            body,
        ),
        Paragraph(_pdf_text(detail.get("record_notice")), small),
        Spacer(1, 4 * mm),
    ]

    profile_rows = [
        ["İl / İlçe", "Temsilci", "Çalışan", "Tehlike", "Yetki no", "Başlangıç", "Bitiş"],
        [
            Paragraph(f"{_pdf_text(detail.get('province'))} / {_pdf_text(detail.get('district'))}", small),
            Paragraph(_pdf_text(detail.get("authorized_representative")), small),
            str(detail.get("employee_count") or 0),
            Paragraph(_pdf_text(detail.get("hazard_class")), small),
            Paragraph(_pdf_text(detail.get("authorization_number")), small),
            detail.get("authorization_start_date") or "—",
            detail.get("authorization_expiry_date") or "—",
        ],
    ]
    profile_table = Table(profile_rows, colWidths=[42 * mm, 55 * mm, 20 * mm, 31 * mm, 42 * mm, 32 * mm, 32 * mm], repeatRows=1)
    profile_table.setStyle(_table_style())
    story.extend([profile_table, Paragraph("Şeffaf uygunluk skoru", heading)])

    category_rows = [["Kategori", "Puan", "Ağırlık", "Detay", "Önerilen aksiyon"]]
    for item in score.get("categories") or []:
        category_rows.append([
            Paragraph(_pdf_text(item.get("label")), small), str(item.get("score") or 0),
            str(item.get("weight") or 0), Paragraph(_pdf_text(item.get("detail")), small),
            Paragraph(_pdf_text(item.get("recommended_action")), small),
        ])
    category_table = Table(category_rows, colWidths=[44 * mm, 18 * mm, 18 * mm, 92 * mm, 94 * mm], repeatRows=1)
    category_table.setStyle(_table_style())
    story.extend([category_table, Paragraph("Belge geçerliliği", heading)])

    document_rows = [["Belge", "Tür", "Zorunlu", "Başlangıç", "Bitiş", "Durum", "Kalan gün"]]
    for item in detail.get("documents") or []:
        validity = item.get("validity") or {}
        document_rows.append([
            Paragraph(_pdf_text(item.get("title")), small), Paragraph(_pdf_text(item.get("document_type")), small),
            "Evet" if item.get("mandatory") else "Hayır", item.get("start_date") or "—",
            item.get("expiry_date") or "—", Paragraph(_pdf_text(validity.get("label")), small),
            str(validity.get("days_left") if validity.get("days_left") is not None else "—"),
        ])
    document_table = Table(document_rows, colWidths=[58 * mm, 35 * mm, 22 * mm, 31 * mm, 31 * mm, 55 * mm, 24 * mm], repeatRows=1)
    document_table.setStyle(_table_style())
    story.extend([document_table, PageBreak(), Paragraph("Profesyonel uygunluğu", heading)])

    professional_rows = [["Profesyonel", "Rol", "Belge no", "Bitiş", "Durum", "Puan", "Hizmet dk (Z/P/G)"]]
    for item in detail.get("professionals") or []:
        professional_rows.append([
            Paragraph(_pdf_text(item.get("full_name")), small), Paragraph(_pdf_text(item.get("professional_type")), small),
            Paragraph(_pdf_text(item.get("certificate_number")), small), item.get("certificate_expiry_date") or "—",
            Paragraph(_pdf_text(item.get("status_label")), small), str(item.get("score") or 0),
            f"{item.get('required_minutes_monthly') or 0}/{item.get('planned_minutes_monthly') or 0}/{item.get('actual_minutes_monthly') or 0}",
        ])
    professional_table = Table(professional_rows, colWidths=[55 * mm, 45 * mm, 35 * mm, 30 * mm, 42 * mm, 18 * mm, 41 * mm], repeatRows=1)
    professional_table.setStyle(_table_style())
    story.extend([professional_table, Paragraph("Eksik ve önerilen aksiyon listesi", heading)])

    blockers = score.get("critical_blockers") or []
    action_rows = [["Öncelik", "Başlık", "Açıklama", "Önerilen aksiyon"]]
    for item in blockers:
        action_rows.append([
            "Kritik", Paragraph(_pdf_text(item.get("title")), small), Paragraph(_pdf_text(item.get("detail")), small),
            Paragraph(_pdf_text(item.get("recommended_action")), small),
        ])
    for item in score.get("failed_checks") or []:
        if any(blocker.get("code") == item.get("code") for blocker in blockers):
            continue
        action_rows.append([
            "Gelişim", Paragraph(_pdf_text(item.get("label")), small), Paragraph(_pdf_text(item.get("detail")), small),
            Paragraph(_pdf_text(item.get("recommended_action")), small),
        ])
    if len(action_rows) == 1:
        action_rows.append(["Bilgi", "Açık eksik yok", "Mevcut kontroller geçti.", "Periyodik gözden geçirmeyi sürdürün."])
    action_table = Table(action_rows, colWidths=[22 * mm, 48 * mm, 96 * mm, 100 * mm], repeatRows=1)
    action_table.setStyle(_table_style())
    story.extend([
        action_table,
        Spacer(1, 5 * mm),
        Paragraph(
            f"Skor yöntemi: {_pdf_text(score.get('calculation'))}. Kara kutu kullanılmaz. "
            "Gizlilik: sağlık verileri yalnız anonim toplamlar olarak kullanılır; kişi ve klinik ayrıntı bu dosyada yer almaz.",
            small,
        ),
    ])
    doc.build(story)
    return stream.getvalue()


def build_status_report_excel(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Yetkili Firma Durumu"
    ws.append(["Firma", "İşyeri", "İl", "İlçe", "Aktif", "Tehlike", "Yetki bitiş", "Belge durumu", "Profesyonel", "Uygunluk", "Kalite", "Hazırlık", "Kritik engel"])
    _style_header(ws[1])
    for item in rows:
        ws.append([
            _excel_safe(item.get("firm_name")), _excel_safe(item.get("company_name")),
            _excel_safe(item.get("province")), _excel_safe(item.get("district")),
            "Evet" if item.get("is_active") else "Hayır", _excel_safe(item.get("hazard_class")),
            item.get("authorization_expiry_date") or "", _excel_safe(item.get("document_status")),
            item.get("professional_count"), item.get("compliance_score"), item.get("quality_score"),
            _excel_safe(item.get("readiness_label")), item.get("critical_blocker_count"),
        ])
    _finish_sheet(ws, (36, 36, 18, 20, 10, 16, 15, 18, 14, 12, 10, 20, 14))
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def build_inspection_package(detail: dict[str, Any]) -> bytes:
    """Firma dosyası + makine-okunur checklist; resmî başvuru/gönderim yapmaz."""
    base = safe_filename(detail.get("firm_name"))
    manifest = {
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "profile_id": detail.get("id"),
        "company_id": detail.get("company_id"),
        "firm_name": detail.get("firm_name"),
        "record_notice": detail.get("record_notice"),
        "score": detail.get("compliance_score"),
        "alerts": detail.get("alerts"),
        "privacy": {"health_mode": "aggregate_only", "sensitive_fields_exposed": False},
        "external_submission_performed": False,
    }
    stream = BytesIO()
    with zipfile.ZipFile(stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(f"{base}-firma-dosyasi.pdf", build_authorized_firm_pdf(detail))
        archive.writestr(f"{base}-firma-dosyasi.xlsx", build_authorized_firm_excel(detail))
        archive.writestr("eksik-kontrol-listesi.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        archive.writestr(
            "ACIKLAMA.txt",
            "Bu paket kurum içi denetim hazırlığı içindir. Harici gönderim veya resmî doğrulama yapılmamıştır.\n"
            "Sağlık verileri yalnız anonim toplamlarla değerlendirilmiştir.\n",
        )
    return stream.getvalue()
