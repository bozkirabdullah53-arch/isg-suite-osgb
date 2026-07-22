"""0.9.134 — Acil durum ekipleri Excel / PDF / görevlendirme yazısı çıktıları."""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

PDF_FONT = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"
_ASSETS = Path(__file__).resolve().parent.parent / "assets" / "fonts"
CREATOR_LINE = "İSG Suite OSGB"

CERT_LABEL = {
    "green": "Geçerli",
    "yellow": "30 gün içinde",
    "red": "Süresi dolmuş",
    "grey": "Kayıt yok",
}


def _register_pdf_fonts() -> None:
    global PDF_FONT, PDF_FONT_BOLD
    candidates = [
        (_ASSETS / "DejaVuSans.ttf", _ASSETS / "DejaVuSans-Bold.ttf", "EmgDejaVu", "EmgDejaVu-Bold"),
        (Path(r"C:\Windows\Fonts\arial.ttf"), Path(r"C:\Windows\Fonts\arialbd.ttf"), "EmgArial", "EmgArial-Bold"),
        (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
            "EmgDejaVu",
            "EmgDejaVu-Bold",
        ),
    ]
    for regular, bold, name, bold_name in candidates:
        if not regular.exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont(name, str(regular)))
            PDF_FONT = name
            if bold.exists():
                pdfmetrics.registerFont(TTFont(bold_name, str(bold)))
                PDF_FONT_BOLD = bold_name
            else:
                PDF_FONT_BOLD = name
            try:
                registerFontFamily(name, normal=name, bold=PDF_FONT_BOLD, italic=name, boldItalic=PDF_FONT_BOLD)
            except Exception:
                pass
            return
        except Exception:
            continue


_register_pdf_fonts()


def _fmt_date(d) -> str:
    if not d:
        return "—"
    if hasattr(d, "strftime"):
        return d.strftime("%d.%m.%Y")
    s = str(d)
    if len(s) >= 10 and s[4] == "-":
        y, m, day = s[:10].split("-")
        return f"{day}.{m}.{y}"
    return s


def _add_pdf_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont(PDF_FONT, 7)
    canvas.setFillColor(colors.HexColor("#6c757d"))
    canvas.drawString(doc.leftMargin, 10 * mm, f"Program tasarımı ve raporlama: {CREATOR_LINE}")
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 10 * mm, f"Sayfa {doc.page}")
    canvas.restoreState()


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def build_teams_excel(*, company, teams_data: list[dict]) -> bytes:
    """teams_data: [{team, assignments:[{...cert_status...}]}]"""
    wb = openpyxl.Workbook()
    wb.properties.creator = CREATOR_LINE
    wb.properties.title = "Acil Durum Ekipleri"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws = wb.active
    ws.title = "Ekip Üyeleri"
    ws.merge_cells("A1:L1")
    ws["A1"] = f"ACİL DURUM EKİPLERİ / DESTEK ELEMANLARI - {getattr(company, 'name', '')}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1a5276")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:L2")
    ws["A2"] = (
        f"SGK Sicil: {getattr(company, 'sgk_registry_no', None) or '—'} | "
        f"Tehlike Sınıfı: {getattr(company, 'hazard_class', None) or '—'} | "
        f"Tarih: {datetime.now().strftime('%d.%m.%Y')} | Program: {CREATOR_LINE}"
    )
    ws["A2"].font = Font(size=9, italic=True, color="6c757d")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "Ekip", "Ad Soyad", "Görev/Unvan", "Üyelik", "Lider",
        "Vardiya", "Bölüm", "Sicil No", "Telefon",
        "Belge Durumu", "Belge Bitiş", "Görev. Yazı No",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin

    row = 5
    for block in teams_data:
        team = block["team"]
        team_name = getattr(team, "name", "")
        members = block.get("assignments") or []
        if not members:
            ws.cell(row=row, column=1, value=team_name).border = thin
            ws.cell(row=row, column=2, value="(üye yok)").border = thin
            row += 1
            continue
        for m in members:
            data = [
                team_name,
                m.get("employee_name") or "—",
                m.get("role_title") or "—",
                "Asıl" if m.get("membership") == "asil" else "Yedek",
                "Evet" if m.get("is_leader") else "—",
                m.get("shift") or "—",
                m.get("section") or "—",
                m.get("personnel_no") or "—",
                m.get("phone") or "—",
                CERT_LABEL.get(m.get("cert_status", "grey"), "—"),
                _fmt_date(m.get("cert_valid_until")),
                m.get("letter_no") or "—",
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin
                cell.font = Font(size=9)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    for i, w in enumerate([20, 22, 18, 8, 7, 12, 16, 12, 15, 14, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    for sheet in wb.worksheets:
        sheet.oddFooter.center.text = CREATOR_LINE
        sheet.oddFooter.right.text = "Sayfa &P / &N"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# --------------------------------------------------------------------------- #
# PDF — ekip listesi
# --------------------------------------------------------------------------- #
def build_teams_pdf(*, company, teams_data: list[dict], specialist: str | None = None) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=15 * mm, leftMargin=15 * mm, topMargin=18 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = PDF_FONT
    title_style = ParagraphStyle(
        "EmgTitle", parent=styles["Title"], fontSize=16, fontName=PDF_FONT_BOLD,
        spaceAfter=4, textColor=colors.HexColor("#1a5276"), alignment=TA_CENTER,
    )
    subtitle = ParagraphStyle(
        "EmgSub", parent=styles["Normal"], fontSize=9, fontName=PDF_FONT,
        spaceAfter=2, alignment=TA_CENTER, textColor=colors.HexColor("#2c3e50"),
    )
    info = ParagraphStyle("EmgInfo", parent=styles["Normal"], fontSize=9, fontName=PDF_FONT, spaceAfter=2)
    section = ParagraphStyle("EmgSection", parent=styles["Normal"], fontSize=12, fontName=PDF_FONT_BOLD, spaceAfter=6)
    cell = ParagraphStyle("EmgCell", parent=styles["Normal"], fontSize=8, fontName=PDF_FONT, leading=10)

    elements = [
        Paragraph("ACİL DURUM EKİPLERİ / DESTEK ELEMANLARI", title_style),
        Paragraph(f"Hazırlanma Tarihi: {datetime.now().strftime('%d.%m.%Y')}", subtitle),
        Paragraph(f"Program: {CREATOR_LINE}", subtitle),
        Spacer(1, 6 * mm),
        Paragraph(f"<b>İşyeri:</b> {getattr(company, 'name', '—')}", info),
        Paragraph(f"<b>SGK Sicil No:</b> {getattr(company, 'sgk_registry_no', None) or '—'}", info),
        Paragraph(f"<b>Tehlike Sınıfı:</b> {getattr(company, 'hazard_class', None) or '—'}", info),
        Paragraph(f"<b>Adres:</b> {getattr(company, 'address', None) or '—'}", info),
        Paragraph(f"<b>İSG Uzmanı / Hazırlayan:</b> {specialist or '—'}", info),
        Spacer(1, 5 * mm),
    ]

    for block in teams_data:
        team = block["team"]
        members = block.get("assignments") or []
        status = block.get("status") or {}
        elements.append(
            Paragraph(
                f"<b>{getattr(team, 'name', '')}</b>  "
                f"({len(members)} üye · durum: {status.get('label', '—')})",
                section,
            )
        )
        rows = [["Ad Soyad", "Görev", "Üyelik", "Lider", "Vardiya", "Belge", "Bitiş"]]
        if not members:
            rows.append(["(üye yok)", "—", "—", "—", "—", "—", "—"])
        for m in members:
            rows.append([
                Paragraph(str(m.get("employee_name") or "—"), cell),
                Paragraph(str(m.get("role_title") or "—"), cell),
                "Asıl" if m.get("membership") == "asil" else "Yedek",
                "Evet" if m.get("is_leader") else "—",
                Paragraph(str(m.get("shift") or "—"), cell),
                CERT_LABEL.get(m.get("cert_status", "grey"), "—"),
                _fmt_date(m.get("cert_valid_until")),
            ])
        table = Table(rows, colWidths=[95, 90, 40, 35, 60, 65, 55])
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 4 * mm))

    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(
        "Not: Bu liste yol gösterici bir özettir; ekip mevcudu ve belge geçerlilikleri "
        "işyeri koşullarına göre kontrol edilmesi önerilir.",
        ParagraphStyle("EmgNote", parent=styles["Normal"], fontSize=8, fontName=PDF_FONT,
                       textColor=colors.HexColor("#6c757d")),
    ))

    doc.build(elements, onFirstPage=_add_pdf_footer, onLaterPages=_add_pdf_footer)
    buf.seek(0)
    return buf.read()


# --------------------------------------------------------------------------- #
# PDF — görevlendirme yazısı (tek üye)
# --------------------------------------------------------------------------- #
def build_assignment_letter_pdf(*, company, team, assignment, employee_name: str) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=20 * mm, leftMargin=20 * mm, topMargin=22 * mm, bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = PDF_FONT
    title_style = ParagraphStyle(
        "LetterTitle", parent=styles["Title"], fontSize=14, fontName=PDF_FONT_BOLD,
        spaceAfter=10, textColor=colors.HexColor("#1a5276"), alignment=TA_CENTER,
    )
    body = ParagraphStyle(
        "LetterBody", parent=styles["Normal"], fontSize=11, fontName=PDF_FONT,
        spaceAfter=8, alignment=TA_JUSTIFY, leading=16,
    )
    info = ParagraphStyle("LetterInfo", parent=styles["Normal"], fontSize=10, fontName=PDF_FONT, spaceAfter=3)

    membership_label = "asıl" if getattr(assignment, "membership", "asil") == "asil" else "yedek"
    role = getattr(assignment, "role_title", None) or getattr(team, "name", "acil durum ekibi")
    letter_no = getattr(assignment, "letter_no", None) or "—"
    letter_date = _fmt_date(getattr(assignment, "letter_date", None) or date.today())

    elements = [
        Paragraph("ACİL DURUM EKİBİ GÖREVLENDİRME YAZISI", title_style),
        Spacer(1, 4 * mm),
        Paragraph(f"<b>İşyeri:</b> {getattr(company, 'name', '—')}", info),
        Paragraph(f"<b>SGK Sicil No:</b> {getattr(company, 'sgk_registry_no', None) or '—'}", info),
        Paragraph(f"<b>Yazı No:</b> {letter_no}  &nbsp;&nbsp; <b>Tarih:</b> {letter_date}", info),
        Spacer(1, 6 * mm),
        Paragraph(
            f"Sayın <b>{employee_name}</b>,", body,
        ),
        Paragraph(
            f"İşyerimizde oluşturulan <b>{getattr(team, 'name', '')}</b> kapsamında, "
            f"<b>{membership_label}</b> üye olarak <b>{role}</b> görevini yürütmek üzere "
            "görevlendirilmiş bulunmaktasınız.",
            body,
        ),
        Paragraph(
            "Acil durumlarda üzerinize düşen görevleri yerine getirmeniz, ilgili eğitim ve "
            "tatbikatlara katılmanız beklenmektedir. Görev süresince ekip sorumlusu ile "
            "koordineli çalışmanız önerilir.",
            body,
        ),
        Spacer(1, 4 * mm),
    ]

    detail_rows = [
        ["Ekip", getattr(team, "name", "—")],
        ["Üyelik", "Asıl" if membership_label == "asıl" else "Yedek"],
        ["Görev / Unvan", getattr(assignment, "role_title", None) or "—"],
        ["Vardiya", getattr(assignment, "shift", None) or "—"],
        ["Bölüm", getattr(assignment, "section", None) or "—"],
        ["Görev Başlangıç", _fmt_date(getattr(assignment, "assign_start", None))],
    ]
    table = Table(detail_rows, colWidths=[120, 320])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
        ("FONTNAME", (0, 0), (0, -1), PDF_FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 16 * mm))

    sign = Table(
        [["İşveren / Vekili", "Görevlendirilen"],
         [getattr(assignment, "assigned_by", None) or " ", employee_name]],
        colWidths=[220, 220],
    )
    sign.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
        ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 1), (-1, 1), 22),
        ("LINEBELOW", (0, 1), (-1, 1), 0.5, colors.grey),
    ]))
    elements.append(sign)

    doc.build(elements, onFirstPage=_add_pdf_footer, onLaterPages=_add_pdf_footer)
    buf.seek(0)
    return buf.read()
