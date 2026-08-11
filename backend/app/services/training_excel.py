# -*- coding: utf-8 -*-
"""Excel/CSV çalışan listesi — ISG Pro 2026 egitim/app.py birebir parser portu.

Dönüş satırları Suite şeması: full_name, national_id_masked, job_title, department.
"""
from __future__ import annotations

import csv
import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.services.national_id_format import normalize_national_id


def _cell(v: Any) -> str:
    if v is None:
        return ""
    text = str(v).strip()
    # Excel bozulmuş boşluk / BOM
    text = text.replace("\ufeff", "").replace("\xa0", " ").strip()
    return "" if text.lower() in ("none", "nan") else text


def _norm(text: str) -> str:
    """Pro _norm_text birebir: Türkçe İ/ı + NFKD (combining dot temizliği)."""
    import unicodedata

    t = _cell(text).strip()
    # Büyük İ/I önce — Python lower() İ → i+combining-dot üretebilir
    t = t.replace("İ", "i").replace("I", "i").replace("ı", "i")
    t = t.lower().replace("ı", "i")
    t = unicodedata.normalize("NFKD", t)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    for a, b in (
        (" ", ""),
        ("_", ""),
        ("-", ""),
        (".", ""),
        ("/", ""),
        ("\\", ""),
        ("ğ", "g"),
        ("ü", "u"),
        ("ş", "s"),
        ("ö", "o"),
        ("ç", "c"),
        ("*", ""),
        ("(", ""),
        (")", ""),
    ):
        t = t.replace(a, b)
    return re.sub(r"[^a-z0-9]+", "", t)


def _tc_format(tc: str) -> str:
    normalized = normalize_national_id(tc)
    digits = re.sub(r"\D", "", normalized)
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}.{digits[9:]}"
    return normalized


# Pro _HEADER_ALIASES → Suite field keys
_HEADER_ALIASES: dict[str, str] = {
    "adsoyad": "full_name",
    "adisoyadi": "full_name",
    "adsoyadi": "full_name",
    "isimsoyisim": "full_name",
    "isimsoyad": "full_name",
    "advesoyad": "full_name",
    "adivesoyadi": "full_name",
    "personeladsoyad": "full_name",
    "personeladisoyadi": "full_name",
    "calisanadsoyad": "full_name",
    "calisanadisoyadi": "full_name",
    "namesurname": "full_name",
    "fullname": "full_name",
    "isim": "full_name",
    "personeladi": "full_name",
    "calisanadi": "full_name",
    "adisoyad": "full_name",
    "adiivesoyadi": "full_name",
    "calisaninadisoyadi": "full_name",
    "personelinadisoyadi": "full_name",
    "iscininadisoyadi": "full_name",
    "katilimci": "full_name",
    "katilimciadi": "full_name",
    "katilimciadisoyadi": "full_name",
    "ad": "_first",
    "adi": "_first",
    "firstname": "_first",
    "first": "_first",
    "soyad": "_last",
    "soyadi": "_last",
    "surname": "_last",
    "lastname": "_last",
    "last": "_last",
    "tc": "national_id_masked",
    "tckimlik": "national_id_masked",
    "tckimlikno": "national_id_masked",
    "tckimliknumarasi": "national_id_masked",
    "tcno": "national_id_masked",
    "tckn": "national_id_masked",
    "kimlik": "national_id_masked",
    "kimlikno": "national_id_masked",
    "kimliknumarasi": "national_id_masked",
    "bransgorev": "job_title",
    "gorev": "job_title",
    "gorevi": "job_title",
    "gorevmeslek": "job_title",
    "gorevimeslegi": "job_title",
    "meslek": "job_title",
    "meslegi": "job_title",
    "pozisyon": "job_title",
    "brans": "job_title",
    "unvan": "job_title",
    "unvani": "job_title",
    "jobtitle": "job_title",
    "gorevunvani": "job_title",
    "bolum": "department",
    "bolumu": "department",
    "departman": "department",
    "birim": "department",
    "calistigibolum": "department",
    "calismabolumu": "department",
    "isbolumu": "department",
    "lokasyon": "department",
    "isletmebolumu": "department",
    "department": "department",
    "servis": "department",
}


def _header_field(value: Any) -> str:
    n = _norm(str(value or ""))
    if not n:
        return ""
    if n in _HEADER_ALIASES:
        return _HEADER_ALIASES[n]
    # Bulanık: "Personel Adı ve Soyadı", "Çalışanın Adı Soyadı" vb.
    if n in ("soyad", "soyadi"):
        return "_last"
    if "soyad" in n:
        if n in ("ad", "adi") or n.startswith("ad") or "isim" in n or "personel" in n or "calisan" in n or "katilim" in n:
            return "full_name"
        return "full_name"
    if n in ("ad", "adi", "isim"):
        return "_first"
    return ""


def _sheet_rows(ws) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        while values and not _cell(values[-1]):
            values.pop()
        if any(_cell(v) for v in values):
            rows.append(values)
    return rows


def _row_is_repeated_header(row: list[Any], field_map: dict[int, str]) -> bool:
    matches = 0
    for idx in field_map:
        if idx < len(row) and _header_field(row[idx]):
            matches += 1
    return matches >= max(1, min(2, len(field_map)))


def _veri_cek_rows(headers: list[Any], rows: list[list[Any]]) -> list[dict]:
    field_map: dict[int, str] = {}
    for idx, col in enumerate(headers):
        field = _header_field(col)
        if field:
            field_map[idx] = field

    data: list[dict] = []
    for row in rows:
        if _row_is_repeated_header(row, field_map):
            continue
        item = {"full_name": "", "national_id_masked": "", "job_title": "", "department": ""}
        first = last = ""
        for idx, alan in field_map.items():
            value = _cell(row[idx] if idx < len(row) else "")
            if alan == "_first":
                first = value
            elif alan == "_last":
                last = value
            else:
                item[alan] = value
        if not item["full_name"]:
            item["full_name"] = " ".join(p for p in (first, last) if p).strip()
        if not item["full_name"]:
            continue
        if _header_field(item["full_name"]):
            continue
        if item["national_id_masked"]:
            item["national_id_masked"] = _tc_format(item["national_id_masked"])
        data.append(item)
    return data


def _extract_education_metadata(rows: list[list[Any]]) -> dict[str, str]:
    info: dict[str, str] = {}
    for row in rows[:40]:
        label = _norm(row[0] if row else "")
        values = [_cell(v) for v in row[1:] if _cell(v)]
        value = values[0] if values else ""
        if "egitiminadi" in label or label == "egitimadi":
            info["title"] = value
        elif "egitiminsaati" in label or "egitiminsuresi" in label or label == "egitimsuresi":
            info["duration"] = value
        elif "egitimturu" in label:
            info["training_type"] = value
        elif "egitimsekli" in label:
            info["delivery_method"] = value
    return info


def _best_employee_table(sheet_rows: list[list[Any]]) -> tuple[list[dict], int | None]:
    best_data: list[dict] = []
    best_score = -1
    best_header_index: int | None = None
    scan_limit = min(80, len(sheet_rows))
    for header_index, row in enumerate(sheet_rows[:scan_limit]):
        fields = [_header_field(value) for value in row]
        has_name = "full_name" in fields or ("_first" in fields and "_last" in fields)
        if not has_name:
            continue
        mapped_count = sum(1 for field in fields if field)
        data = _veri_cek_rows(row, sheet_rows[header_index + 1 :])
        score = len(data) * 100 + mapped_count * 10 - header_index
        if data and score > best_score:
            best_data = data
            best_score = score
            best_header_index = header_index
    if best_data:
        return best_data, best_header_index

    populated_columns: set[int] = set()
    for row in sheet_rows:
        for idx, value in enumerate(row):
            if _cell(value):
                populated_columns.add(idx)
    if len(populated_columns) == 1:
        index = next(iter(populated_columns))
        names: list[dict] = []
        ignored = {
            "ad",
            "adi",
            "adisoyadi",
            "adsoyad",
            "isim",
            "isimlistesi",
            "personel",
            "sira",
            "sirano",
            "no",
            "tckimlik",
            "sertifika",
            "egitimadi",
            "adisoyad",
            "soyad",
            "soyadi",
        }
        for row in sheet_rows:
            value = _cell(row[index] if index < len(row) else "")
            normalized = _norm(value)
            if not value or normalized in ignored or normalized.startswith("toplam"):
                continue
            if _header_field(value):
                continue
            names.append(
                {
                    "full_name": value,
                    "national_id_masked": "",
                    "job_title": "",
                    "department": "",
                }
            )
        if names:
            return names, 0

    # Son çare: iki+ kelimeli hücreleri ad-soyad say (başlık bulunamadıysa)
    fallback: list[dict] = []
    for row in sheet_rows:
        for value in row:
            text = _cell(value)
            if not text or _header_field(text):
                continue
            parts = [p for p in text.replace(",", " ").split() if p]
            if len(parts) >= 2 and len(text) <= 80 and not re.search(r"\d{6,}", text):
                fallback.append(
                    {
                        "full_name": text,
                        "national_id_masked": "",
                        "job_title": "",
                        "department": "",
                    }
                )
                break
    if len(fallback) >= 2:
        return fallback, 0
    return [], None


def _decode_csv_bytes(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1254", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _csv_rows_from_bytes(raw: bytes) -> list[list[str]]:
    text = _decode_csv_bytes(raw)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    return [
        row
        for row in csv.reader(text.splitlines(), delimiter=delimiter)
        if any(_cell(v) for v in row)
    ]


def extract_first_excel_logo_bytes(content: bytes) -> bytes | None:
    """Pro _extract_first_excel_logo — bellek içi ilk gömülü görsel."""
    try:
        wb = load_workbook(BytesIO(content), data_only=False)
    except Exception:
        return None
    try:
        for ws in wb.worksheets:
            images = getattr(ws, "_images", None) or []
            if not images:
                continue
            try:
                data = images[0]._data()
                if data:
                    return data
            except Exception:
                continue
    finally:
        wb.close()
    return None


def parse_employee_upload(
    content: bytes,
    filename: str | None = None,
) -> tuple[list[dict], dict[str, str], bytes | None]:
    """Pro _parse_uploaded_employee_file eşdeğeri.

    Dönüş: (satırlar, excel_meta, logo_bytes|None)
    """
    name = (filename or "").lower().strip()
    ext = name.rsplit(".", 1)[-1] if "." in name else "xlsx"

    if ext == "csv":
        rows = _csv_rows_from_bytes(content)
        data, _ = _best_employee_table(rows)
        if not data:
            raise ValueError(
                "CSV dosyasında katılımcı bulunamadı. Ad Soyad (veya Adı + Soyadı) sütunu gerekli."
            )
        return data, _extract_education_metadata(rows), None

    if ext not in ("xlsx", "xlsm"):
        raise ValueError(
            "Geçersiz dosya! Lütfen .xlsx, .xlsm veya .csv uzantılı bir çalışan listesi yükleyin."
        )

    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(
            "Excel dosyası okunamadı. Dosyayı Excel’de .xlsx olarak kaydedip tekrar yükleyin."
        ) from exc

    try:
        best_data: list[dict] = []
        best_info: dict[str, str] = {}
        best_score = -1
        for worksheet in workbook.worksheets:
            rows = _sheet_rows(worksheet)
            data, header_index = _best_employee_table(rows)
            score = len(data) * 1000 - (header_index if header_index is not None else 999)
            if data and score > best_score:
                best_data = data
                best_info = _extract_education_metadata(rows)
                best_score = score
    finally:
        workbook.close()

    if not best_data:
        raise ValueError(
            "Excel dosyasında katılımcı bulunamadı. Ad Soyad sütunu gerekli "
            "(Ad Soyad / Adı Soyadı). Açıklama satırlarının üstünde başlık satırı olmalı."
        )

    logo = extract_first_excel_logo_bytes(content)
    return best_data, best_info, logo


def parse_employees_xlsx(content: bytes) -> list[dict]:
    """Geriye uyumlu: yalnız satır listesi."""
    rows, _meta, _logo = parse_employee_upload(content, "liste.xlsx")
    return rows
