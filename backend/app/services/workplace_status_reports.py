"""İşyeri Durum Merkezi PDF ve Excel çıktıları."""
from __future__ import annotations

from html import escape
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_ASSETS = Path(__file__).resolve().parent.parent / "assets" / "fonts"
PDF_FONT = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"


def _excel_safe(value):
    """Kullanıcı metninin Excel formülü olarak çalışmasını engeller."""
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    text = value
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _pdf_text(value) -> str:
    return escape(str(value if value not in (None, "") else "—"))


def _register_fonts() -> None:
    global PDF_FONT, PDF_FONT_BOLD
    regular = _ASSETS / "DejaVuSans.ttf"
    bold = _ASSETS / "DejaVuSans-Bold.ttf"
    if regular.exists():
        pdfmetrics.registerFont(TTFont("StatusDejaVu", str(regular)))
        pdfmetrics.registerFont(TTFont("StatusDejaVu-Bold", str(bold if bold.exists() else regular)))
        PDF_FONT = "StatusDejaVu"
        PDF_FONT_BOLD = "StatusDejaVu-Bold"


def build_workplace_status_excel(payload: dict) -> bytes:
    company = payload.get("company") or {}
    center = payload.get("status_center") or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Durum Özeti"
    blue = PatternFill("solid", fgColor="0F4C81")
    white_bold = Font(color="FFFFFF", bold=True)

    ws.append(["İŞYERİ DURUM RAPORU", _excel_safe(company.get("name") or "—")])
    ws.append(["SGK Sicil", _excel_safe(company.get("sgk_registry_no") or "—")])
    ws.append(["Genel durum", _excel_safe(center.get("overall_label") or "—")])
    ws.append(["Tamamlanma", f"%{center.get('completion_pct', 0)}"])
    ws.append(["Üretim zamanı (UTC)", center.get("generated_at") or "—"])
    ws.append(["İBYS doğrulama", "Resmî doğrulama bekleniyor — hazır beyanı değildir"])
    for cell in ws[1]:
        cell.fill = blue
        cell.font = white_bold
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70

    items_ws = wb.create_sheet("Süreç Durumları")
    headers = ["Kod", "Süreç", "Durum", "Detay", "Sorumlu", "Kaynak", "Modül", "Kritik"]
    items_ws.append(headers)
    for cell in items_ws[1]:
        cell.fill = blue
        cell.font = white_bold
    for item in center.get("items") or []:
        items_ws.append([
            _excel_safe(item.get("code")), _excel_safe(item.get("title")),
            _excel_safe(item.get("status_label")), _excel_safe(item.get("detail")),
            _excel_safe(item.get("responsible_role")), _excel_safe(item.get("source")),
            _excel_safe(item.get("module")),
            "Evet" if item.get("critical") else "Hayır",
        ])
    items_ws.freeze_panes = "A2"
    items_ws.auto_filter.ref = items_ws.dimensions
    for width, column in zip((22, 34, 18, 72, 34, 34, 24, 12), "ABCDEFGH"):
        items_ws.column_dimensions[column].width = width
    for row in items_ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    deadline_ws = wb.create_sheet("Terminler")
    headers = ["Kaynak", "Başlık", "Termin", "Kalan Gün", "Durum", "Sorumlu", "Modül"]
    deadline_ws.append(headers)
    for cell in deadline_ws[1]:
        cell.fill = blue
        cell.font = white_bold
    for item in center.get("deadlines") or []:
        deadline_ws.append([
            _excel_safe(item.get("source")), _excel_safe(item.get("title")),
            _excel_safe(item.get("due_date")), item.get("days_left"),
            _excel_safe(item.get("status")), _excel_safe(item.get("responsible_role")),
            _excel_safe(item.get("module")),
        ])
    deadline_ws.freeze_panes = "A2"
    deadline_ws.auto_filter.ref = deadline_ws.dimensions
    for width, column in zip((24, 54, 15, 14, 16, 34, 24), "ABCDEFG"):
        deadline_ws.column_dimensions[column].width = width
    for row in deadline_ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def build_workplace_status_pdf(payload: dict) -> bytes:
    _register_fonts()
    company = payload.get("company") or {}
    center = payload.get("status_center") or {}
    stream = BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"İşyeri Durum Raporu - {company.get('name') or ''}",
        author="İSG Suite",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "StatusTitle", parent=styles["Title"], fontName=PDF_FONT_BOLD, fontSize=18,
        textColor=colors.HexColor("#0F4C81"), spaceAfter=6,
    )
    normal = ParagraphStyle("StatusNormal", parent=styles["BodyText"], fontName=PDF_FONT, fontSize=8.2, leading=10)
    small = ParagraphStyle("StatusSmall", parent=normal, fontSize=7.2, leading=9)
    story = [
        Paragraph("İŞYERİ DURUM RAPORU", title_style),
        Paragraph(
            f"<b>{_pdf_text(company.get('name'))}</b> · SGK Sicil: {_pdf_text(company.get('sgk_registry_no'))} · "
            f"Genel durum: <b>{_pdf_text(center.get('overall_label'))}</b> · Tamamlanma: <b>%{int(center.get('completion_pct') or 0)}</b>",
            normal,
        ),
        Spacer(1, 5 * mm),
    ]
    rows = [["Süreç", "Durum", "Detay", "Sorumlu", "Kaynak / Modül"]]
    for item in center.get("items") or []:
        rows.append([
            Paragraph(_pdf_text(item.get("title")), small),
            Paragraph(_pdf_text(item.get("status_label")), small),
            Paragraph(_pdf_text(item.get("detail")), small),
            Paragraph(_pdf_text(item.get("responsible_role")), small),
            Paragraph(f"{_pdf_text(item.get('source'))} / {_pdf_text(item.get('module'))}", small),
        ])
    table = Table(rows, colWidths=[42 * mm, 25 * mm, 88 * mm, 53 * mm, 58 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F4C81")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), PDF_FONT),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([table, Spacer(1, 5 * mm)])

    deadlines = center.get("deadlines") or []
    if deadlines:
        story.append(Paragraph("Yaklaşan ve gecikmiş terminler", title_style))
        deadline_rows = [["Kaynak", "Başlık", "Termin", "Kalan gün", "Sorumlu"]]
        for item in deadlines[:40]:
            deadline_rows.append([
                Paragraph(_pdf_text(item.get("source")), small),
                Paragraph(_pdf_text(item.get("title")), small),
                item.get("due_date") or "—",
                str(item.get("days_left") if item.get("days_left") is not None else "—"),
                Paragraph(_pdf_text(item.get("responsible_role")), small),
            ])
        deadline_table = Table(deadline_rows, colWidths=[38 * mm, 95 * mm, 28 * mm, 25 * mm, 80 * mm], repeatRows=1)
        deadline_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
            ("FONTNAME", (0, 1), (-1, -1), PDF_FONT),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("FONTSIZE", (0, 0), (-1, -1), 7.2),
        ]))
        story.append(deadline_table)

    story.extend([
        Spacer(1, 5 * mm),
        Paragraph(
            "Gizlilik: Sağlık verileri yalnız anonim toplamlar olarak gösterilir. Bu rapor resmî İBYS uygunluk onayı veya “İBYS Ready” beyanı değildir; resmî doğrulama ve kabul beklenmektedir.",
            small,
        ),
    ])
    doc.build(story)
    return stream.getvalue()
