"""Selected professional detail XLSX with a fixed ten-column layout."""
from __future__ import annotations

from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from sqlalchemy.orm import Session

from app.services.osgb_oversight import build_professional_performance
from app.services.professional_performance_excel import (
    EXPORT_VERSION,
    _GREEN,
    _GREEN_TEXT,
    _RED,
    _RED_TEXT,
    _TEAL_DARK,
    _TEAL_LIGHT,
    _THIN,
    _add_table,
    _apply_score_rules,
    _period_label,
    _role_label,
    _safe,
    _save,
    _set_title,
    _set_widths,
    _status_label,
    _style_rows,
    _write_header,
    _write_kpis,
)


def _control_sheet(
    wb: Workbook,
    *,
    title: str,
    sheet_name: str,
    professional_name: str,
    period_label: str,
    rows: list[dict[str, Any]],
    table_name: str,
    completed: bool,
) -> None:
    sheet = wb.create_sheet(sheet_name)
    _set_title(sheet, title, f"Profesyonel: {professional_name} · Dönem: {period_label}", 6)
    headers = ["Sıra", "İşyeri", "Kontrol", "Açıklama", "Mevzuat Dayanağı", "Sonuç"]
    _write_header(sheet, 4, headers)
    for index, item in enumerate(rows, start=1):
        values = [
            index,
            item.get("company_name") or "—",
            item.get("check_title") or item.get("check_code") or "—",
            item.get("detail") or "—",
            item.get("legal") or "—",
            "Tamamlandı" if completed else "Eksik",
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(4 + index, col, _safe(value))
        result_cell = sheet.cell(4 + index, 6)
        result_cell.fill = PatternFill("solid", fgColor=_GREEN if completed else _RED)
        result_cell.font = Font(color=_GREEN_TEXT if completed else _RED_TEXT, bold=True)
    if rows:
        _style_rows(sheet, 5, 4 + len(rows), 6, wrap_cols={2, 3, 4, 5})
        _add_table(sheet, f"A4:F{4 + len(rows)}", table_name)
    else:
        sheet.cell(5, 1, "Kayıt bulunamadı.")
        sheet.merge_cells("A5:F5")
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    _set_widths(sheet, {"A": 7, "B": 30, "C": 28, "D": 58, "E": 58, "F": 16})


def build_professional_performance_detail_xlsx_safe(
    db: Session,
    professional_id: int,
) -> tuple[bytes, str]:
    report = build_professional_performance(db, professional_id)
    pro = report.get("professional") or {}
    perf = report.get("performance") or {}
    period_label = _period_label(report.get("period") or {})
    generated = datetime.now().strftime("%d.%m.%Y %H:%M")
    name = str(pro.get("full_name") or f"Profesyonel {professional_id}")

    wb = Workbook()
    summary = wb.active
    summary.title = "Profesyonel Özeti"
    _set_title(
        summary,
        f"Profesyonel Performans Raporu — {name}",
        f"Dönem: {period_label} · Oluşturulma: {generated} · Sürüm: {EXPORT_VERSION}",
        10,
    )
    row = _write_kpis(
        summary,
        4,
        [
            ("Performans Puanı", f"%{int(perf.get('score') or 0)}"),
            ("Durum", _status_label(perf.get("status"))),
            ("İşyeri", int(perf.get("firm_count") or 0)),
            ("Tamamlanan", int(perf.get("completed_checks") or 0)),
            ("Eksik", int(perf.get("gap_count") or 0)),
        ],
        10,
    )
    identity = [
        ("Ad Soyad", name),
        ("Unvan", pro.get("role_label") or _role_label(pro.get("professional_type"))),
        ("Belge Sınıfı", pro.get("certificate_class") or "—"),
        ("Belge Numarası", pro.get("certificate_number") or "—"),
        ("Belge Tarihi", pro.get("certificate_date") or "—"),
        ("E-posta", pro.get("email") or "—"),
        ("Telefon", pro.get("phone") or "—"),
        ("Kayıt Durumu", "Aktif" if pro.get("is_active", True) else "Pasif"),
        ("Dönem", period_label),
        ("Tamamlanma", f"%{int(perf.get('completion_pct') or 0)}"),
    ]
    _write_header(summary, row, ["Alan", "Bilgi", "Alan", "Bilgi", "Alan", "Bilgi", "Alan", "Bilgi", "Alan", "Bilgi"])
    info_row = row + 1
    for index in range(0, len(identity), 5):
        chunk = identity[index:index + 5]
        col = 1
        for label, value in chunk:
            summary.cell(info_row, col, _safe(label))
            summary.cell(info_row, col + 1, _safe(value))
            summary.cell(info_row, col).font = Font(bold=True, color=_TEAL_DARK)
            summary.cell(info_row, col).fill = PatternFill("solid", fgColor=_TEAL_LIGHT)
            for target in (summary.cell(info_row, col), summary.cell(info_row, col + 1)):
                target.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
                target.alignment = Alignment(vertical="top", wrap_text=True)
            col += 2
        info_row += 1
    summary.sheet_view.showGridLines = False
    _set_widths(
        summary,
        {"A": 17, "B": 25, "C": 17, "D": 23, "E": 17, "F": 24, "G": 17, "H": 25, "I": 17, "J": 22},
    )

    _control_sheet(
        wb,
        title="Eksik / Tamamlanmayan Kontroller",
        sheet_name="Eksik Kontroller",
        professional_name=name,
        period_label=period_label,
        rows=list(report.get("incomplete") or []),
        table_name="ProfessionalDetailGaps",
        completed=False,
    )
    _control_sheet(
        wb,
        title="Tamamlanan Kontroller",
        sheet_name="Tamamlanan",
        professional_name=name,
        period_label=period_label,
        rows=list(report.get("completed") or []),
        table_name="ProfessionalDetailCompleted",
        completed=True,
    )

    firms = wb.create_sheet("Firma Checklist")
    _set_title(firms, "Firma Bazlı Checklist", f"Profesyonel: {name} · Dönem: {period_label}", 8)
    headers = ["Sıra", "İşyeri", "Tehlike Sınıfı", "Firma Puanı (%)", "Kontrol", "Sonuç", "Açıklama", "Mevzuat Dayanağı"]
    _write_header(firms, 4, headers)
    next_row = 5
    item_count = 0
    for firm in report.get("firms") or []:
        for check in firm.get("checks") or []:
            item_count += 1
            values = [
                item_count,
                firm.get("company_name") or "—",
                firm.get("hazard_class") or "—",
                int(firm.get("score") or 0),
                check.get("title") or check.get("code") or "—",
                "Tamamlandı" if check.get("passed") else "Eksik",
                check.get("detail") or "—",
                check.get("legal") or "—",
            ]
            for col, value in enumerate(values, start=1):
                firms.cell(next_row, col, _safe(value))
            result_cell = firms.cell(next_row, 6)
            result_cell.fill = PatternFill("solid", fgColor=_GREEN if check.get("passed") else _RED)
            result_cell.font = Font(color=_GREEN_TEXT if check.get("passed") else _RED_TEXT, bold=True)
            next_row += 1
    if item_count:
        _style_rows(firms, 5, next_row - 1, 8, wrap_cols={2, 3, 5, 7, 8})
        _add_table(firms, f"A4:H{next_row - 1}", "ProfessionalFirmChecklist")
        _apply_score_rules(firms, f"D5:D{next_row - 1}")
    else:
        firms.cell(5, 1, "Atanmış işyeri veya firma kontrol verisi bulunamadı.")
        firms.merge_cells("A5:H5")
    firms.freeze_panes = "A5"
    firms.sheet_view.showGridLines = False
    _set_widths(firms, {"A": 7, "B": 30, "C": 18, "D": 18, "E": 28, "F": 16, "G": 58, "H": 58})

    stamp = date.today().isoformat()
    return _save(wb), f"csgb-profesyonel-performans-detay-{professional_id}-{stamp}.xlsx"
