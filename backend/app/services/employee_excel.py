# -*- coding: utf-8 -*-
"""Personel Excel içe aktarma — esnek başlık + şablon."""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.services.national_id_format import normalize_national_id


def _cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    text = str(v).strip().replace("\ufeff", "").replace("\xa0", " ").strip()
    return "" if text.lower() in ("none", "nan") else text


def _norm(text: str) -> str:
    t = _cell(text).strip()
    t = t.replace("İ", "i").replace("I", "i").replace("ı", "i")
    t = t.lower().replace("ı", "i")
    t = unicodedata.normalize("NFKD", t)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    for a, b in (
        (" ", ""),
        ("_", ""),
        ("-", ""),
        (".", ""),
        ("/", ""),
        ("\\", ""),
        ("ğ", "g"),
        ("ü", "u"),
        ("ş", "s"),
        ("ö", "o"),
        ("ç", "c"),
        ("*", ""),
    ):
        t = t.replace(a, b)
    return re.sub(r"[^a-z0-9]+", "", t)


_HEADER_ALIASES: dict[str, str] = {
    "adsoyad": "full_name",
    "adisoyadi": "full_name",
    "adsoyadi": "full_name",
    "adivesoyadi": "full_name",
    "advesoyad": "full_name",
    "isimsoyisim": "full_name",
    "isimsoyad": "full_name",
    "namesurname": "full_name",
    "fullname": "full_name",
    "isim": "full_name",
    "personeladisoyadi": "full_name",
    "personeladsoyad": "full_name",
    "calisanadisoyadi": "full_name",
    "calisanadsoyad": "full_name",
    "tc": "national_id_masked",
    "tckimlik": "national_id_masked",
    "tckimlikno": "national_id_masked",
    "tckimliknumarasi": "national_id_masked",
    "tcno": "national_id_masked",
    "tckn": "national_id_masked",
    "kimlik": "national_id_masked",
    "kimlikno": "national_id_masked",
    "gorev": "job_title",
    "gorevi": "job_title",
    "bransgorev": "job_title",
    "bransgorevi": "job_title",
    "unvan": "job_title",
    "unvani": "job_title",
    "meslek": "job_title",
    "pozisyon": "job_title",
    "isegiristarihi": "start_date",
    "isegiris": "start_date",
    "giristarihi": "start_date",
    "baslangictarihi": "start_date",
    "startdate": "start_date",
    "engellihukumludurumu": "special_status",
    "engellihukumlu": "special_status",
    "engellihukumludurum": "special_status",
    "ozeldurum": "special_status",
    "specialstatus": "special_status",
    "departman": "department",
    "bolum": "department",
    "bolumu": "department",
    "birim": "department",
}


def map_header(value: Any) -> str:
    n = _norm(str(value or ""))
    if not n:
        return ""
    if n in _HEADER_ALIASES:
        return _HEADER_ALIASES[n]
    if "soyad" in n and ("ad" in n or "isim" in n or "personel" in n or "calisan" in n):
        return "full_name"
    if n in ("ad", "adi"):
        return "_first"
    if n in ("soyad", "soyadi"):
        return "_last"
    if "tc" in n or "kimlik" in n:
        return "national_id_masked"
    if "gorev" in n or "unvan" in n or "meslek" in n:
        return "job_title"
    if "giris" in n and "tarih" in n:
        return "start_date"
    if "engelli" in n or "hukumlu" in n or "ozeldurum" in n:
        return "special_status"
    if "departman" in n or n == "bolum" or n == "bolumu":
        return "department"
    return ""


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _personnel_sheet(wb):
    for name in wb.sheetnames:
        if _norm(name) == "personel":
            return wb[name]
    return wb.active


def _is_placeholder_name(value: str) -> bool:
    n = _norm(value)
    return n in {
        "ornek",
        "orneksatir",
        "ornekpersonel",
        "adisoyadiyaziniz",
        "adsoyadyaziniz",
        "ornek1",
        "ornek2",
    }


def _normalize_special_status(value: str | None) -> str | None:
    text = (value or "").strip()
    if not text:
        return None
    n = _norm(text)
    if n in {"yok", "hayir", "degil", "-", "bos", "yoktur"}:
        return None
    if n in {"engelli"}:
        return "Engelli"
    if n in {"hukumlu"}:
        return "Hükümlü"
    if n in {"engellivehukumlu", "engellihukumlu"}:
        return "Engelli ve Hükümlü"
    return text[:80]


def parse_employees_workbook(content: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(content), data_only=True)
    try:
        ws = _personnel_sheet(wb)
        rows_raw: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            while vals and not _cell(vals[-1]):
                vals.pop()
            if any(_cell(v) for v in vals):
                rows_raw.append(vals)
    finally:
        wb.close()
    if not rows_raw:
        return []

    # İlk 30 satırda gerçek sütun başlığını ara (banner/açıklama satırlarını atla).
    header_idx = None
    field_map: dict[int, str] = {}
    for i, row in enumerate(rows_raw[:30]):
        mapping = {idx: map_header(v) for idx, v in enumerate(row)}
        mapping = {k: v for k, v in mapping.items() if v}
        mapped = set(mapping.values())
        has_name = "full_name" in mapped or ("_first" in mapped and "_last" in mapped)
        if has_name and len(mapped) >= 2:
            header_idx = i
            field_map = mapping
            break
    if header_idx is None:
        raise ValueError(
            "Excel dosyasında 'Adı Soyadı' (veya Adı + Soyadı) sütunu bulunmalıdır. "
            "Şablonu indirip aynı başlıklarla doldurun."
        )

    out: list[dict] = []
    for row in rows_raw[header_idx + 1 :]:
        item = {
            "full_name": "",
            "national_id_masked": None,
            "job_title": None,
            "department": None,
            "start_date": None,
            "special_status": None,
        }
        first = last = ""
        for idx, key in field_map.items():
            raw = row[idx] if idx < len(row) else None
            if key == "_first":
                first = _cell(raw)
            elif key == "_last":
                last = _cell(raw)
            elif key == "full_name":
                item["full_name"] = _cell(raw)
            elif key == "national_id_masked":
                tc = normalize_national_id(_cell(raw))
                item["national_id_masked"] = tc or None
            elif key == "job_title":
                item["job_title"] = _cell(raw) or None
            elif key == "department":
                item["department"] = _cell(raw) or None
            elif key == "start_date":
                item["start_date"] = _parse_date(raw)
            elif key == "special_status":
                item["special_status"] = _normalize_special_status(_cell(raw))
        if not item["full_name"]:
            item["full_name"] = " ".join(p for p in (first, last) if p).strip()
        if not item["full_name"]:
            continue
        if map_header(item["full_name"]) or _is_placeholder_name(item["full_name"]):
            continue
        out.append(item)
    return out


TEMPLATE_HEADERS = [
    "Adı Soyadı",
    "TC Kimlik No",
    "Görevi",
    "İşe Giriş Tarihi",
    "Engelli/Hükümlü",
]

TEMPLATE_DATA_START = 5
TEMPLATE_DATA_ROWS = 100
SPECIAL_STATUS_CHOICES = ("Yok", "Engelli", "Hükümlü", "Engelli ve Hükümlü")

_TEAL = "0F766E"
_TEAL_DARK = "115E59"
_TEAL_LIGHT = "CCFBF1"
_HEADER_BG = "0F4C5C"
_WHITE = "FFFFFF"
_MUTED = "64748B"
_BORDER = "CBD5E1"
_ZEBRA = "F8FAFC"
_THIN = Side(style="thin", color=_BORDER)


def _style_header_cell(cell, *, required: bool = False) -> None:
    cell.font = Font(bold=True, color=_WHITE, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    if required:
        cell.font = Font(bold=True, color="FEF3C7", name="Calibri", size=11)


def build_import_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Personel"
    last_col = get_column_letter(len(TEMPLATE_HEADERS))
    last_data = TEMPLATE_DATA_START + TEMPLATE_DATA_ROWS - 1

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "İSG Suite OSGB — Personel Aktarım Şablonu"
    ws["A1"].font = Font(bold=True, color=_WHITE, name="Calibri", size=16)
    ws["A1"].fill = PatternFill("solid", fgColor=_TEAL)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        "Başlık satırını silmeyin. Sarı sütun zorunludur. Örnek satırları Ornek sayfasından kopyalayabilirsiniz. "
        "Doldurduğunuz bu dosyayı Personel Yönetimi → Excel Yükle ile aktarın."
    )
    ws["A2"].font = Font(color=_TEAL_DARK, name="Calibri", size=10)
    ws["A2"].fill = PatternFill("solid", fgColor=_TEAL_LIGHT)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = "Sütunlar: Adı Soyadı *  |  TC Kimlik No  |  Görevi  |  İşe Giriş Tarihi (GG.AA.YYYY)  |  Engelli/Hükümlü"
    ws["A3"].font = Font(color=_MUTED, name="Calibri", size=9, italic=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for idx, title in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(TEMPLATE_DATA_START - 1, idx, title)
        _style_header_cell(cell, required=(idx == 1))

    for row_idx in range(TEMPLATE_DATA_START, last_data + 1):
        for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
            cell = ws.cell(row_idx, col_idx, None)
            cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
            cell.alignment = Alignment(vertical="center")
            cell.font = Font(name="Calibri", size=11, color="0F172A")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_ZEBRA)
        ws.cell(row_idx, 2).number_format = "@"
        ws.cell(row_idx, 4).number_format = "DD.MM.YYYY"
        ws.cell(row_idx, 5).alignment = Alignment(horizontal="center", vertical="center")

    table = Table(displayName="PersonelListesi", ref=f"A{TEMPLATE_DATA_START - 1}:{last_col}{last_data}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    tc_len = DataValidation(
        type="textLength",
        operator="lessThanOrEqual",
        formula1="11",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="TC Kimlik No",
        error="TC Kimlik No 11 haneli olmalıdır. Boş bırakılabilir.",
        promptTitle="TC Kimlik No",
        prompt="11 haneli T.C. kimlik numarası. Excel'in bilimsel gösterime çevirmemesi için metin olarak yazın.",
        showInputMessage=True,
    )
    tc_len.add(f"B{TEMPLATE_DATA_START}:B{last_data}")
    ws.add_data_validation(tc_len)

    date_dv = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(1950,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="İşe Giriş Tarihi",
        error="Tarihi GG.AA.YYYY olarak girin (ör. 15.03.2024).",
        promptTitle="İşe Giriş Tarihi",
        prompt="GG.AA.YYYY — örnek: 15.03.2024",
        showInputMessage=True,
    )
    date_dv.add(f"D{TEMPLATE_DATA_START}:D{last_data}")
    ws.add_data_validation(date_dv)

    status_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(SPECIAL_STATUS_CHOICES) + '"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Engelli/Hükümlü",
        error="Listeden seçin: Yok, Engelli, Hükümlü, Engelli ve Hükümlü.",
        promptTitle="Engelli/Hükümlü",
        prompt="Açılır listeden seçin. Yok veya boş = özel durum yok.",
        showInputMessage=True,
    )
    status_dv.add(f"E{TEMPLATE_DATA_START}:E{last_data}")
    ws.add_data_validation(status_dv)

    ws.freeze_panes = f"A{TEMPLATE_DATA_START}"
    ws.auto_filter.ref = f"A{TEMPLATE_DATA_START - 1}:{last_col}{last_data}"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[TEMPLATE_DATA_START - 1].height = 24
    widths = (28, 18, 24, 20, 22)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = "1:4"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.protection.autoFilter = True
    ws.protection.sheet = False

    example = wb.create_sheet("Ornek")
    example["A1"] = "Örnek doldurma — bu sayfa yüklenmez"
    example["A1"].font = Font(bold=True, color=_WHITE, name="Calibri", size=14)
    example["A1"].fill = PatternFill("solid", fgColor=_TEAL)
    example.merge_cells("A1:E1")
    example["A2"] = "Bu satırları kopyalayıp Personel sayfasına yapıştırabilirsiniz. Yükleme yalnızca Personel sayfasını okur."
    example["A2"].font = Font(color=_TEAL_DARK, name="Calibri", size=10)
    example.merge_cells("A2:E2")
    for idx, title in enumerate(TEMPLATE_HEADERS, start=1):
        cell = example.cell(4, idx, title)
        _style_header_cell(cell, required=(idx == 1))
    samples = [
        ("Ali Veli", "12345678901", "Kaynakçı", date(2024, 1, 15), "Yok"),
        ("Ayşe Yılmaz", "98765432109", "Operatör", date(2023, 3, 15), "Engelli"),
        ("Mehmet Demir", "", "Forklift Operatörü", date(2022, 11, 1), "Hükümlü"),
    ]
    for offset, row in enumerate(samples):
        for col, value in enumerate(row, start=1):
            cell = example.cell(5 + offset, col, value)
            cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
            cell.font = Font(name="Calibri", size=11)
            if col == 2:
                cell.number_format = "@"
            if col == 4:
                cell.number_format = "DD.MM.YYYY"
    for idx, width in enumerate(widths, start=1):
        example.column_dimensions[get_column_letter(idx)].width = width
    example.row_dimensions[1].height = 26
    example.freeze_panes = "A5"
    example.sheet_view.showGridLines = False

    note = wb.create_sheet("Aciklama")
    note["A1"] = "Nasıl kullanılır"
    note["A1"].font = Font(bold=True, color=_WHITE, name="Calibri", size=14)
    note["A1"].fill = PatternFill("solid", fgColor=_TEAL)
    note.merge_cells("A1:C1")
    guide = [
        ("1", "Personel sayfasında sarı başlıklı Adı Soyadı sütununu doldurun."),
        ("2", "TC, görev, işe giriş ve Engelli/Hükümlü isteğe bağlıdır; boş bırakılabilir."),
        ("3", "İşe giriş tarihini 15.03.2024 biçiminde yazın veya Excel tarih seçicisini kullanın."),
        ("4", "Engelli/Hükümlü hücresindeki oka basıp listeden seçin."),
        ("5", "Başlık satırını, tabloyu ve sayfa adını değiştirmeyin."),
        ("6", "Dosyayı .xlsx olarak kaydedip uygulamada Excel Yükle ile aktarın."),
    ]
    note["A3"] = "Adım"
    note["B3"] = "İşlem"
    note["C3"] = "Zorunlu"
    for col in range(1, 4):
        _style_header_cell(note.cell(3, col))
    for idx, (step, text) in enumerate(guide, start=4):
        note.cell(idx, 1, step)
        note.cell(idx, 2, text)
        note.cell(idx, 3, "Evet" if idx == 4 else "Hayır")
        for col in range(1, 4):
            note.cell(idx, col).border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
            note.cell(idx, col).font = Font(name="Calibri", size=11)
    note["A11"] = "Sütun"
    note["B11"] = "Zorunlu"
    note["C11"] = "Açıklama"
    for col in range(1, 4):
        _style_header_cell(note.cell(11, col))
    columns = [
        ("Adı Soyadı", "Evet", "Personelin adı ve soyadı. Ad + Soyad ayrı sütun da kabul edilir."),
        ("TC Kimlik No", "Hayır", "11 hane. Excel bilimsel gösterime çevirmesin diye metin biçimindedir."),
        ("Görevi", "Hayır", "Branş / görev / unvan. Örn. Kaynakçı, Operatör."),
        ("İşe Giriş Tarihi", "Hayır", "GG.AA.YYYY veya YYYY-AA-GG."),
        ("Engelli/Hükümlü", "Hayır", "Yok, Engelli, Hükümlü, Engelli ve Hükümlü. Yok/boş = özel durum yok."),
    ]
    for idx, row in enumerate(columns, start=12):
        for col, value in enumerate(row, start=1):
            cell = note.cell(idx, col, value)
            cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    note.column_dimensions["A"].width = 22
    note.column_dimensions["B"].width = 12
    note.column_dimensions["C"].width = 78
    note.row_dimensions[1].height = 26
    note.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
