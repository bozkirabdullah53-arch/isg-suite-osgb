"""0.9.138 — Yıllık plan değerlendirme Excel/PDF (plan kayıtlarını değiştirmez)."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

_ASSETS = Path(__file__).resolve().parents[1] / "assets"
_FONT_REG = "EvalDejaVu"
_FONT_BOLD = "EvalDejaVu-Bold"

_CAT_GROUPS = (
    ("egitim", "Eğitim faaliyetleri", ("egitim", "isbasi", "egitimler")),
    ("saglik", "Sağlık gözetimi", ("saglik", "saglik_gozetimi")),
    ("risk", "Risk değerlendirme", ("risk", "risk_degerlendirme")),
    ("saha", "Saha gözetim / denetim", ("saha", "denetim", "uygunsuzluk")),
    ("periyodik", "Periyodik / ortam", ("periyodik", "olcum", "ortam")),
    ("tatbikat", "Acil durum / tatbikat", ("tatbikat", "acil")),
    ("olay", "İş kazası / ramak kala", ("kaza", "ramak", "olay")),
)


def _register_fonts() -> tuple[str, str]:
    candidates = [
        (_ASSETS / "DejaVuSans.ttf", _ASSETS / "DejaVuSans-Bold.ttf"),
        (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ),
        (Path("C:/Windows/Fonts/arial.ttf"), Path("C:/Windows/Fonts/arialbd.ttf")),
    ]
    for regular, bold in candidates:
        if not regular.exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont(_FONT_REG, str(regular)))
            pdfmetrics.registerFont(TTFont(_FONT_BOLD, str(bold if bold.exists() else regular)))
            registerFontFamily(_FONT_REG, normal=_FONT_REG, bold=_FONT_BOLD)
            return _FONT_REG, _FONT_BOLD
        except Exception:
            continue
    return "Helvetica", "Helvetica-Bold"


def build_eval_xlsx(
    *,
    company_name: str,
    year: int,
    items: list[dict[str, Any]],
    unplanned: list[dict[str, Any]],
    kpis: dict[str, Any],
    capas: list[dict[str, Any]] | None = None,
    suggestions: list[dict[str, Any]] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Genel ozet"
    ws.append(["Firma", company_name])
    ws.append(["Yil", year])
    for k, v in kpis.items():
        if k == "formulas":
            continue
        ws.append([k, str(v)])

    w2 = wb.create_sheet("Faaliyetler")
    w2.append([
        "Plan ID", "Kategori", "Faaliyet", "Ay", "Hedef", "Sorumlu",
        "Durum", "Gerceklesme", "Oran", "Gecikme", "Kanit", "Sonuc",
    ])
    for it in items:
        plan = it.get("plan") or {}
        w2.append([
            plan.get("id"),
            plan.get("category"),
            plan.get("activity"),
            plan.get("month"),
            str(plan.get("target_date") or ""),
            plan.get("responsible_name"),
            it.get("outcome_status"),
            str(it.get("actual_end") or ""),
            it.get("completion_pct"),
            it.get("delay_days"),
            it.get("evidence_count"),
            (it.get("result_text") or "")[:500],
        ])

    w_miss = wb.create_sheet("Gerceklesmeyen")
    w_miss.append(["Plan ID", "Faaliyet", "Durum", "Gerekce", "Sonraki yil onerisi"])
    for it in items:
        if it.get("outcome_status") not in ("gerceklesmedi", "ertelendi"):
            continue
        plan = it.get("plan") or {}
        w_miss.append([
            plan.get("id"),
            plan.get("activity"),
            it.get("outcome_status"),
            (it.get("deviation_reason") or "")[:400],
            (it.get("next_year_suggestion") or "")[:400],
        ])

    w3 = wb.create_sheet("Plan disi")
    w3.append(["Faaliyet", "Kategori", "Tarih", "Neden", "Sonuc", "Sonraki yila oner"])
    for u in unplanned:
        w3.append([
            u.get("activity"),
            u.get("category"),
            str(u.get("done_date") or ""),
            (u.get("reason") or "")[:300],
            (u.get("result_text") or "")[:300],
            u.get("suggest_next_year"),
        ])

    w4 = wb.create_sheet("Duzeltici faaliyetler")
    w4.append(["Baslik", "Sorumlu", "Hedef", "Oncelik", "Durum", "Kok neden"])
    for c in capas or []:
        w4.append([
            c.get("title"),
            c.get("responsible"),
            str(c.get("due_date") or ""),
            c.get("priority"),
            c.get("status"),
            (c.get("root_cause") or "")[:300],
        ])

    w5 = wb.create_sheet("Sonraki yil onerileri")
    w5.append(["Faaliyet", "Neden", "Oneri"])
    for s in suggestions or []:
        w5.append([s.get("activity"), s.get("reason"), (s.get("suggestion") or "")[:400]])

    w6 = wb.create_sheet("Performans")
    w6.append(["Gosterge", "Deger", "Aciklama"])
    formulas = kpis.get("formulas") or {}
    for key in ("completion_rate", "on_time_rate", "evidence_rate", "planned_total", "tamam", "missing_evidence"):
        w6.append([key, kpis.get(key), formulas.get(key, "")])

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def build_eval_pdf(
    *,
    company_name: str,
    year: int,
    kpis: dict[str, Any],
    items: list[dict[str, Any]],
    unplanned: list[dict[str, Any]] | None = None,
    suggestions: list[dict[str, Any]] | None = None,
    capas: list[dict[str, Any]] | None = None,
    meta: dict[str, Any] | None = None,
    related: dict[str, Any] | None = None,
    compare: dict[str, Any] | None = None,
) -> bytes:
    font, font_b = _register_fonts()
    stream = BytesIO()
    pdf = canvas.Canvas(stream, pagesize=A4)
    w, h = A4
    y = h - 40
    meta = meta or {}
    page_no = [1]

    def footer():
        pdf.setFont(font, 7)
        pdf.drawString(40, 24, f"Doküman: YPD-{year} | Rev: {meta.get('report_status') or '-'} | Sayfa {page_no[0]}")
        pdf.drawRightString(w - 40, 24, "Plan alanları bu raporda değiştirilmez.")

    def line(txt: str, size: int = 9, bold: bool = False, gap: int = 12):
        nonlocal y
        if y < 56:
            footer()
            pdf.showPage()
            page_no[0] += 1
            y = h - 40
        pdf.setFont(font_b if bold else font, size)
        pdf.drawString(40, y, (txt or "")[:112])
        y -= gap

    # Kapak / işyeri
    line(f"Yıllık Çalışma Planı Değerlendirme Raporu — {year}", 14, True, 18)
    line(f"Firma / işyeri: {company_name}", 10)
    line(f"SGK: {meta.get('sgk_registry_no') or '—'} | Tehlike: {meta.get('hazard_class') or '—'} | Çalışan: {meta.get('employee_count') or '—'}", 9)
    line(f"Adres: {(meta.get('address') or '—')[:90]}", 9)
    line(
        f"Uzman: {meta.get('specialist_name') or '—'} | Hekim: {meta.get('physician_name') or '—'} | "
        f"İşveren: {meta.get('employer_name') or '—'}",
        9,
    )
    line(f"Değerlendirme durumu: {meta.get('report_status') or '—'} | Rapor tarihi: {meta.get('report_date') or '—'}", 9, gap=14)
    line("1. Amaç ve kapsam", 11, True, 14)
    line(f"Bu rapor {year} yılı yıllık çalışma planı faaliyetlerinin gerçekleşme, sapma ve kanıt durumunu değerlendirir.", 9)
    line("2. Gerçekleşme özeti", 11, True, 14)
    line(
        f"Planlanan: {kpis.get('planned_total')} | Tamam: {kpis.get('tamam')} | Kısmi: {kpis.get('kismi')} | "
        f"Ertelenen: {kpis.get('ertelendi')} | Gerçekleşmeyen: {kpis.get('gerceklesmedi')}",
        9,
    )
    line(
        f"Gerçekleşme: {kpis.get('completion_rate')}% | Zamanında: {kpis.get('on_time_rate')}% | "
        f"Kanıtlı: {kpis.get('evidence_rate')}% | Plan dışı: {kpis.get('unplanned')}",
        9,
        gap=14,
    )
    if compare:
        line("2.1 Yıllar arası karşılaştırma", 10, True, 12)
        line(
            f"{compare.get('prev_year')}: gerçekleşme {compare.get('prev_rate')}% "
            f"({compare.get('prev_planned')} plan) → {year}: {compare.get('curr_rate')}% "
            f"({compare.get('curr_planned')} plan)",
            9,
            gap=14,
        )

    line("3. Faaliyet bazlı değerlendirme", 11, True, 14)
    for it in items:
        plan = it.get("plan") or {}
        line(
            f"• {plan.get('activity', '')[:46]} | {it.get('outcome_status')} | "
            f"{it.get('actual_end') or '-'} | kanıt:{it.get('evidence_count') or 0}"
        )

    sec = 4
    for _code, title, keys in _CAT_GROUPS:
        subset = [
            it for it in items
            if any(k in (it.get("plan") or {}).get("category") or "" for k in keys)
        ]
        line(f"{sec}. {title}", 11, True, 14)
        if not subset:
            line("Bu kategoride plan kalemi yok / eşleşmedi.")
        for it in subset[:25]:
            plan = it.get("plan") or {}
            line(f"• {plan.get('activity', '')[:55]} — {it.get('outcome_status')}")
        sec += 1

    if related:
        line(f"{sec}. Modül özetleri (toplulaştırılmış)", 11, True, 14)
        hs = related.get("health_summary") or {}
        line(
            f"Eğitim {related.get('trainings', {}).get('count', 0)} | Tatbikat {related.get('drills', {}).get('count', 0)} | "
            f"Olay {related.get('incidents', {}).get('count', 0)} | Risk {related.get('risks', {}).get('count', 0)} | "
            f"Sağlık muayene {hs.get('exams_completed', 0)}",
            9,
            gap=14,
        )
        sec += 1

    line(f"{sec}. Uygunsuzluklar / düzeltici faaliyetler", 11, True, 14)
    if not capas:
        line("Açık/kayıtlı DÖF yok.")
    for c in capas or []:
        line(f"• {c.get('title', '')[:60]} — {c.get('status')} / {c.get('due_date') or '-'}")
    sec += 1

    line(f"{sec}. Plan dışı gerçekleştirilen çalışmalar", 11, True, 14)
    if not unplanned:
        line("Kayıt yok.")
    for u in unplanned or []:
        line(f"• {u.get('activity', '')[:60]} ({u.get('done_date') or '-'})")
    sec += 1

    line(f"{sec}. Gerçekleştirilemeyen faaliyetler", 11, True, 14)
    miss = [it for it in items if it.get("outcome_status") in ("gerceklesmedi", "ertelendi")]
    if not miss:
        line("Kayıt yok.")
    for it in miss:
        plan = it.get("plan") or {}
        line(f"• {plan.get('activity', '')[:50]} — {(it.get('deviation_reason') or '')[:40]}")
    sec += 1

    line(f"{sec}. Bir sonraki yıl önerileri", 11, True, 14)
    if not suggestions:
        line("Öneri yok.")
    for s in suggestions or []:
        line(f"• {s.get('activity', '')[:50]} — {s.get('reason')}")
    sec += 1

    line(f"{sec}. Sonuç ve onay", 11, True, 14)
    line("Plan dışı faaliyetler gerçekleşme oranına dahil edilmemiştir.", 9)
    y -= 8
    line("İş Güvenliği Uzmanı: ____________________    Tarih: ________", 9)
    line("İşyeri Hekimi: __________________________    Tarih: ________", 9)
    line("İşveren / Vekil: ________________________    Tarih: ________", 9)
    footer()
    pdf.save()
    return stream.getvalue()
