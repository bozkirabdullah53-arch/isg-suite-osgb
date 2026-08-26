"""Risk değerlendirme PDF / Excel raporları — İSG PRO reports.py Suite uyarlaması."""
from __future__ import annotations

import json
import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image as ReportLabImage,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.pdfgen import canvas

from app.services.risk_validity import METHOD_LABEL, document_meta_rows
from app.services.risk_hazop import guide_word_label, normalize_hazop_data, priority_details
from app.services.risk_methods import resolve_method
from app.services.risk_scoring import canonical_risk_level, fine_kinney_level_details

PDF_FONT = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"
_ASSETS = Path(__file__).resolve().parent.parent / "assets" / "fonts"
CREATOR_LINE = "İSG Suite OSGB"
_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}
logger = logging.getLogger(__name__)


def _upload_root() -> Path:
    try:
        from app.core.config import settings

        root = Path(settings.upload_dir).resolve()
    except Exception:
        root = Path("uploads").resolve()
    root.mkdir(parents=True, exist_ok=True)
    return root


def _risk_excel_photo_path(risk) -> str | None:
    """Risk kaydına bağlı ilk uygun fotoğraf dosyasını döndürür (PRO parity)."""
    try:
        media_items = list(getattr(risk, "media_files", None) or [])
    except Exception:
        media_items = []
    root = _upload_root()
    for media in media_items:
        name = getattr(media, "original_name", "") or ""
        rel = getattr(media, "storage_path", "") or ""
        ctype = (getattr(media, "content_type", "") or "").lower()
        ext = Path(name or rel).suffix.lower()
        if not (ctype.startswith("image/") or ext in _IMAGE_EXTS):
            continue
        candidate = (root / rel).resolve()
        if root in candidate.parents and candidate.exists():
            return str(candidate)
    return None


def _item_value(item, key: str, default=None):
    if isinstance(item, dict):
        return item.get(key, default)
    return getattr(item, key, default)


def _media_is_photo(media) -> bool:
    file_type = str(_item_value(media, "file_type", "") or "").lower()
    content_type = str(_item_value(media, "content_type", "") or "").lower()
    name = str(_item_value(media, "original_name", "") or "")
    rel = str(_item_value(media, "storage_path", "") or "")
    return file_type == "photo" or content_type.startswith("image/") or Path(name or rel).suffix.lower() in _IMAGE_EXTS


def _read_media_bytes(media) -> bytes | None:
    """Read an attached image from inline draft bytes or the active object store."""
    inline = _item_value(media, "inline_bytes")
    if isinstance(inline, (bytes, bytearray)):
        return bytes(inline)
    key = str(_item_value(media, "storage_path", "") or "").replace("\\", "/").strip("/")
    if not key or ".." in key.split("/"):
        return None
    try:
        from app.services.object_store import get_object_store

        return get_object_store().get_bytes(key)
    except Exception:
        logger.warning("Risk raporu fotoğrafı okunamadı: %s", key, exc_info=True)
        return None


def _normalize_photo_bytes(raw: bytes) -> bytes:
    """Normalize common mobile image formats to a safe, embeddable JPEG."""
    from PIL import Image, ImageOps

    with Image.open(BytesIO(raw)) as source:
        image = ImageOps.exif_transpose(source).convert("RGB")
        image.thumbnail((2200, 1800), Image.Resampling.LANCZOS)
        output = BytesIO()
        image.save(output, format="JPEG", quality=88, optimize=True)
        return output.getvalue()


def _photo_flowable(media, *, max_width: float = 220, max_height: float = 88 * mm):
    raw = _read_media_bytes(media)
    if not raw:
        return None
    try:
        normalized = _normalize_photo_bytes(raw)
        reader = ImageReader(BytesIO(normalized))
        source_width, source_height = reader.getSize()
        if not source_width or not source_height:
            return None
        scale = min(max_width / source_width, max_height / source_height, 1)
        image = ReportLabImage(
            BytesIO(normalized),
            width=max(1, source_width * scale),
            height=max(1, source_height * scale),
        )
        image.hAlign = "CENTER"
        return image
    except Exception:
        logger.warning("Risk raporu fotoğrafı PDF'e gömülemedi", exc_info=True)
        return None


def _fmt_datetime(value) -> str:
    if not value:
        return "—"
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y %H:%M")
    raw = str(value)
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return raw


def _confidence_percent(value) -> int:
    try:
        return max(0, min(100, round(float(value or 0) * 100)))
    except (TypeError, ValueError):
        return 0


def _media_caption(media, index: int, caption_style) -> Paragraph:
    from html import escape

    name = escape(str(_item_value(media, "original_name", "") or f"Saha fotoğrafı {index}"))
    lines = [f"<b>Fotoğraf {index}</b> · {name}"]
    captured_at = _item_value(media, "captured_at")
    if captured_at:
        lines.append(f"Çekim: {escape(_fmt_datetime(captured_at))}")
    lat = _item_value(media, "gps_lat")
    lng = _item_value(media, "gps_lng")
    if lat is not None and lng is not None:
        accuracy = _item_value(media, "gps_accuracy_m")
        suffix = f" · ±{accuracy} m" if accuracy is not None else ""
        lines.append(f"GPS: {escape(str(lat))}, {escape(str(lng))}{suffix}")
    tags = _item_value(media, "tags")
    if not tags:
        raw_tags = _item_value(media, "tags_json")
        if raw_tags:
            try:
                parsed = json.loads(raw_tags) if isinstance(raw_tags, str) else raw_tags
                tags = parsed.get("selected") if isinstance(parsed, dict) else parsed
            except (TypeError, ValueError, json.JSONDecodeError):
                tags = []
    if tags:
        labels = ", ".join(str(tag) for tag in list(tags)[:12])
        lines.append(f"Etiket: {escape(labels)}")
    return Paragraph("<br/>".join(lines), caption_style)


def _photo_evidence_table(media_items, caption_style):
    photos = [media for media in list(media_items or []) if _media_is_photo(media)]
    if not photos:
        return None
    cells = []
    for index, media in enumerate(photos, 1):
        single_photo = len(photos) == 1
        image = _photo_flowable(
            media,
            max_width=440 if single_photo else 220,
            max_height=120 * mm if single_photo else 88 * mm,
        )
        if image is None:
            cell = [
                Paragraph(
                    f"<b>Fotoğraf {index}</b><br/>Fotoğraf dosyası depolamadan okunamadı.",
                    caption_style,
                )
            ]
        else:
            cell = [image, Spacer(1, 2 * mm), _media_caption(media, index, caption_style)]
        cells.append(cell)
    rows = []
    for start in range(0, len(cells), 2):
        row = cells[start:start + 2]
        if len(row) == 1:
            row.append("")
        rows.append(row)
    table = Table(rows, colWidths=[230, 230], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    if len(cells) % 2 == 1:
        table.setStyle(TableStyle([("SPAN", (0, len(rows) - 1), (1, len(rows) - 1))]))
    return table


def _register_pdf_fonts() -> None:
    global PDF_FONT, PDF_FONT_BOLD
    candidates = [
        (_ASSETS / "DejaVuSans.ttf", _ASSETS / "DejaVuSans-Bold.ttf", "RiskDejaVu", "RiskDejaVu-Bold"),
        (Path(r"C:\Windows\Fonts\arial.ttf"), Path(r"C:\Windows\Fonts\arialbd.ttf"), "RiskArial", "RiskArial-Bold"),
        (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
            "RiskDejaVu",
            "RiskDejaVu-Bold",
        ),
    ]
    for regular, bold, name, bold_name in candidates:
        if not regular.exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont(name, str(regular)))
            PDF_FONT = name
            if bold.exists():
                pdfmetrics.registerFont(TTFont(bold_name, str(bold)))
                PDF_FONT_BOLD = bold_name
            else:
                PDF_FONT_BOLD = name
            try:
                registerFontFamily(name, normal=name, bold=PDF_FONT_BOLD, italic=name, boldItalic=PDF_FONT_BOLD)
            except Exception:
                pass
            return
        except Exception:
            continue


_register_pdf_fonts()


def _fmt_date(d) -> str:
    if not d:
        return "—"
    if hasattr(d, "strftime"):
        return d.strftime("%d.%m.%Y")
    s = str(d)
    if len(s) >= 10 and s[4] == "-":
        y, m, day = s[:10].split("-")
        return f"{day}.{m}.{y}"
    return s


def _dept(risk) -> str:
    return getattr(risk, "department_name", None) or "—"


def _hazard_name(risk, hazard_map: dict) -> str:
    h = hazard_map.get(getattr(risk, "hazard_id", None))
    return h.name if h else "—"


def _hazard_code(risk, hazard_map: dict) -> str:
    h = hazard_map.get(getattr(risk, "hazard_id", None))
    return h.code if h else "—"


def _risk_method(risk, fallback_code: str = "5x5_l") -> dict:
    return resolve_method(getattr(risk, "method_code", None) or fallback_code)


def _normalized_risk_level(risk, fallback_code: str = "5x5_l") -> str:
    """Return a method-aware normalized level for reports and colors."""
    method_code = getattr(risk, "method_code", None) or fallback_code
    return (
        canonical_risk_level(
            method_code,
            getattr(risk, "risk_score", None),
            getattr(risk, "risk_level", None),
        )
        or "—"
    )


def _hazop_data(risk) -> dict:
    raw = getattr(risk, "hazop_data", None)
    if hasattr(raw, "model_dump"):
        raw = raw.model_dump()
    if raw is None:
        raw = getattr(risk, "hazop_data_json", None)
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except json.JSONDecodeError:
                raw = None
    if not isinstance(raw, dict):
        return {}
    try:
        return normalize_hazop_data(raw)
    except ValueError:
        return {str(key): str(value or "") for key, value in raw.items()}


def _score_text(value) -> str:
    if value is None:
        return "—"
    numeric = float(value)
    return str(int(numeric)) if numeric.is_integer() else f"{numeric:.2f}".rstrip("0").rstrip(".")


def _level_color(level: str):
    if "Kabul" in (level or ""):
        return colors.HexColor("#95a5a6")
    if "Düşük" in (level or ""):
        return colors.HexColor("#2ecc71")
    if "Orta" in (level or ""):
        return colors.HexColor("#f1c40f")
    if "Çok" in (level or ""):
        return colors.HexColor("#e74c3c")
    if "Yüksek" in (level or ""):
        return colors.HexColor("#f39c12")
    return colors.white


def _add_pdf_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont(PDF_FONT, 7)
    canvas.setFillColor(colors.HexColor("#6c757d"))
    canvas.drawString(doc.leftMargin, 10 * mm, f"Program tasarımı ve raporlama: {CREATOR_LINE}")
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 10 * mm, f"Sayfa {doc.page}")
    canvas.restoreState()


class _RiskPdfCanvas(canvas.Canvas):
    """Her sayfada ekip imza + sayfa no; son sayfada sayfa sayısı beyanı."""

    def __init__(self, *args, team_line: str = "", **kwargs):
        self._team_line = team_line or ""
        self._saved_page_states: list[dict] = []
        super().__init__(*args, **kwargs)

    def showPage(self):
        self._saved_page_states.append(
            {k: v for k, v in self.__dict__.items() if k != "_saved_page_states"}
        )
        self._startPage()

    def save(self):
        total = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_chrome(total)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def _draw_chrome(self, total: int):
        page = self._pageNumber
        self.saveState()
        self.setFont(PDF_FONT, 6)
        self.setFillColor(colors.HexColor("#6c757d"))
        width = self._pagesize[0]
        self.drawString(14 * mm, 12 * mm, (self._team_line or "")[:115])
        self.drawRightString(width - 14 * mm, 12 * mm, f"Sayfa {page} / {total}")
        self.drawString(14 * mm, 8 * mm, CREATOR_LINE)
        if page == total and total > 0:
            self.setFont(PDF_FONT_BOLD, 8)
            self.setFillColor(colors.HexColor("#1a5276"))
            self.drawCentredString(
                width / 2,
                18 * mm,
                f"İş bu risk değerlendirme raporu {total} sayfadan oluşur.",
            )
        self.restoreState()


def build_risk_pdf(
    *,
    company,
    risks,
    hazard_map: dict | None = None,
    prepared_by: str | None = None,
    sgk_no: str | None = None,
    workplace_physician: str | None = None,
    employer_representative: str | None = None,
    employee_representative: str | None = None,
    support_staff: str | None = None,
    validity: dict | None = None,
    team_details: dict | None = None,
    employee_count: int | None = None,
    document_no: str | None = None,
    revision_no: str | None = None,
    revision_reason: str | None = None,
    scope_note: str | None = None,
    tax_number: str | None = None,
    nace_code: str | None = None,
    nace_roadmap: dict | None = None,
) -> bytes:
    """Firma risk değerlendirme PDF raporu — kapak, yöntem, ekip, kayıt, imza."""
    from app.services.risk_methods import (
        CONTROL_HIERARCHY,
        DEFINITIONS,
        LEGAL_BASIS,
        PURPOSE,
        SCOPE,
        resolve_method,
    )

    hazard_map = hazard_map or {}
    team_details = team_details or {}
    method = resolve_method((validity or {}).get("method_code") or getattr(company, "risk_method", None))
    doc_no = document_no or getattr(company, "risk_document_no", None) or f"RD-{getattr(company, 'id', 0)}"
    rev_no = revision_no or getattr(company, "risk_revision_no", None) or "00"
    rev_reason = revision_reason or getattr(company, "risk_revision_reason", None)
    scope_extra = scope_note or getattr(company, "risk_scope_note", None)
    prepared_on = datetime.now().strftime("%d.%m.%Y")

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=16 * mm,
        bottomMargin=22 * mm,
    )
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = PDF_FONT
    styles["Title"].fontName = PDF_FONT_BOLD

    title_style = ParagraphStyle(
        "RiskTitle",
        parent=styles["Title"],
        fontSize=16,
        fontName=PDF_FONT_BOLD,
        spaceAfter=4,
        textColor=colors.HexColor("#1a5276"),
        alignment=TA_CENTER,
    )
    subtitle = ParagraphStyle(
        "RiskSub",
        parent=styles["Normal"],
        fontSize=9,
        fontName=PDF_FONT,
        spaceAfter=2,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#2c3e50"),
    )
    info = ParagraphStyle("RiskInfo", parent=styles["Normal"], fontSize=9, fontName=PDF_FONT, spaceAfter=2, leading=12)
    section = ParagraphStyle(
        "RiskSection",
        parent=styles["Normal"],
        fontSize=11,
        fontName=PDF_FONT_BOLD,
        spaceBefore=4,
        spaceAfter=6,
        textColor=colors.HexColor("#1a5276"),
    )
    body = ParagraphStyle("RiskBody", parent=styles["Normal"], fontSize=8.5, fontName=PDF_FONT, leading=11, spaceAfter=3)
    cell = ParagraphStyle("RiskCell", parent=styles["Normal"], fontSize=8, fontName=PDF_FONT, leading=10)
    photo_caption = ParagraphStyle(
        "RiskPhotoCaption",
        parent=styles["Normal"],
        fontSize=7.5,
        fontName=PDF_FONT,
        leading=9,
        textColor=colors.HexColor("#475569"),
    )

    def _person_line(role_key: str, fallback_name: str | None, fallback_title: str) -> str:
        detail = team_details.get(role_key) or {}
        name = detail.get("full_name") or fallback_name or "—"
        title = detail.get("title") or fallback_title
        cert = detail.get("certificate_number") or detail.get("certificate_class")
        if cert:
            return f"{name} — {title} (Sertifika: {cert})"
        return f"{name} — {title}"

    elements: list = [
        Paragraph("RİSK DEĞERLENDİRME RAPORU", title_style),
        Paragraph(str(getattr(company, "name", "") or ""), subtitle),
        Paragraph("6331 sayılı İş Sağlığı ve Güvenliği Kanunu ve Risk Değerlendirmesi Yönetmeliği kapsamında", subtitle),
        Paragraph(
            f"Belge No: {doc_no}  |  Revizyon: {rev_no}  |  Düzenleme: {prepared_on}  |  Program: {CREATOR_LINE}",
            subtitle,
        ),
        Spacer(1, 5 * mm),
        Paragraph("1. İŞYERİ KÜNYESİ", section),
        Paragraph(f"<b>İşyeri Ünvanı:</b> {getattr(company, 'name', None) or '—'}", info),
        Paragraph(
            f"<b>SGK Sicil No:</b> {sgk_no or getattr(company, 'sgk_registry_no', None) or '—'}  |  "
            f"<b>Vergi No:</b> {tax_number or getattr(company, 'tax_number', None) or '—'}  |  "
            f"<b>NACE:</b> {nace_code or getattr(company, 'nace_code', None) or '—'}",
            info,
        ),
        Paragraph(f"<b>Adres:</b> {getattr(company, 'address', None) or '—'}", info),
        Paragraph(
            f"<b>Telefon:</b> {getattr(company, 'phone', None) or '—'}  |  "
            f"<b>Tehlike Sınıfı:</b> {getattr(company, 'hazard_class', None) or '—'}  |  "
            f"<b>Çalışan Sayısı:</b> {employee_count if employee_count is not None else '—'}",
            info,
        ),
        Paragraph(
            f"<b>İşveren / Yetkili:</b> {employer_representative or getattr(company, 'authorized_person', None) or '—'}",
            info,
        ),
        Spacer(1, 3 * mm),
        Paragraph("2. BELGE KONTROLÜ VE GEÇERLİLİK", section),
    ]

    # NACE roadmap is an additive report section.  It is intentionally read-only:
    # the report lists the evidence and work sequence but never invents a risk
    # record from the NACE code.
    if nace_roadmap:
        identity = nace_roadmap.get("identity") or {}
        status_label = nace_roadmap.get("status_label") or "—"
        elements.extend(
            [
                Paragraph("2A. NACE ODAKLI RİSK KAPSAMI VE YOL HARİTASI", section),
                Paragraph(f"<b>NACE durumu:</b> {status_label}", info),
                Paragraph(
                    f"<b>NACE kimliği:</b> {identity.get('code') or nace_roadmap.get('entered_nace_code') or '—'} — "
                    f"{identity.get('description') or 'Tam katalog açıklaması yok.'}",
                    info,
                ),
            ]
        )
        if identity.get("hazard_class"):
            elements.append(
                Paragraph(
                    f"<b>NACE bölümü / tehlike sınıfı:</b> {identity.get('section_code') or '—'} — "
                    f"{identity.get('section_name') or '—'} / {identity.get('hazard_class')}",
                    info,
                )
            )
        elements.append(
            Paragraph(
                "NACE eşleşmesi başlangıç kapsamıdır; saha gözlemi, gerçek proses ve ekip incelemesi olmadan tek başına hukuken tam risk değerlendirmesi sayılmaz.",
                body,
            )
        )
        domains = list(nace_roadmap.get("technical_risk_tags") or []) + list(nace_roadmap.get("special_risks") or [])
        elements.append(Paragraph("<b>NACE teknik/özel risk başlıkları:</b>", info))
        if domains:
            for domain in domains:
                elements.append(
                    Paragraph(
                        f"• {domain.get('label') or domain.get('key') or '—'} — {domain.get('description') or 'Saha doğrulaması gerekir.'}",
                        body,
                    )
                )
        else:
            elements.append(Paragraph("• Teknik eşleştirme mevcut değil; genel kontrol listesi uygulanır ve uzman incelemesi beklenir.", body))
        elements.append(Paragraph("<b>Bu raporda bulunması gereken asgari başlıklar:</b>", info))
        for item in nace_roadmap.get("report_checklist") or []:
            elements.append(
                Paragraph(
                    f"• <b>{item.get('title') or item.get('key') or '—'}</b>: {item.get('description') or '—'} "
                    f"({item.get('legal_basis') or 'mevzuat / ekip doğrulaması'})",
                    body,
                )
            )
        elements.append(Paragraph("<b>Uygulama yol haritası:</b>", info))
        for idx, step in enumerate(nace_roadmap.get("roadmap") or [], 1):
            elements.append(
                Paragraph(
                    f"{idx}. <b>{step.get('title') or step.get('key') or '—'}</b>: {step.get('description') or '—'}",
                    body,
                )
            )
        for warning in nace_roadmap.get("warnings") or []:
            elements.append(Paragraph(f"<i>Uyarı: {warning}</i>", body))

    for label, value in document_meta_rows(
        validity=validity,
        prepared_by=prepared_by,
        workplace_physician=workplace_physician,
        employer_representative=employer_representative,
        employee_representative=employee_representative,
        support_staff=support_staff,
        document_no=doc_no,
        revision_no=rev_no,
        revision_reason=rev_reason,
    ):
        elements.append(Paragraph(f"<b>{label}:</b> {value}", info))
    if validity and validity.get("message"):
        elements.append(Paragraph(f"<i>{validity['message']}</i>", body))
    triggers = (validity or {}).get("renewal_triggers") or []
    if triggers:
        elements.append(Paragraph("<b>Süre beklenmeden yenileme tetikleyicileri (md.12/2):</b>", info))
        for t in triggers:
            elements.append(Paragraph(f"• {t}", body))

    elements.extend(
        [
            Spacer(1, 3 * mm),
            Paragraph("3. AMAÇ, KAPSAM VE MEVZUAT DAYANAĞI", section),
            Paragraph(f"<b>Amaç:</b> {PURPOSE}", body),
            Paragraph(f"<b>Kapsam:</b> {scope_extra or SCOPE}", body),
            Paragraph("<b>Yasal dayanak:</b>", info),
        ]
    )
    for law in LEGAL_BASIS:
        elements.append(Paragraph(f"• {law}", body))

    elements.extend(
        [
            Spacer(1, 3 * mm),
            Paragraph("4. TANIMLAR VE KISALTMALAR", section),
        ]
    )
    for term, meaning in DEFINITIONS:
        elements.append(Paragraph(f"<b>{term}:</b> {meaning}", body))

    elements.extend(
        [
            Spacer(1, 3 * mm),
            Paragraph("5. YÖNTEM VE SKORLAMA KRİTERLERİ", section),
            Paragraph(f"<b>Seçilen yöntem:</b> {method['label']}", info),
            Paragraph(f"<b>Formül:</b> {method['formula']}", info),
            Paragraph(method["narrative"], body),
            Paragraph("<b>Olasılık / maruziyet tanımları:</b>", info),
        ]
    )
    for val, txt in method.get("probability_defs") or []:
        elements.append(Paragraph(f"• {val}: {txt}", body))
    if not method.get("probability_defs"):
        elements.append(Paragraph("• Nitel değerlendirme (yönteme özgü skorlama ekip tutanağında).", body))
    elements.append(Paragraph("<b>Şiddet tanımları:</b>", info))
    for val, txt in method.get("severity_defs") or []:
        elements.append(Paragraph(f"• {val}: {txt}", body))
    if not method.get("severity_defs"):
        elements.append(Paragraph("• Nitel değerlendirme (yönteme özgü).", body))
    if method.get("frequency_defs"):
        elements.append(Paragraph("<b>Frekans / maruziyet tanımları:</b>", info))
        for val, txt in method.get("frequency_defs") or []:
            elements.append(Paragraph(f"• {val}: {txt}", body))
    elements.append(Paragraph("<b>Risk seviyesi / öncelik:</b>", info))
    for rng, level, note in method.get("levels") or []:
        elements.append(Paragraph(f"• {rng} → <b>{level}</b>: {note}", body))

    elements.extend(
        [
            Spacer(1, 3 * mm),
            Paragraph("6. ÖNLEM HİYERARŞİSİ VE DÜZELTİCİ FAALİYET İLKELERİ", section),
        ]
    )
    for line in CONTROL_HIERARCHY:
        elements.append(Paragraph(f"• {line}", body))
    elements.append(
        Paragraph(
            "İlave önlemler sonrası artık risk yeniden değerlendirilir; DÖF ile sorumlu, termin ve durum izlenir. "
            "Onaylı/imzalı raporlar tarihsel belge olarak korunur; işyeri kartındaki sonraki değişiklikler "
            "geçmiş raporları otomatik değiştirmez.",
            body,
        )
    )

    elements.append(PageBreak())
    elements.append(Paragraph("7. RİSK DEĞERLENDİRME EKİBİ (md.15)", section))
    elements.append(
        Paragraph(
            "Aşağıdaki kişiler yönetmelik gereği değerlendirmeye katılır / onaylar. "
            "Görevlendirme kayıtlarından gelen unvan ve sertifika bilgileri otomatik doldurulur.",
            body,
        )
    )
    team_rows = [
        ["Rol / Görev", "Ad Soyad / Unvan / Sertifika", "Sorumluluk", "İmza"],
        [
            "İşveren / Vekil",
            Paragraph(_person_line("employer", employer_representative, "İşveren / Vekil"), cell),
            "Onay / yetkilendirme",
            "",
        ],
        [
            "İSG Uzmanı",
            Paragraph(
                _person_line("safety_specialist", prepared_by, "İş Güvenliği Uzmanı"),
                cell,
            ),
            "Hazırlama / koordinasyon",
            "",
        ],
        [
            "İşyeri Hekimi",
            Paragraph(
                _person_line("workplace_physician", workplace_physician, "İşyeri Hekimi"),
                cell,
            ),
            "Sağlık boyutu / inceleme",
            "",
        ],
        [
            "Çalışan Temsilcisi",
            Paragraph(employee_representative or "—", cell),
            "Katılım / görüş",
            "",
        ],
        [
            "Destek Elemanı",
            Paragraph(support_staff or "—", cell),
            "Destek / uygulama",
            "",
        ],
        [
            "Diğer Sağlık Personeli",
            Paragraph(_person_line("other_health_personnel", None, "DSP"), cell),
            "Gerektiğinde katılım",
            "",
        ],
    ]
    team_table = Table(team_rows, colWidths=[85, 200, 100, 70])
    team_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a5276")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 1), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 10),
            ]
        )
    )
    elements.append(team_table)

    # Summary
    total = len(risks)
    risk_levels: dict[str, int] = {}
    for r in risks:
        rl = _normalized_risk_level(r)
        risk_levels[rl] = risk_levels.get(rl, 0) + 1

    elements.append(Spacer(1, 5 * mm))
    elements.append(Paragraph("8. RİSK ÖZETİ", section))
    summary_data = [["Risk Seviyesi", "Adet", "Yüzde"]]
    for level in ["Çok Yüksek", "Yüksek", "Orta", "Düşük", "Kabul Edilebilir"]:
        count = risk_levels.get(level, 0)
        pct = f"{(count / total * 100):.1f}%" if total else "%0"
        summary_data.append([level, str(count), pct])
    summary_table = Table(summary_data, colWidths=[120, 50, 50])
    summary_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a5276")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 0), (2, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 1), (0, 1), colors.HexColor("#e74c3c")),
                ("BACKGROUND", (0, 2), (0, 2), colors.HexColor("#f39c12")),
                ("BACKGROUND", (0, 3), (0, 3), colors.HexColor("#f1c40f")),
                ("BACKGROUND", (0, 4), (0, 4), colors.HexColor("#2ecc71")),
                ("BACKGROUND", (0, 5), (0, 5), colors.HexColor("#95a5a6")),
            ]
        )
    )
    elements.append(summary_table)

    elements.append(PageBreak())
    elements.append(Paragraph("9. RİSK KAYIT LİSTESİ", section))
    elements.append(
        Paragraph(
            f"Belge yöntemi: {method.get('label')}. Kayıtlar yöntem koduna göre değerlendirilir; "
            "tarihsel 5×5 kayıtları Fine-Kinney skoruna dönüştürülmez ve mevcut değerleri korunur.",
            body,
        )
    )
    elements.append(Spacer(1, 2 * mm))

    for risk in risks:
        risk_method = _risk_method(risk, method.get("code", "5x5_l"))
        risk_level_label = _normalized_risk_level(risk, method.get("code", "5x5_l"))
        if risk_method.get("code") == "fine_kinney":
            _, risk_level_label, _ = fine_kinney_level_details(float(risk.risk_score or 0))
        elif risk_method.get("code") == "hazop":
            risk_level_label = priority_details(_hazop_data(risk).get("priority"))["label"]
        regs = []
        h = hazard_map.get(getattr(risk, "hazard_id", None))
        if h and getattr(h, "regulations", None):
            try:
                import json

                raw = h.regulations
                regs = json.loads(raw) if isinstance(raw, str) else (raw or [])
            except Exception:
                regs = []
        risk_data = [
            ["Risk Kodu", risk.risk_code],
            ["Yöntem", risk_method.get("label") or "—"],
            ["Bölüm", _dept(risk)],
            ["Faaliyet", Paragraph(str(risk.activity or "—"), cell)],
            ["Tehlike", f"{_hazard_code(risk, hazard_map)} — {_hazard_name(risk, hazard_map)}"],
            ["Risk Tanımı", Paragraph(str(risk.risk_definition or "—"), cell)],
            ["Etkilenenler", risk.affected_people or getattr(risk, "affected_group", None) or "—"],
        ]
        if risk_method.get("code") == "hazop":
            hazop = _hazop_data(risk)
            priority = priority_details(hazop.get("priority"))
            risk_data.extend(
                [
                    ["Proses düğümü", Paragraph(hazop.get("node") or "—", cell)],
                    ["Tasarım amacı", Paragraph(hazop.get("design_intent") or "—", cell)],
                    ["Parametre", hazop.get("parameter") or "—"],
                    ["Kılavuz kelime", guide_word_label(hazop.get("guide_word"))],
                    ["Sapma", Paragraph(hazop.get("deviation") or "—", cell)],
                    ["Nedenler", Paragraph(hazop.get("causes") or "—", cell)],
                    ["Sonuçlar", Paragraph(hazop.get("consequences") or "—", cell)],
                    ["Mevcut korumalar", Paragraph(hazop.get("safeguards") or "—", cell)],
                    ["Öneriler", Paragraph(hazop.get("recommendations") or "—", cell)],
                    ["HAZOP önceliği", priority["label"]],
                    ["Yöntem aksiyonu", priority["action"]],
                ]
            )
        else:
            risk_data.append([risk_method.get("probability_axis") or "Olasılık", _score_text(risk.probability)])
            if risk_method.get("code") == "fine_kinney":
                risk_data.append([risk_method.get("frequency_axis") or "Frekans / Maruziyet", _score_text(getattr(risk, "frequency", None))])
            risk_data.extend(
                [
                    [risk_method.get("severity_axis") or "Şiddet", _score_text(risk.severity)],
                    ["Risk Skoru", _score_text(risk.risk_score)],
                    ["Risk Seviyesi", risk_level_label],
                    ["Yöntem aksiyonu", next((note for _, label, note in risk_method.get("levels", []) if label in {risk_level_label, getattr(risk, "risk_level", None)}), "—")],
                ]
            )
        risk_data.extend(
            [
            ["Termin Tarihi", _fmt_date(risk.term_date)],
            ["Durum", risk.status or "Açık"],
            ["DÖF sayısı", str(len(getattr(risk, "dofs", None) or []))],
            ["Revizyon (kayıt)", str(getattr(risk, "revision_no", None) or "—")],
            ["Mevzuat ref.", Paragraph(", ".join(regs) if regs else "—", cell)],
            ]
        )
        if risk_method.get("code") != "hazop" and getattr(risk, "residual_score", None) is not None:
            residual_label = getattr(risk, "residual_level", None) or "—"
            if risk_method.get("code") == "fine_kinney":
                _, residual_label, _ = fine_kinney_level_details(float(risk.residual_score or 0))
            risk_data.append(["Artık risk skoru / seviyesi", f"{_score_text(risk.residual_score)} / {residual_label}"])
        risk_table = Table(risk_data, colWidths=[100, 370])
        bg = _level_color(_normalized_risk_level(risk, risk_method.get("code", "5x5_l")))
        score_label = "HAZOP önceliği" if risk_method.get("code") == "hazop" else "Risk Skoru"
        score_row = next((index for index, item in enumerate(risk_data) if item[0] == score_label), 0)
        risk_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTNAME", (0, 0), (0, -1), PDF_FONT_BOLD),
                    ("BACKGROUND", (1, score_row), (1, score_row + 1), bg),
                ]
            )
        )
        elements.append(risk_table)
        elements.append(Spacer(1, 2 * mm))

        residual_text = (
            "İlave önlem / DÖF tamamlandıktan sonra yeniden değerlendirilir "
            "(önlem sonrası skor sahada güncellenir)."
        )
        if getattr(risk, "residual_score", None) is not None:
            residual_text = (
                f"Kayıtlı artık risk: {_score_text(risk.residual_score)} / "
                f"{getattr(risk, 'residual_level', None) or '—'}. "
                "Kontrollerin sahada etkinliği ayrıca doğrulanır."
            )
        measures = [
            ["Mevcut Önlemler", Paragraph(str(risk.existing_measures or "—"), cell)],
            ["İlave Önlemler", Paragraph(str(risk.additional_measures or "—"), cell)],
            [
                "Artık risk",
                Paragraph(
                    residual_text,
                    cell,
                ),
            ],
        ]
        measures_table = Table(measures, colWidths=[100, 370])
        measures_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(measures_table)

        dofs = list(getattr(risk, "dofs", None) or [])
        if dofs:
            elements.append(Spacer(1, 2 * mm))
            dof_rows = [["DÖF No", "Yapılacak İş", "Sorumlu", "Termin", "Durum"]]
            for d in dofs:
                dof_rows.append(
                    [
                        d.dof_code,
                        Paragraph(str(d.description or "—")[:200], cell),
                        d.responsible_person or "—",
                        _fmt_date(d.term_date),
                        d.status or "Açık",
                    ]
                )
            dof_table = Table(dof_rows, colWidths=[55, 200, 80, 60, 55])
            dof_table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                        ("FONTSIZE", (0, 0), (-1, -1), 7),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            elements.append(dof_table)

        photo_table = _photo_evidence_table(getattr(risk, "media_files", None), photo_caption)
        if photo_table:
            elements.append(
                KeepTogether(
                    [Spacer(1, 3 * mm), Paragraph("FOTOĞRAF KANITLARI", section), photo_table]
                )
            )

        elements.append(Spacer(1, 6 * mm))

    elements.append(PageBreak())
    elements.append(Paragraph("10. İMZA / ONAY", section))
    elements.append(
        Paragraph(
            "Bu belgeyi hazırlayan, inceleyen, katılan ve onaylayan kişiler aşağıda imza altına alır. "
            "İşveren onayı olmadan risk değerlendirmesi resmi sayılmaz.",
            body,
        )
    )
    elements.append(Spacer(1, 3 * mm))
    # Yatay imza: satır1 unvanlar · satır2 adlar · satır3 imza alanı
    sign_people = _risk_sign_people(
        prepared_by=prepared_by,
        workplace_physician=workplace_physician,
        employer_representative=employer_representative,
        employee_representative=employee_representative,
        support_staff=support_staff,
    )
    dsp = (team_details.get("other_health_personnel") or {}).get("full_name")
    if dsp:
        sign_people.append(("Diğer Sağlık\nPersoneli", dsp))
    titles = [p[0].replace("\n", " ") for p in sign_people]
    names = [p[1] for p in sign_people]
    stamps = ["Kaşe / İmza / Tarih"] * len(sign_people)
    col_w = max(70, int(470 / max(1, len(sign_people))))
    sign_table = Table([titles, names, stamps], colWidths=[col_w] * len(sign_people))
    sign_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a5276")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 2), (-1, 2), 16),
                ("BOTTOMPADDING", (0, 2), (-1, 2), 16),
            ]
        )
    )
    elements.append(sign_table)

    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph("11. DAĞITIM, SAKLAMA VE ARŞİV", section))
    elements.append(
        Paragraph(
            "Dağıtım: İşveren / vekil, İSG uzmanı, işyeri hekimi, çalışan temsilcisi, ilgili birim sorumluları. "
            "Saklama: İşyerinde erişilebilir; denetim ve müfettiş talebinde sunulur. "
            "Arşiv: Onaylı PDF/Excel kopyaları tarihsel belge olarak saklanır; sonraki işyeri kartı "
            "değişiklikleri onaylı sürümü otomatik değiştirmez.",
            body,
        )
    )
    elements.append(
        Paragraph(
            f"Belge kontrol: {doc_no} / Rev {rev_no} / {prepared_on} / {CREATOR_LINE}",
            body,
        )
    )

    team_line = _risk_team_footer_line(
        prepared_by=prepared_by,
        workplace_physician=workplace_physician,
        employer_representative=employer_representative,
        employee_representative=employee_representative,
        support_staff=support_staff,
    )
    doc.build(
        elements,
        canvasmaker=lambda *args, **kwargs: _RiskPdfCanvas(*args, team_line=team_line, **kwargs),
    )
    buf.seek(0)
    return buf.read()


def build_field_inspection_pdf(
    *,
    company,
    prepared_by: str | None = None,
    department_name: str | None = None,
    location: str | None = None,
    category_name: str | None = None,
    hazard_code: str | None = None,
    hazard_name: str | None = None,
    summary: str | None = None,
    existing_measures: str | None = None,
    action: str | None = None,
    responsible_person: str | None = None,
    term_date: str | None = None,
    probability: float | None = None,
    severity: float | None = None,
    observed_at=None,
    gps_lat: float | None = None,
    gps_lng: float | None = None,
    gps_accuracy_m: float | None = None,
    photos: list[dict] | None = None,
    vision_results: list[dict] | None = None,
) -> bytes:
    """Kayıt öncesi saha formu için fotoğraflı, geçici PDF raporu."""
    from html import escape

    photos = list(photos or [])
    vision_results = [item for item in list(vision_results or []) if isinstance(item, dict)]
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = PDF_FONT
    title = ParagraphStyle(
        "FieldReportTitle",
        parent=styles["Title"],
        fontSize=17,
        leading=21,
        fontName=PDF_FONT_BOLD,
        textColor=colors.HexColor("#0f766e"),
        alignment=TA_CENTER,
        spaceAfter=4,
    )
    subtitle = ParagraphStyle(
        "FieldReportSubtitle",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        fontName=PDF_FONT,
        textColor=colors.HexColor("#475569"),
        alignment=TA_CENTER,
        spaceAfter=3,
    )
    section = ParagraphStyle(
        "FieldReportSection",
        parent=styles["Normal"],
        fontSize=11,
        leading=14,
        fontName=PDF_FONT_BOLD,
        textColor=colors.HexColor("#0f766e"),
        spaceBefore=6,
        spaceAfter=6,
    )
    body = ParagraphStyle(
        "FieldReportBody",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        fontName=PDF_FONT,
        spaceAfter=4,
    )
    small = ParagraphStyle(
        "FieldReportSmall",
        parent=styles["Normal"],
        fontSize=7.5,
        leading=9,
        fontName=PDF_FONT,
        textColor=colors.HexColor("#475569"),
    )
    label = ParagraphStyle(
        "FieldReportLabel",
        parent=small,
        fontName=PDF_FONT_BOLD,
        textColor=colors.HexColor("#0f172a"),
    )

    def safe(value, fallback="—", limit=2200):
        text = str(value or "").strip()
        return escape(text[:limit] if text else fallback)

    gps = "—"
    if gps_lat is not None and gps_lng is not None:
        accuracy = f" · ±{gps_accuracy_m} m" if gps_accuracy_m is not None else ""
        gps = f"{gps_lat}, {gps_lng}{accuracy}"
    score = "—"
    if probability is not None and severity is not None:
        score = f"{_score_text(probability)} × {_score_text(severity)} = {_score_text(float(probability) * float(severity))}"

    info_rows = [
        [Paragraph("İşyeri", label), Paragraph(safe(getattr(company, "name", None)), body)],
        [Paragraph("Bölüm / saha alanı", label), Paragraph(safe(department_name), body)],
        [Paragraph("Gözlem konumu", label), Paragraph(safe(location), body)],
        [Paragraph("Gözlem tarihi", label), Paragraph(safe(_fmt_datetime(observed_at)), body)],
        [Paragraph("Hazırlayan", label), Paragraph(safe(prepared_by), body)],
        [Paragraph("GPS", label), Paragraph(safe(gps), body)],
        [Paragraph("Risk puanı", label), Paragraph(safe(score), body)],
        [Paragraph("Kategori / tehlike", label), Paragraph(safe(" · ".join(part for part in [category_name, hazard_code, hazard_name] if part)), body)],
    ]
    info_table = Table(info_rows, colWidths=[125, 335], hAlign="LEFT")
    info_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#cbd5e1")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5f9")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    elements: list = [
        Paragraph("FOTOĞRAFLI SAHA DENETİM RAPORU", title),
        Paragraph(safe(getattr(company, "name", None)), subtitle),
        Paragraph(
            f"Düzenleme: {escape(_fmt_datetime(datetime.now()))} · {CREATOR_LINE}",
            subtitle,
        ),
        Spacer(1, 4 * mm),
        info_table,
        Spacer(1, 3 * mm),
        Paragraph("1. SAHA BULGUSU", section),
        Paragraph(f"<b>Uygunsuzluk / risk tanımı:</b> {safe(summary)}", body),
        Paragraph(f"<b>Mevcut önlemler:</b> {safe(existing_measures)}", body),
        Paragraph("2. DÖF / AKSİYON", section),
        Paragraph(f"<b>Aksiyon:</b> {safe(action)}", body),
        Paragraph(f"<b>Sorumlu kişi:</b> {safe(responsible_person)}", body),
        Paragraph(f"<b>Termin tarihi:</b> {safe(term_date)}", body),
    ]

    photo_caption = ParagraphStyle(
        "FieldReportPhotoCaption",
        parent=small,
        fontSize=7.5,
        leading=9,
    )
    photo_table = _photo_evidence_table(photos, photo_caption)
    elements.append(
        KeepTogether(
            [
                Paragraph("3. FOTOĞRAF KANITLARI", section),
                photo_table or Paragraph("Fotoğraf kanıtı eklenmedi.", body),
            ]
        )
    )

    if vision_results:
        elements.append(Paragraph("4. AI FOTOĞRAF ANALİZİ", section))
        elements.append(
            Paragraph(
                "AI çıktısı fotoğraf üzerinden ön değerlendirmedir; nihai risk kararı yetkili İSG profesyoneli tarafından kontrol edilmelidir.",
                body,
            )
        )
        for index, analysis in enumerate(vision_results, 1):
            summary_text = safe(analysis.get("summary"), "Analiz özeti yok.", 1200)
            elements.append(Paragraph(f"<b>Fotoğraf {index}:</b> {summary_text}", body))
            hazards = analysis.get("hazards") or []
            if not isinstance(hazards, list) or not hazards:
                continue
            hazard_rows = [["Tehlike", "Fotoğrafta görülen durum", "Şiddet", "Güven"]]
            for hazard in hazards[:12]:
                if not isinstance(hazard, dict):
                    continue
                hazard_rows.append(
                    [
                        Paragraph(safe(hazard.get("category")), small),
                        Paragraph(safe(hazard.get("observed") or hazard.get("note")), small),
                        safe(hazard.get("severity")),
                        f"%{_confidence_percent(hazard.get('confidence'))}",
                    ]
                )
            if len(hazard_rows) > 1:
                hazard_table = Table(hazard_rows, colWidths=[105, 245, 50, 60], hAlign="LEFT")
                hazard_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 5),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                            ("TOPPADDING", (0, 0), (-1, -1), 4),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ]
                    )
                )
                elements.append(KeepTogether([hazard_table, Spacer(1, 2 * mm)]))

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
    )
    doc.build(elements, onFirstPage=_add_pdf_footer, onLaterPages=_add_pdf_footer)
    buf.seek(0)
    return buf.read()


def _risk_sign_people(
    *,
    prepared_by: str | None = None,
    workplace_physician: str | None = None,
    employer_representative: str | None = None,
    employee_representative: str | None = None,
    support_staff: str | None = None,
) -> list[tuple[str, str]]:
    """(kısa unvan, ad) listesi — yatay imza bloğu için."""
    people = [
        ("İSG Uzmanı\n/ Hazırlayan", (prepared_by or "—").strip() or "—"),
        ("İşyeri\nHekimi", (workplace_physician or "—").strip() or "—"),
        ("İşveren\n/ Vekili", (employer_representative or "—").strip() or "—"),
        ("Çalışan\nTemsilcisi", (employee_representative or "—").strip() or "—"),
    ]
    if (support_staff or "").strip():
        people.append(("Destek\nElemanı", support_staff.strip()))
    return people


def _risk_team_footer_line(
    *,
    prepared_by: str | None = None,
    workplace_physician: str | None = None,
    employer_representative: str | None = None,
    employee_representative: str | None = None,
    support_staff: str | None = None,
) -> str:
    """Yazdırma alt bilgisi — her sayfada yatay ekip imza satırı."""
    parts = [
        f"İGU: {(prepared_by or '—').strip() or '—'}",
        f"İH: {(workplace_physician or '—').strip() or '—'}",
        f"İşv: {(employer_representative or '—').strip() or '—'}",
        f"ÇT: {(employee_representative or '—').strip() or '—'}",
    ]
    if (support_staff or "").strip():
        parts.append(f"Destek: {support_staff.strip()}")
    # Excel footer'da \\n satır kırar; yatay tek satır + ikinci satır imza alanı
    return "İMZA/ONAY (yatay)  " + "  ·  ".join(parts) + "\nKaşe/İmza/Tarih: ________  ________  ________  ________"


def _apply_excel_page_chrome(ws, *, team_line: str, doc_label: str = "") -> None:
    """Her yazdırılan sayfada sayfa no + yatay ekip imza alt bilgisi."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 1.05
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.65
    ws.oddHeader.center.text = doc_label[:120] if doc_label else ""
    ws.oddHeader.center.font = "Calibri"
    ws.oddHeader.center.size = 8
    # Alt bilgi: yatay ekip + imza boşlukları · sayfa no
    ws.oddFooter.left.text = (team_line or "")[:255]
    ws.oddFooter.left.font = "Calibri"
    ws.oddFooter.left.size = 6
    ws.oddFooter.center.text = "Sayfa &P / &N"
    ws.oddFooter.center.font = "Calibri"
    ws.oddFooter.center.size = 8
    ws.oddFooter.right.text = CREATOR_LINE
    ws.oddFooter.right.font = "Calibri"
    ws.oddFooter.right.size = 7
    ws.evenFooter.left.text = (team_line or "")[:255]
    ws.evenFooter.center.text = "Sayfa &P / &N"
    ws.evenFooter.right.text = CREATOR_LINE
    ws.print_options.horizontalCentered = True


def _write_excel_horizontal_sign_block(
    ws,
    *,
    start_row: int,
    people: list[tuple[str, str]],
    header_font,
    header_fill,
    thin,
    merge_cols: int = 17,
) -> int:
    """Yatay imza/onay: her kişi bir sütun (Unvan / Ad / İmza kutusu). Sonraki boş satırı döner."""
    n = max(1, len(people))
    # Başlık satırı
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=merge_cols)
    ws.cell(row=start_row, column=1, value="İMZA / ONAY — RİSK DEĞERLENDİRME EKİBİ (yatay)")
    ws.cell(row=start_row, column=1).font = Font(name="Calibri", bold=True, size=11, color="1a5276")

    note_r = start_row + 1
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=merge_cols)
    ws.cell(
        row=note_r,
        column=1,
        value=(
            "Hazırlayan / inceleyen / onaylayan kişiler aşağıda imza altına alır. "
            "İşveren onayı olmadan risk değerlendirmesi resmi sayılmaz. "
            "Aynı yatay imza düzeni her yazdırılan sayfanın alt bilgisinde de yer alır."
        ),
    )
    ws.cell(row=note_r, column=1).font = Font(name="Calibri", size=8, color="475569")
    ws.cell(row=note_r, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_r].height = 28

    # Sütun genişliği: kişi başına ~3 sütun birleştir (A-C, D-F, ...)
    span = max(2, merge_cols // n)
    title_row = note_r + 2
    name_row = title_row + 1
    sign_row = title_row + 2

    for i, (title, name) in enumerate(people):
        c0 = 1 + i * span
        c1 = min(c0 + span - 1, merge_cols)
        if c0 > merge_cols:
            break
        if c1 > c0:
            ws.merge_cells(start_row=title_row, start_column=c0, end_row=title_row, end_column=c1)
            ws.merge_cells(start_row=name_row, start_column=c0, end_row=name_row, end_column=c1)
            ws.merge_cells(start_row=sign_row, start_column=c0, end_row=sign_row, end_column=c1)

        t_cell = ws.cell(row=title_row, column=c0, value=title.replace("\n", " "))
        t_cell.font = header_font
        t_cell.fill = header_fill
        t_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        t_cell.border = thin
        for c in range(c0, c1 + 1):
            ws.cell(row=title_row, column=c).border = thin
            ws.cell(row=title_row, column=c).fill = header_fill

        n_cell = ws.cell(row=name_row, column=c0, value=name)
        n_cell.font = Font(name="Calibri", size=9, bold=True)
        n_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in range(c0, c1 + 1):
            ws.cell(row=name_row, column=c).border = thin

        s_cell = ws.cell(row=sign_row, column=c0, value="Kaşe / İmza / Tarih\n\n")
        s_cell.font = Font(name="Calibri", size=8, color="6c757d")
        s_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        for c in range(c0, c1 + 1):
            ws.cell(row=sign_row, column=c).border = thin

    ws.row_dimensions[title_row].height = 28
    ws.row_dimensions[name_row].height = 22
    ws.row_dimensions[sign_row].height = 48
    return sign_row + 2


def _estimate_risk_sheet_pages(risks: list) -> int:
    """Yazdırma önizlemesine yakın sayfa tahmini (imza sayfası dahil)."""
    import math

    n = len(risks)
    if n <= 0:
        return 1
    photo_bonus = 0
    for r in risks:
        if _risk_excel_photo_path(r):
            photo_bonus += 1
    # Yatay A4, fit-width: başlık+filtre satırları sonrası ~18 satır/sayfa; foto satırı ~2
    units = n + photo_bonus
    data_pages = max(1, math.ceil(units / 18))
    return data_pages + 1  # imza / onay sayfası


def build_risk_excel(
    *,
    company,
    risks,
    hazard_map: dict | None = None,
    validity: dict | None = None,
    prepared_by: str | None = None,
    workplace_physician: str | None = None,
    employer_representative: str | None = None,
    employee_representative: str | None = None,
    support_staff: str | None = None,
    nace_roadmap: dict | None = None,
) -> bytes:
    """Excel: Risk tablosu + DÖF listesi + istatistikler (+ sayfa imza / sayfa no)."""
    hazard_map = hazard_map or {}
    wb = openpyxl.Workbook()
    wb.properties.creator = CREATOR_LINE
    wb.properties.title = "İSG Risk Değerlendirme Raporu"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    level_fills = {
        "Kabul Edilebilir": PatternFill(start_color="95a5a6", end_color="95a5a6", fill_type="solid"),
        "Düşük": PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid"),
        "Orta": PatternFill(start_color="f1c40f", end_color="f1c40f", fill_type="solid"),
        "Yüksek": PatternFill(start_color="f39c12", end_color="f39c12", fill_type="solid"),
        "Çok Yüksek": PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid"),
    }

    team_line = _risk_team_footer_line(
        prepared_by=prepared_by,
        workplace_physician=workplace_physician,
        employer_representative=employer_representative,
        employee_representative=employee_representative,
        support_staff=support_staff,
    )
    page_count = _estimate_risk_sheet_pages(risks)
    doc_label = f"Risk Değerlendirme Raporu — {getattr(company, 'name', '')}"
    has_hazop = any(getattr(risk, "method_code", None) == "hazop" for risk in risks)
    has_fine_kinney = any(getattr(risk, "method_code", None) == "fine_kinney" for risk in risks)
    column_count = 21 if has_hazop else (19 if has_fine_kinney else 17)
    last_column = get_column_letter(column_count)

    ws = wb.active
    ws.title = "Risk Değerlendirme"
    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = f"RİSK DEĞERLENDİRME RAPORU - {company.name}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1a5276")
    ws["A1"].alignment = Alignment(horizontal="center")

    doc_no = getattr(company, "risk_document_no", None) or f"RD-{getattr(company, 'id', '')}"
    rev_no = getattr(company, "risk_revision_no", None) or "00"
    ws.merge_cells(f"A2:{last_column}2")
    ws["A2"] = (
        f"Yetkili: {getattr(company, 'authorized_person', None) or '—'} | "
        f"Tel: {getattr(company, 'phone', None) or '—'} | "
        f"Tehlike Sınıfı: {getattr(company, 'hazard_class', None) or '—'} | "
        f"SGK: {getattr(company, 'sgk_registry_no', None) or '—'} | "
        f"NACE: {getattr(company, 'nace_code', None) or '—'} | "
        f"Belge: {doc_no} / Rev {rev_no} | "
        f"Tarih: {datetime.now().strftime('%d.%m.%Y')}"
    )
    ws["A2"].font = Font(size=9, color="2c3e50")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A3:{last_column}3")
    method = (validity or {}).get("method") or METHOD_LABEL
    valid_line = ""
    if validity and validity.get("valid_until"):
        valid_line = f" · Geçerlilik: {_fmt_date(validity['valid_until'])}"
    assess_line = ""
    if validity and validity.get("assessment_date"):
        assess_line = f" · Değerlendirme: {_fmt_date(validity['assessment_date'])}"
    ws["A3"] = (
        f"Yöntem: {method}{assess_line}{valid_line} · "
        f"Program: {CREATOR_LINE} · 6331 sayılı Kanun / Risk Değerlendirmesi Yönetmeliği"
    )
    ws["A3"].font = Font(size=9, italic=True, color="6c757d")
    ws["A3"].alignment = Alignment(horizontal="center")

    headers = (
        [
            "Risk Kodu", "Yöntem", "Bölüm", "Faaliyet", "Proses Düğümü", "Tasarım Amacı",
            "Parametre", "Kılavuz Kelime", "Sapma", "Tehlike", "Tehlike Kodu", "Nedenler",
            "Sonuçlar", "Mevcut Korumalar", "Öneriler", "Etkilenenler", "HAZOP Önceliği",
            "Termin Tarihi", "Durum", "DÖF Sayısı", "Fotoğraf",
        ]
        if has_hazop
        else [
            "Risk Kodu", "Yöntem", "Bölüm", "Faaliyet", "Tehlike", "Tehlike Kodu",
            "Risk Tanımı", "Etkilenenler", "Olasılık", "Frekans / Maruziyet", "Şiddet",
            "Risk Skoru", "Risk Seviyesi", "Termin Tarihi", "Mevcut Önlemler",
            "İlave Önlemler", "Durum", "DÖF Sayısı", "Fotoğraf",
        ]
        if has_fine_kinney
        else [
            "Risk Kodu", "Bölüm", "Faaliyet", "Tehlike", "Tehlike Kodu", "Risk Tanımı",
            "Etkilenenler", "Olasılık (1-5)", "Şiddet (1-5)", "Risk Skoru", "Risk Seviyesi",
            "Termin Tarihi", "Mevcut Önlemler", "İlave Önlemler", "Durum", "DÖF Sayısı", "Fotoğraf",
        ]
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin

    for idx, risk in enumerate(risks, 5):
        dofs = list(getattr(risk, "dofs", None) or [])
        risk_method = _risk_method(risk, (validity or {}).get("method_code") or "5x5_l")
        if risk_method.get("code") == "fine_kinney":
            _, risk_level_label, _ = fine_kinney_level_details(float(risk.risk_score or 0))
        elif risk_method.get("code") == "hazop":
            risk_level_label = priority_details(_hazop_data(risk).get("priority"))["label"]
        else:
            risk_level_label = _normalized_risk_level(risk, (validity or {}).get("method_code") or "5x5_l")
        hazop = _hazop_data(risk) if risk_method.get("code") == "hazop" else {}
        data = (
            [
                risk.risk_code, risk_method.get("label") or "—", _dept(risk), risk.activity,
                hazop.get("node") or "—", hazop.get("design_intent") or "—", hazop.get("parameter") or "—",
                guide_word_label(hazop.get("guide_word")), hazop.get("deviation") or "—",
                _hazard_name(risk, hazard_map), _hazard_code(risk, hazard_map), hazop.get("causes") or "—",
                hazop.get("consequences") or "—", hazop.get("safeguards") or "—",
                hazop.get("recommendations") or "—", risk.affected_people or "—", risk_level_label,
                _fmt_date(risk.term_date), risk.status or "Açık", len(dofs), "",
            ]
            if has_hazop
            else [
                risk.risk_code, risk_method.get("label") or "—", _dept(risk), risk.activity,
                _hazard_name(risk, hazard_map), _hazard_code(risk, hazard_map), risk.risk_definition,
                risk.affected_people or "—", _score_text(risk.probability), _score_text(getattr(risk, "frequency", None)),
                _score_text(risk.severity), _score_text(risk.risk_score), risk_level_label,
                _fmt_date(risk.term_date), risk.existing_measures or "—", risk.additional_measures or "—",
                risk.status or "Açık", len(dofs), "",
            ]
            if has_fine_kinney
            else [
                risk.risk_code, _dept(risk), risk.activity, _hazard_name(risk, hazard_map),
                _hazard_code(risk, hazard_map), risk.risk_definition, risk.affected_people or "—",
                _score_text(risk.probability), _score_text(risk.severity), _score_text(risk.risk_score),
                risk_level_label, _fmt_date(risk.term_date), risk.existing_measures or "—",
                risk.additional_measures or "—", risk.status or "Açık", len(dofs), "",
            ]
        )
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=9)
        level = _normalized_risk_level(risk, (validity or {}).get("method_code") or "5x5_l")
        if level in level_fills:
            if has_hazop:
                ws.cell(row=idx, column=17).fill = level_fills[level]
            else:
                score_column = 12 if has_fine_kinney else 10
                ws.cell(row=idx, column=score_column).fill = level_fills[level]
                ws.cell(row=idx, column=score_column + 1).fill = level_fills[level]

        photo_path = _risk_excel_photo_path(risk)
        if photo_path:
            try:
                img = XLImage(photo_path)
                max_w, max_h = 140, 95
                if img.width and img.height:
                    ratio = min(max_w / img.width, max_h / img.height, 1)
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)
                img.anchor = f"{last_column}{idx}"
                ws.add_image(img)
                ws.row_dimensions[idx].height = max(ws.row_dimensions[idx].height or 15, 72)
            except Exception:
                ws.cell(row=idx, column=column_count, value="Fotoğraf var / eklenemedi")

    widths = (
        [12, 24, 15, 20, 20, 24, 18, 18, 30, 18, 12, 30, 30, 30, 30, 18, 18, 12, 10, 10, 18]
        if has_hazop
        else (
            [12, 20, 15, 20, 18, 12, 28, 16, 10, 16, 10, 10, 22, 12, 28, 28, 10, 10, 18]
            if has_fine_kinney
            else [12, 15, 20, 18, 12, 28, 16, 10, 10, 10, 14, 12, 28, 28, 10, 10, 18]
        )
    )
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    last_data_row = max(4, 4 + len(risks))
    ws.auto_filter.ref = f"A4:{last_column}{last_data_row}"
    ws.print_title_rows = "1:4"

    # Son sayfa: yatay imza bloğu + sayfa sayısı beyanı
    sign_start = last_data_row + 2
    ws.row_breaks.append(Break(id=last_data_row + 1))
    people = _risk_sign_people(
        prepared_by=prepared_by,
        workplace_physician=workplace_physician,
        employer_representative=employer_representative,
        employee_representative=employee_representative,
        support_staff=support_staff,
    )
    after_sign = _write_excel_horizontal_sign_block(
        ws,
        start_row=sign_start,
        people=people,
        header_font=header_font,
        header_fill=header_fill,
        thin=thin,
        merge_cols=column_count,
    )

    declare_row = after_sign
    ws.merge_cells(start_row=declare_row, start_column=1, end_row=declare_row, end_column=column_count)
    declare = (
        f"İş bu risk değerlendirme raporu {page_count} sayfadan oluşur. "
        f"(Yazdırma önizlemesinde Sayfa X / N satırı ile doğrulanır; N farklıysa N esas alınır.)"
    )
    cell = ws.cell(row=declare_row, column=1, value=declare)
    cell.font = Font(name="Calibri", bold=True, size=11, color="1a5276")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[declare_row].height = 32

    note_row = declare_row + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=column_count)
    ws.cell(
        row=note_row,
        column=1,
        value=(
            "Her yazdırılan sayfanın altında aynı yatay İMZA/ONAY ekip satırı ve sayfa numarası yer alır. "
            f"Belge: {getattr(company, 'risk_document_no', None) or ('RD-' + str(getattr(company, 'id', '') or ''))} · "
            f"Düzenleme: {datetime.now().strftime('%d.%m.%Y %H:%M')} · {CREATOR_LINE}"
        ),
    )
    ws.cell(row=note_row, column=1).font = Font(name="Calibri", size=8, italic=True, color="6c757d")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)

    _apply_excel_page_chrome(ws, team_line=team_line, doc_label=doc_label)

    # DÖF sheet
    ws2 = wb.create_sheet("DÖF Listesi")
    dof_headers = [
        "DÖF No",
        "Risk Kodu",
        "Faaliyet",
        "Tehlike",
        "Yapılacak İş",
        "Sorumlu",
        "Sorumlu Bölüm",
        "Termin",
        "Maliyet",
        "Durum",
        "Tamamlanma Notu",
    ]
    for col, header in enumerate(dof_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin

    row = 2
    for risk in risks:
        for dof in getattr(risk, "dofs", None) or []:
            data = [
                dof.dof_code,
                risk.risk_code,
                risk.activity,
                _hazard_name(risk, hazard_map),
                dof.description,
                dof.responsible_person or "—",
                dof.responsible_department or "—",
                _fmt_date(dof.term_date),
                dof.cost_estimate if dof.cost_estimate is not None else "—",
                dof.status or "Açık",
                dof.completion_note or "—",
            ]
            for col, value in enumerate(data, 1):
                cell = ws2.cell(row=row, column=col, value=value)
                cell.border = thin
                cell.font = Font(size=9)
                cell.alignment = Alignment(wrap_text=True)
            row += 1
    for i, w in enumerate([12, 12, 18, 16, 32, 16, 16, 12, 10, 12, 24], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Stats
    ws3 = wb.create_sheet("İstatistikler")
    ws3.merge_cells("A1:C1")
    ws3["A1"] = "RİSK İSTATİSTİKLERİ"
    ws3["A1"].font = Font(bold=True, size=14, color="1a5276")
    ws3["A3"] = "Risk Seviyesi"
    ws3["B3"] = "Adet"
    ws3["C3"] = "Yüzde"
    for col in ["A3", "B3", "C3"]:
        ws3[col].font = header_font
        ws3[col].fill = header_fill
        ws3[col].border = thin

    risk_levels: dict[str, int] = {}
    for r in risks:
        rl = _normalized_risk_level(r)
        risk_levels[rl] = risk_levels.get(rl, 0) + 1
    total = len(risks)
    for i, level in enumerate(["Çok Yüksek", "Yüksek", "Orta", "Düşük", "Kabul Edilebilir"], 4):
        count = risk_levels.get(level, 0)
        ws3.cell(row=i, column=1, value=level).border = thin
        ws3.cell(row=i, column=2, value=count).border = thin
        ws3.cell(row=i, column=3, value=f"{(count / total * 100):.1f}%" if total else "%0").border = thin

    open_dofs = sum(1 for r in risks for d in (getattr(r, "dofs", None) or []) if not d.is_completed)
    done_dofs = sum(1 for r in risks for d in (getattr(r, "dofs", None) or []) if d.is_completed)
    ws3["A10"] = "Açık DÖF"
    ws3["B10"] = open_dofs
    ws3["A11"] = "Tamamlanan DÖF"
    ws3["B11"] = done_dofs
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 10

    if nace_roadmap:
        # Separate sheet keeps the existing risk/DÖF/statistics sheets stable
        # while making the NACE report scope auditable and exportable.
        ws4 = wb.create_sheet("NACE Yol Haritası")
        ws4.merge_cells("A1:E1")
        ws4["A1"] = "NACE ODAKLI RİSK KAPSAMI VE YOL HARİTASI"
        ws4["A1"].font = Font(name="Calibri", bold=True, size=14, color="1a5276")
        ws4["A1"].alignment = Alignment(horizontal="center")
        identity = nace_roadmap.get("identity") or {}
        ws4.merge_cells("A2:E2")
        ws4["A2"] = (
            f"İşyeri: {getattr(company, 'name', '')} | "
            f"NACE: {identity.get('code') or nace_roadmap.get('entered_nace_code') or '—'} | "
            f"Durum: {nace_roadmap.get('status_label') or '—'}"
        )
        ws4["A2"].font = Font(size=9, color="2c3e50")
        ws4["A2"].alignment = Alignment(wrap_text=True)
        ws4.merge_cells("A3:E3")
        ws4["A3"] = (
            f"Açıklama: {identity.get('description') or '—'} | "
            f"Bölüm: {identity.get('section_name') or '—'} | "
            "NACE tek başına saha risk değerlendirmesi yerine geçmez."
        )
        ws4["A3"].font = Font(size=9, italic=True, color="6c757d")
        ws4["A3"].alignment = Alignment(wrap_text=True)

        for col, header in enumerate(["Sıra / Faz", "Başlık", "Açıklama", "Mevzuat", "Modül / Durum"], 1):
            cell = ws4.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin

        row = 6
        for index, item in enumerate(nace_roadmap.get("report_checklist") or [], 1):
            values = [
                f"Rapor {index}",
                item.get("title") or item.get("key") or "—",
                item.get("description") or "—",
                item.get("legal_basis") or "—",
                item.get("module") or "—",
            ]
            for col, value in enumerate(values, 1):
                cell = ws4.cell(row=row, column=col, value=value)
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(size=9)
            row += 1

        row += 1
        for col, header in enumerate(["Yol Haritası", "Başlık", "Açıklama", "Mevzuat", "Modül / Durum"], 1):
            cell = ws4.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin
        row += 1
        for index, item in enumerate(nace_roadmap.get("roadmap") or [], 1):
            values = [
                f"{index}. {item.get('phase') or '—'}",
                item.get("title") or item.get("key") or "—",
                item.get("description") or "—",
                item.get("legal_basis") or "—",
                item.get("module") or "—",
            ]
            for col, value in enumerate(values, 1):
                cell = ws4.cell(row=row, column=col, value=value)
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(size=9)
            row += 1

        row += 1
        ws4.cell(row=row, column=1, value="NACE teknik risk başlıkları").font = header_font
        ws4.cell(row=row, column=1).fill = header_fill
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        domains = list(nace_roadmap.get("technical_risk_tags") or []) + list(nace_roadmap.get("special_risks") or [])
        if domains:
            for item in domains:
                values = [
                    item.get("kind") or "technical",
                    item.get("label") or item.get("key") or "—",
                    item.get("description") or "—",
                    item.get("category") or "—",
                    item.get("source") or "—",
                ]
                for col, value in enumerate(values, 1):
                    cell = ws4.cell(row=row, column=col, value=value)
                    cell.border = thin
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.font = Font(size=9)
                row += 1
        else:
            ws4.cell(row=row, column=1, value="Teknik risk eşleştirmesi yok; uzman saha incelemesi gerekli.")
            ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1

        row += 1
        ws4.cell(row=row, column=1, value="Kapsam durumu").font = header_font
        ws4.cell(row=row, column=1).fill = header_fill
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        coverage = nace_roadmap.get("coverage") or {}
        for label, key in (
            ("Risk kaydı", "risk_records"),
            ("Bölüm", "departments"),
            ("Açık DÖF", "open_dofs"),
            ("Tamamlanan DÖF", "completed_dofs"),
        ):
            ws4.cell(row=row, column=1, value=label).border = thin
            ws4.cell(row=row, column=2, value=coverage.get(key, 0)).border = thin
            row += 1
        for warning in nace_roadmap.get("warnings") or []:
            ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws4.cell(row=row, column=1, value=f"Uyarı: {warning}").font = Font(size=9, italic=True, color="9a3412")
            ws4.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1

        for col, width in enumerate([18, 30, 60, 34, 22], 1):
            ws4.column_dimensions[get_column_letter(col)].width = width
        ws4.freeze_panes = "A6"
        ws4.auto_filter.ref = f"A5:E{max(5, row - 1)}"
        _apply_excel_page_chrome(ws4, team_line=team_line, doc_label=f"NACE Yol Haritası — {getattr(company, 'name', '')}")

    _apply_excel_page_chrome(ws2, team_line=team_line, doc_label=f"DÖF Listesi — {getattr(company, 'name', '')}")
    _apply_excel_page_chrome(ws3, team_line=team_line, doc_label=f"Risk İstatistikleri — {getattr(company, 'name', '')}")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def build_dof_excel(
    *,
    company,
    risks: list,
    hazard_map: dict | None = None,
    method_code: str | None = None,
) -> bytes:
    """PRO /rapor/dof-excel — yalnız DÖF listesi."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    hazard_map = hazard_map or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "DÖF Listesi"
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wb.properties.creator = CREATOR_LINE
    wb.properties.title = "DÖF Listesi"

    ws.merge_cells("A1:I1")
    ws["A1"] = f"DÜZELTİCİ-ÖNLEYİCİ FAALİYET LİSTESİ - {getattr(company, 'name', '')}"
    ws["A1"].font = Font(bold=True, size=14, color="1A5276")
    ws.merge_cells("A2:I2")
    ws["A2"] = CREATOR_LINE
    ws["A2"].font = Font(size=9, italic=True, color="6C757D")
    ws["A2"].alignment = Alignment(horizontal="center")
    if method_code:
        ws.merge_cells("A3:I3")
        ws["A3"] = f"Yöntem filtresi: {resolve_method(method_code)['short']}"
        ws["A3"].font = Font(size=9, italic=True, color="6C757D")
        ws["A3"].alignment = Alignment(horizontal="center")

    headers = [
        "DÖF No",
        "Risk Kodu",
        "Bölüm",
        "Tehlike",
        "Yapılacak İş",
        "Sorumlu",
        "Termin Tarihi",
        "Durum",
        "Tamamlanma Tarihi",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_i = 5
    for risk in risks:
        haz = hazard_map.get(getattr(risk, "hazard_id", None))
        haz_name = getattr(haz, "name", None) if haz else "—"
        for dof in getattr(risk, "dofs", None) or []:
            values = [
                dof.dof_code,
                getattr(risk, "risk_code", "") or "—",
                getattr(risk, "department_name", None) or "—",
                haz_name or "—",
                dof.description,
                dof.responsible_person or "—",
                _fmt_date(dof.term_date),
                dof.status or "Açık",
                _fmt_date(dof.completion_date),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.border = thin
                cell.font = Font(size=9)
                cell.alignment = Alignment(wrap_text=True)
            row_i += 1

    for i, w in enumerate([12, 12, 15, 18, 30, 18, 12, 12, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.oddFooter.center.text = CREATOR_LINE

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
