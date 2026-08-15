"""PDF and Excel outputs for the remote-training participation register."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


STATUS_LABELS = {
    "not_started": "Başlamadı",
    "in_progress": "Devam ediyor",
    "completed": "Başarılı / Tamamlandı",
    "failed": "Başarısız",
    "expired": "Süresi geçti",
    "revoked": "Silinmiş atama",
}

_FONT_DIR = Path(__file__).resolve().parent.parent / "assets" / "fonts"
_PDF_FONT = "Helvetica"
_PDF_FONT_BOLD = "Helvetica-Bold"


def _register_pdf_fonts() -> None:
    global _PDF_FONT, _PDF_FONT_BOLD
    candidates = (
        (_FONT_DIR / "DejaVuSans.ttf", _FONT_DIR / "DejaVuSans-Bold.ttf"),
        (Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"), Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")),
        (Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"), Path("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf")),
    )
    for regular, bold in candidates:
        if not regular.exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont("RemoteReportSans", str(regular)))
            pdfmetrics.registerFont(TTFont("RemoteReportSans-Bold", str(bold if bold.exists() else regular)))
            _PDF_FONT = "RemoteReportSans"
            _PDF_FONT_BOLD = "RemoteReportSans-Bold"
            return
        except Exception:
            continue


_register_pdf_fonts()


def _text(value: Any, fallback: str = "—") -> str:
    clean = str(value or "").strip()
    return clean or fallback


def _status_label(value: Any) -> str:
    return STATUS_LABELS.get(str(value or "").strip(), _text(value))


def _report_rows(rows: list[dict[str, Any]]) -> list[list[str]]:
    result: list[list[str]] = []
    for row in rows:
        summary = row.get("summary") or {}
        completed = int(summary.get("completed_video_count") or 0)
        required = int(summary.get("required_video_count") or 0)
        score = row.get("examination_score")
        result.append(
            [
                _text(row.get("company_name")),
                _text(row.get("branch_name"), "Firma geneli"),
                _text(row.get("employee_name")),
                _text(row.get("program_title")),
                _status_label(row.get("status")),
                f"{completed}/{required}",
                f"%{score}" if score is not None else "—",
                "Hazır" if row.get("certificate_ready") else "Hazır değil",
                _text(row.get("certificate_number")),
                _text(row.get("completed_at")),
            ]
        )
    return result


def build_remote_training_status_xlsx(
    rows: list[dict[str, Any]],
    *,
    company_name: str,
    branch_name: str | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    """Build a company-scoped participation/status register."""
    generated_at = generated_at or datetime.utcnow()
    wb = Workbook()
    ws = wb.active
    ws.title = "Eğitim Katılım"
    ws.sheet_view.showGridLines = False

    title = f"Eğitim Katılım ve Belgelendirme Raporu — {_text(company_name)}"
    if branch_name:
        title += f" / {branch_name}"
    ws.append([title])
    ws.append([f"Oluşturulma: {generated_at.strftime('%d.%m.%Y %H:%M')} UTC", f"Kayıt sayısı: {len(rows)}"])
    ws.append([])
    headers = [
        "Firma",
        "İşyeri / Şube",
        "Çalışan",
        "Eğitim",
        "Durum",
        "Video ilerlemesi",
        "Sınav puanı",
        "Katılım belgesi",
        "Belge numarası",
        "Tamamlanma tarihi",
    ]
    ws.append(headers)
    for row in _report_rows(rows):
        ws.append(row)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(headers))
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="123B59")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="496174")
    ws["B2"].font = Font(name="Calibri", size=10, italic=True, color="496174")

    header_row = 4
    header_fill = PatternFill("solid", fgColor="123B59")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E5EC")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        status_cell = row[4]
        if status_cell.value == "Başarılı / Tamamlandı":
            status_cell.font = Font(name="Calibri", bold=True, color="087443")
        elif status_cell.value == "Başarısız":
            status_cell.font = Font(name="Calibri", bold=True, color="B42318")

    widths = [26, 24, 24, 34, 23, 16, 14, 18, 28, 21]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:J{max(ws.max_row, 4)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def build_remote_training_status_pdf(
    rows: list[dict[str, Any]],
    *,
    company_name: str,
    branch_name: str | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    """Build a compact printable company participation/status register."""
    generated_at = generated_at or datetime.utcnow()
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=11 * mm,
        bottomMargin=11 * mm,
        title="Eğitim Katılım ve Belgelendirme Raporu",
        author="İSG SUITE",
    )
    title_style = ParagraphStyle(
        "RemoteReportTitle",
        fontName=_PDF_FONT_BOLD,
        fontSize=15,
        leading=19,
        textColor=colors.HexColor("#123B59"),
        spaceAfter=3 * mm,
    )
    meta_style = ParagraphStyle(
        "RemoteReportMeta",
        fontName=_PDF_FONT,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#496174"),
        spaceAfter=2 * mm,
    )
    cell_style = ParagraphStyle(
        "RemoteReportCell",
        fontName=_PDF_FONT,
        fontSize=7.4,
        leading=9,
        textColor=colors.HexColor("#17324D"),
    )
    header_style = ParagraphStyle(
        "RemoteReportHeader",
        parent=cell_style,
        fontName=_PDF_FONT_BOLD,
        textColor=colors.white,
        alignment=1,
    )
    report_title = f"Eğitim Katılım ve Belgelendirme Raporu — {_text(company_name)}"
    if branch_name:
        report_title += f" / {branch_name}"
    story = [
        Paragraph(escape(report_title), title_style),
        Paragraph(
            escape(
                f"Kayıt sayısı: {len(rows)} · Oluşturulma: {generated_at.strftime('%d.%m.%Y %H:%M')} UTC · "
                "Başarılı, başarısız ve devam eden çalışan eğitim kayıtları"
            ),
            meta_style,
        ),
        Spacer(1, 2 * mm),
    ]
    headers = ["Firma", "İşyeri / Şube", "Çalışan", "Eğitim", "Durum", "Video", "Sınav", "Belge"]
    table_data = [[Paragraph(escape(value), header_style) for value in headers]]
    for row in rows:
        summary = row.get("summary") or {}
        score = row.get("examination_score")
        values = [
            _text(row.get("company_name")),
            _text(row.get("branch_name"), "Firma geneli"),
            _text(row.get("employee_name")),
            _text(row.get("program_title")),
            _status_label(row.get("status")),
            f"{int(summary.get('completed_video_count') or 0)}/{int(summary.get('required_video_count') or 0)}",
            f"%{score}" if score is not None else "—",
            "Hazır" if row.get("certificate_ready") else "Hazır değil",
        ]
        table_data.append([Paragraph(escape(value), cell_style) for value in values])
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[32 * mm, 31 * mm, 31 * mm, 57 * mm, 32 * mm, 18 * mm, 18 * mm, 24 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123B59")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E5EC")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FBFD")]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return output.getvalue()
