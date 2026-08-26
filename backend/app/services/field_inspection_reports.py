"""Onaylanmış görsel saha denetimi için PDF ve Excel çıktıları."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.object_store import get_object_store


_ASSETS = Path(__file__).resolve().parent.parent / "assets" / "fonts"
_FONT = "Helvetica"
_BOLD = "Helvetica-Bold"
_GPS_WARNING = "Bu bulgu için kesin mevzuat maddesi otomatik olarak doğrulanamadı. İş güvenliği uzmanı tarafından mevzuat kontrolü yapılmalıdır."


def _register_fonts() -> None:
    global _FONT, _BOLD
    regular = _ASSETS / "DejaVuSans.ttf"
    bold = _ASSETS / "DejaVuSans-Bold.ttf"
    if not regular.exists():
        regular = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
        bold = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")
    if regular.exists():
        try:
            pdfmetrics.registerFont(TTFont("FieldReportSans", str(regular)))
            pdfmetrics.registerFont(TTFont("FieldReportSans-Bold", str(bold if bold.exists() else regular)))
            _FONT, _BOLD = "FieldReportSans", "FieldReportSans-Bold"
        except Exception:
            pass


_register_fonts()


def _text(value: Any, fallback: str = "—") -> str:
    clean = str(value or "").strip()
    return clean or fallback


def _date(value: Any) -> str:
    if not value:
        return "—"
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y %H:%M") if hasattr(value, "hour") else value.strftime("%d.%m.%Y")
    return _text(value)


def _para(value: Any, *, bold: bool = False, size: int = 8, color: str = "#243447") -> Paragraph:
    style = ParagraphStyle(
        name=f"field-{size}-{bold}-{color}", fontName=_BOLD if bold else _FONT, fontSize=size,
        leading=size + 3, textColor=colors.HexColor(color), spaceAfter=2,
    )
    return Paragraph(escape(_text(value)).replace("\n", "<br/>"), style)


def _section(title: str) -> Table:
    table = Table([[_para(title, bold=True, size=11, color="#ffffff")]], colWidths=[178 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0f4c5c")),
        ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return table


def _header_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont(_FONT, 7)
    canvas.setFillColor(colors.HexColor("#64748b"))
    canvas.drawString(14 * mm, 10 * mm, "İSG Suite OSGB — Görsel Saha Denetimi")
    canvas.drawRightString(A4[0] - 14 * mm, 10 * mm, f"Sayfa {doc.page}")
    canvas.restoreState()


def _photo_flowable(data: bytes | None, max_width: float = 83 * mm, max_height: float = 64 * mm):
    if not data:
        return _para("Görsel bulunamadı", size=8, color="#64748b")
    try:
        with PILImage.open(BytesIO(data)) as image:
            width, height = image.size
        ratio = min(max_width / max(width, 1), max_height / max(height, 1))
        return Image(BytesIO(data), width=max(15 * mm, width * ratio), height=max(15 * mm, height * ratio))
    except Exception:
        return _para("Görsel rapora eklenemedi", size=8, color="#b91c1c")


def _store_bytes(path: str | None) -> bytes | None:
    if not path:
        return None
    try:
        return get_object_store().get_bytes(path)
    except Exception:
        return None


def _gps_text(inspection, *, include_gps: bool) -> str:
    if not include_gps:
        return "Yetki kapsamında GPS ayrıntısı gösterilmedi."
    status = _text(getattr(inspection, "gps_status", None), "GPS alınamadı")
    lat = getattr(inspection, "gps_lat", None)
    lng = getattr(inspection, "gps_lng", None)
    if lat is None or lng is None:
        return f"{status}; açıklama: {_text(getattr(inspection, 'gps_reason', None) or getattr(inspection, 'manual_location_note', None))}"
    accuracy = getattr(inspection, "gps_accuracy_m", None)
    return f"{status}: {float(lat):.7f}, {float(lng):.7f} (±{float(accuracy):.1f} m)" if accuracy is not None else f"{status}: {float(lat):.7f}, {float(lng):.7f}"


def build_field_inspection_pdf(*, inspection, company, site, area, equipment, photos: list, findings: list, actions: list, include_gps: bool = True, generated_at: datetime | None = None) -> bytes:
    """Uzman onayı sonrası oluşturulan, görsel kanıt ve faaliyetli PDF."""
    generated_at = generated_at or datetime.utcnow()
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm,
        topMargin=15 * mm, bottomMargin=17 * mm, title=f"Saha Denetimi {_text(getattr(inspection, 'inspection_no', None))}",
    )
    story: list[Any] = []
    story.extend([
        _para("GÖRSEL SAHA DENETİM RAPORU", bold=True, size=18, color="#0f4c5c"),
        _para(f"Rapor no: {_text(getattr(inspection, 'inspection_no', None))}  |  Revizyon: {_text(getattr(inspection, 'report_revision_no', None), '1')}  |  Oluşturulma: {_date(generated_at)}", size=8, color="#64748b"),
        Spacer(1, 5 * mm), _section("Rapor üst bilgileri"), Spacer(1, 2 * mm),
    ])
    context_rows = [
        [_para("İşyeri", bold=True), _para(getattr(company, "name", None)), _para("İşveren", bold=True), _para(getattr(company, "authorized_person", None))],
        [_para("SGK sicil no", bold=True), _para(getattr(company, "sgk_registry_no", None)), _para("NACE / tehlike", bold=True), _para(f"{_text(getattr(company, 'nace_code', None))} / {_text(getattr(company, 'hazard_class', None))}")],
        [_para("Tesis / saha", bold=True), _para(getattr(site, "name", None)), _para("Saha türü", bold=True), _para(getattr(site, "site_type", None))],
        [_para("Bölüm / alan", bold=True), _para(getattr(area, "name", None)), _para("Ekipman / nokta", bold=True), _para(getattr(equipment, "name", None) if equipment else None)],
        [_para("Denetim zamanı", bold=True), _para(_date(getattr(inspection, "inspection_at", None))), _para("Saat dilimi", bold=True), _para(getattr(inspection, "timezone", None))],
        [_para("Denetimi yapan uzman", bold=True), _para(getattr(inspection, "report_creator_name", None)), _para("Onaylayan uzman", bold=True), _para(getattr(inspection, "report_approver_name", None))],
        [_para("GPS", bold=True), _para(_gps_text(inspection, include_gps=include_gps)), _para("Durum", bold=True), _para(getattr(inspection, "status", None))],
        [_para("AI analiz tarihi", bold=True), _para(_date(getattr(inspection, "ai_analysis_at", None))), _para("AI model / sürüm", bold=True), _para(f"{_text(getattr(inspection, 'ai_model_name', None))} / {_text(getattr(inspection, 'ai_model_version', None))}")],
    ]
    table = Table(context_rows, colWidths=[28 * mm, 61 * mm, 32 * mm, 57 * mm], repeatRows=0)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5f9")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f1f5f9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([table, Spacer(1, 4 * mm), _section("Denetim kapsamı ve AI durumu"), Spacer(1, 2 * mm)])
    categories = getattr(inspection, "selected_category_names", None) or getattr(inspection, "selected_category_ids_json", None)
    if isinstance(categories, (list, tuple)):
        categories = ", ".join(str(item) for item in categories)
    story.append(_para(f"Seçilen tehlike kategorileri: {_text(categories, 'Tüm görünür tehlikeler')}", size=8))
    story.append(_para("AI çıktıları yalnızca uzman yardımcısı taslağıdır. Bu rapordaki bulgular uzman incelemesi ve onayı sonrası yayımlanmıştır; fotoğrafta görünmeyen hususlar otomatik olarak kabul edilmez.", size=8, color="#92400e"))
    if getattr(inspection, "ai_warning", None):
        story.append(_para(f"AI uyarısı: {inspection.ai_warning}", size=8, color="#b45309"))
    story.extend([Spacer(1, 4 * mm), _section("Fotoğraf kanıtları"), Spacer(1, 2 * mm)])
    if not photos:
        story.append(_para("Bu denetimde fotoğraf kanıtı bulunmuyor.", size=8, color="#b91c1c"))
    for number, photo in enumerate(photos, start=1):
        original = _store_bytes(getattr(photo, "original_storage_path", None))
        marked = _store_bytes(getattr(photo, "marked_storage_path", None))
        photo_context = " / ".join(filter(None, (getattr(photo, "report_site_name", None), getattr(photo, "report_area_name", None), getattr(photo, "report_equipment_name", None))))
        photo_table = Table([
            [_para(f"Fotoğraf {number}: {_text(getattr(photo, 'original_name', None))}{(' · ' + photo_context) if photo_context else ''}", bold=True, size=9), ""],
            [_photo_flowable(original), _photo_flowable(marked)],
            [_para("Orijinal", size=7, color="#64748b"), _para("İşaretlenmiş / analiz", size=7, color="#64748b")],
        ], colWidths=[89 * mm, 89 * mm])
        photo_table.setStyle(TableStyle([
            ("SPAN", (0, 0), (1, 0)), ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("INNERGRID", (0, 1), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.extend([photo_table, Spacer(1, 3 * mm)])
    story.extend([PageBreak(), _section("Bulgular ve uzman değerlendirmesi"), Spacer(1, 2 * mm)])
    if not findings:
        story.append(_para("Onaylanan bulgu bulunmuyor.", size=8))
    for number, finding in enumerate(findings, start=1):
        story.extend([
            _para(f"{number}. {_text(getattr(finding, 'hazard_name', None))} — {_text(getattr(finding, 'suggested_priority', None))}", bold=True, size=10, color="#0f4c5c"),
            Table([
                [_para("Görsel kanıt", bold=True), _para(getattr(finding, "visual_evidence", None))],
                [_para("Uygunsuzluk", bold=True), _para(getattr(finding, "nonconformity_description", None))],
                [_para("Olası sonuç / neden", bold=True), _para(f"{_text(getattr(finding, 'possible_harm', None))} / {_text(getattr(finding, 'possible_cause', None))}")],
                [_para("Aksiyon önerisi", bold=True), _para(f"Acil: {_text(getattr(finding, 'urgent_action', None))}\nDüzeltici: {_text(getattr(finding, 'corrective_action', None))}\nÖnleyici: {_text(getattr(finding, 'preventive_action', None))}")],
                [_para("Uzman durumu", bold=True), _para(getattr(finding, "status", None))],
            ], colWidths=[34 * mm, 144 * mm]), Spacer(1, 2 * mm),
        ])
        finding_table = story[-2]
        finding_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f8fafc")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        references = list(getattr(finding, "legal_references", None) or [])
        if references:
            story.append(_para("Mevzuat atıfları (uzman kontrolü)", bold=True, size=8))
            for ref in references:
                story.append(_para(f"• {_text(getattr(ref, 'regulation_name', None))}; madde: {_text(getattr(ref, 'article', None))}; durum: {_text(getattr(ref, 'verification_status', None))}", size=8))
            if not any(getattr(ref, "verification_status", None) == "verified" for ref in references):
                story.append(_para(_GPS_WARNING, size=8, color="#92400e"))
        else:
            story.append(_para(_GPS_WARNING, size=8, color="#92400e"))
        story.append(Spacer(1, 3 * mm))
    story.extend([_section("Düzeltici / önleyici faaliyetler"), Spacer(1, 2 * mm)])
    action_rows = [[_para("Faaliyet", bold=True), _para("Sorumlu", bold=True), _para("Termin", bold=True), _para("Durum", bold=True)]]
    for action in actions:
        responsible = getattr(action, "report_responsible_name", None) or getattr(action, "responsible_person", None)
        action_rows.append([_para(f"{_text(getattr(action, 'title', None))}\n{_text(getattr(action, 'activity', None))}"), _para(f"{_text(responsible)} {_text(getattr(action, 'responsible_role', None), '')}"), _para(_date(getattr(action, "term_date", None))), _para(getattr(action, "status", None))])
    if len(action_rows) == 1:
        action_rows.append([_para("Faaliyet bulunmuyor."), _para("—"), _para("—"), _para("—")])
    actions_table = Table(action_rows, colWidths=[85 * mm, 42 * mm, 25 * mm, 26 * mm], repeatRows=1)
    actions_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(actions_table)
    story.extend([Spacer(1, 4 * mm), _section("Revizyon ve audit geçmişi"), Spacer(1, 2 * mm)])
    revision_rows = [[_para("Tarih", bold=True), _para("İşlem", bold=True), _para("Açıklama", bold=True)]]
    for item in list(getattr(inspection, "revision_history", None) or [])[:50]:
        revision_rows.append([_para(_date(item.get("created_at"))), _para(item.get("action")), _para(item.get("description"))])
    if len(revision_rows) == 1:
        revision_rows.append([_para("Audit kaydı bulunmuyor."), _para("—"), _para("—")])
    revision_table = Table(revision_rows, colWidths=[32 * mm, 48 * mm, 98 * mm], repeatRows=1)
    revision_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(revision_table)
    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    return output.getvalue()


def _excel_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F4C5C")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        index = column[0].column
        longest = max((len(str(cell.value or "")) for cell in column), default=8)
        ws.column_dimensions[get_column_letter(index)].width = min(48, max(12, longest + 2))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_field_inspection_excel(*, inspection, company, site, area, equipment, photos: list, findings: list, actions: list, include_gps: bool = True, generated_at: datetime | None = None) -> bytes:
    generated_at = generated_at or datetime.utcnow()
    wb = Workbook()
    summary = wb.active
    summary.title = "Özet"
    summary.sheet_view.showGridLines = False
    summary_rows = [
        ["Rapor no", _text(getattr(inspection, "inspection_no", None))], ["Revizyon", _text(getattr(inspection, "report_revision_no", None), "1")],
        ["İşyeri", _text(getattr(company, "name", None))], ["İşveren", _text(getattr(company, "authorized_person", None))],
        ["SGK sicil no", _text(getattr(company, "sgk_registry_no", None))], ["NACE", _text(getattr(company, "nace_code", None))],
        ["Tehlike sınıfı", _text(getattr(company, "hazard_class", None))], ["Tesis / saha", _text(getattr(site, "name", None))],
        ["Bölüm / alan", _text(getattr(area, "name", None))], ["Ekipman / nokta", _text(getattr(equipment, "name", None) if equipment else None)],
        ["Denetimi yapan uzman", _text(getattr(inspection, "report_creator_name", None))], ["Onaylayan uzman", _text(getattr(inspection, "report_approver_name", None))],
        ["Denetim zamanı", _date(getattr(inspection, "inspection_at", None))], ["GPS", _gps_text(inspection, include_gps=include_gps)],
        ["Durum", _text(getattr(inspection, "status", None))], ["AI durumu", _text(getattr(inspection, "ai_status", None))],
        ["AI analiz tarihi", _date(getattr(inspection, "ai_analysis_at", None))], ["AI model / sürüm", f"{_text(getattr(inspection, 'ai_model_name', None))} / {_text(getattr(inspection, 'ai_model_version', None))}"],
        ["Oluşturulma", _date(generated_at)], ["Bulgu sayısı", len(findings)], ["Faaliyet sayısı", len(actions)],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 72
    for cell in summary["A"]:
        cell.font = Font(bold=True, color="0F4C5C")
    _excel_sheet(wb, "Bulgular", ["No", "Tehlike", "Kategori", "Görsel kanıt", "Uygunsuzluk", "Olası neden", "Olası sonuç", "Öncelik", "Durum", "Sorumlu rol", "Termin"], [
        [getattr(row, "finding_no", i + 1), _text(getattr(row, "hazard_name", None)), _text(getattr(row, "category_name", None)), _text(getattr(row, "visual_evidence", None)), _text(getattr(row, "nonconformity_description", None)), _text(getattr(row, "possible_cause", None)), _text(getattr(row, "possible_harm", None)), _text(getattr(row, "suggested_priority", None)), _text(getattr(row, "status", None)), _text(getattr(row, "suggested_responsible_role", None)), _date(getattr(row, "suggested_term_date", None))]
        for i, row in enumerate(findings)
    ])
    _excel_sheet(wb, "Faaliyetler", ["No", "Başlık", "Faaliyet", "Acil", "Kalıcı çözüm", "Önleyici", "Sorumlu kişi", "Sorumlu rol", "Termin", "Öncelik", "Durum", "Tamamlanma"], [
        [i + 1, _text(getattr(row, "title", None)), _text(getattr(row, "activity", None)), _text(getattr(row, "urgent_action", None)), _text(getattr(row, "permanent_solution", None)), _text(getattr(row, "preventive_action", None)), _text(getattr(row, "report_responsible_name", None) or getattr(row, "responsible_person", None)), _text(getattr(row, "responsible_role", None)), _date(getattr(row, "term_date", None)), _text(getattr(row, "priority", None)), _text(getattr(row, "status", None)), _date(getattr(row, "completion_date", None))]
        for i, row in enumerate(actions)
    ])
    _excel_sheet(wb, "Fotoğraflar", ["No", "Dosya adı", "Tür", "Boyut", "Genişlik", "Yükseklik", "Tesis/saha", "Bölüm/alan", "Ekipman/nokta", "Çekim zamanı", "GPS durumu", "Enlem", "Boylam", "Doğruluk (m)"], [
        [i + 1, _text(getattr(row, "original_name", None)), _text(getattr(row, "content_type", None)), getattr(row, "file_size", None) or 0, getattr(row, "width", None), getattr(row, "height", None), _text(getattr(row, "report_site_name", None)), _text(getattr(row, "report_area_name", None)), _text(getattr(row, "report_equipment_name", None)), _date(getattr(row, "captured_at", None)), _text(getattr(row, "gps_status", None)), getattr(row, "gps_lat", None) if include_gps else "Gizli", getattr(row, "gps_lng", None) if include_gps else "Gizli", getattr(row, "gps_accuracy_m", None) if include_gps else "Gizli"]
        for i, row in enumerate(photos)
    ])
    _excel_sheet(wb, "Mevzuat", ["Bulgu no", "Mevzuat", "Madde", "Fıkra", "Kaynak", "Sürüm", "Doğrulama", "İlişki açıklaması"], [
        [getattr(getattr(row, "finding", None), "finding_no", "—"), _text(getattr(row, "regulation_name", None)), _text(getattr(row, "article", None)), _text(getattr(row, "paragraph", None)), _text(getattr(row, "source_url", None)), _text(getattr(row, "source_version", None)), _text(getattr(row, "verification_status", None)), _text(getattr(row, "relation_explanation", None))]
        for finding in findings for row in list(getattr(finding, "legal_references", None) or [])
    ] + [["—", _GPS_WARNING, "", "", "", "", "needs_expert_review", ""]] if not any(getattr(row, "legal_references", None) for row in findings) else [])
    _excel_sheet(wb, "Revizyon", ["Tarih", "İşlem", "Varlık", "Kayıt", "Kullanıcı", "Açıklama"], [
        [_date(item.get("created_at")), _text(item.get("action")), _text(item.get("entity_type")), _text(item.get("entity_id")), _text(item.get("user_id")), _text(item.get("description"))]
        for item in list(getattr(inspection, "revision_history", None) or [])
    ] or [["—", "—", "—", "—", "—", "Audit kaydı bulunmuyor."]])
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
