"""Olay / ramak kala PDF ve Excel raporları — PRO rapor uyarlaması."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.incident_meta import EVENT_TYPES, meta_payload

PDF_FONT = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"
_ASSETS = Path(__file__).resolve().parent.parent / "assets" / "fonts"


def _register() -> None:
    global PDF_FONT, PDF_FONT_BOLD
    candidates = [
        (_ASSETS / "DejaVuSans.ttf", _ASSETS / "DejaVuSans-Bold.ttf", "IncDejaVu", "IncDejaVu-Bold"),
        (Path(r"C:\Windows\Fonts\arial.ttf"), Path(r"C:\Windows\Fonts\arialbd.ttf"), "IncArial", "IncArial-Bold"),
    ]
    for regular, bold, name, bold_name in candidates:
        if not regular.exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont(name, str(regular)))
            pdfmetrics.registerFont(TTFont(bold_name, str(bold if bold.exists() else regular)))
            registerFontFamily(name, normal=name, bold=bold_name, italic=name, boldItalic=bold_name)
            PDF_FONT, PDF_FONT_BOLD = name, bold_name
            return
        except Exception:
            continue
    raise RuntimeError("Olay PDF için Unicode font bulunamadı (DejaVu/Arial).")


def _fmt(d) -> str:
    if not d:
        return "—"
    if hasattr(d, "strftime"):
        return d.strftime("%d.%m.%Y")
    return str(d)


def _yes(v: bool | None) -> str:
    return "Evet" if v else "Hayır"


def completeness_gaps(incident, root_cause, dofs: list) -> list[str]:
    gaps: list[str] = []
    if not (incident.short_summary or "").strip() or len(incident.short_summary.strip()) < 20:
        gaps.append("Olay özeti yetersiz (en az 20 karakter).")
    if not (incident.detail or "").strip():
        gaps.append("Olay detayı eksik.")
    if not (incident.classification or "").strip():
        gaps.append("Sınıflandırma seçilmemiş.")
    if not (incident.location or "").strip():
        gaps.append("Olay yeri eksik.")
    if not root_cause or not (root_cause.root_cause_category or "").strip():
        gaps.append("Kök neden kategorisi eksik.")
    if not root_cause or not (root_cause.root_cause or "").strip():
        gaps.append("Kök neden metni eksik.")
    if not dofs:
        gaps.append("Olay DÖF kaydı oluşturulmamış.")
    if incident.event_type == "is_kazasi" and not incident.sgk_reported:
        gaps.append("İş kazasında SGK bildirimi işaretlenmemiş.")
    return gaps


def build_incident_excel(*, rows: list, company_names: dict[int, str] | None = None) -> bytes:
    """Filtreli olay / ramak kala kayıt defteri (Excel)."""
    company_names = company_names or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Olay Kayıtları"
    headers = [
        "Form No",
        "Firma",
        "Tip",
        "Durum",
        "Tarih",
        "Saat",
        "Yer",
        "Bölüm",
        "Alan",
        "Sınıflandırma",
        "Özet",
        "Olasılık",
        "Şiddet",
        "Risk Skoru",
        "Risk Seviyesi",
        "Kök Neden",
        "DÖF Sayısı",
        "SGK Bildirimi",
        "Kolluk",
        "İSG Uzmanı",
        "İşyeri Hekimi",
        "İşveren Vekili",
        "Kaydeden",
    ]
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    tip_map = {k: v.get("adi", k) for k, v in EVENT_TYPES.items()}
    for r in rows:
        rc = r.root_cause
        tip = tip_map.get(r.event_type, r.event_type)
        ws.append(
            [
                r.form_no,
                company_names.get(r.company_id, str(r.company_id)),
                tip,
                r.status,
                _fmt(r.event_date) if r.event_date else "",
                r.event_time or "",
                r.location or "",
                r.department or "",
                r.area or "",
                r.classification or "",
                r.short_summary or "",
                r.probability or 0,
                r.severity or 0,
                r.risk_score or 0,
                r.risk_level or "",
                (rc.root_cause if rc else "") or "",
                len(r.dofs or []),
                _yes(r.sgk_reported),
                _yes(r.police_reported),
                r.safety_specialist or "",
                r.workplace_physician or "",
                r.employer_representative or "",
                r.recorded_by_name or "",
            ]
        )

    for col in ws.columns:
        width = 12
        for cell in col:
            val = str(cell.value or "")
            width = min(max(width, min(len(val) + 2, 42)), 48)
        ws.column_dimensions[col[0].column_letter].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def build_capa_board_excel(*, rows: list[dict]) -> bytes:
    """Olay + risk DÖF birleşik panosu Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DOF Panosu"
    headers = [
        "Kaynak",
        "DÖF No",
        "Bağlı Kayıt",
        "Özet / Faaliyet",
        "Tespit",
        "Düzeltici Faaliyet",
        "Sorumlu",
        "Termin",
        "Durum",
        "Öncelik",
    ]
    header_fill = PatternFill("solid", fgColor="212529")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
    for r in rows:
        ws.append(
            [
                r.get("source") or "",
                r.get("code") or "",
                r.get("parent") or "",
                r.get("parentSummary") or "",
                r.get("title") or "",
                r.get("action") or "",
                r.get("responsible") or "",
                r.get("term") or "",
                r.get("status") or "",
                r.get("priority") or "",
            ]
        )
    for col in ws.columns:
        width = 12
        for cell in col:
            val = str(cell.value or "")
            width = min(max(width, min(len(val) + 2, 40)), 44)
        ws.column_dimensions[col[0].column_letter].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def build_incident_pdf(
    *,
    company_name: str,
    incident,
    root_cause=None,
    dofs: list | None = None,
    company=None,
) -> bytes:
    _register()
    dofs = list(dofs or [])
    gaps = completeness_gaps(incident, root_cause, dofs)
    tip = EVENT_TYPES.get(incident.event_type, {}).get("adi", incident.event_type)
    risk_opts = {o["code"]: o["label"] for o in meta_payload()["risk_analysis_options"]}

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm, topMargin=16 * mm, bottomMargin=14 * mm
    )
    styles = getSampleStyleSheet()
    for s in styles.byName.values():
        s.fontName = PDF_FONT
    title = ParagraphStyle(
        "T", parent=styles["Title"], fontName=PDF_FONT_BOLD, fontSize=14, textColor=colors.white, alignment=TA_CENTER
    )
    h = ParagraphStyle(
        "H",
        parent=styles["Heading2"],
        fontName=PDF_FONT_BOLD,
        fontSize=11,
        textColor=colors.HexColor("#0d6efd"),
        spaceBefore=8,
        spaceAfter=4,
    )
    body = ParagraphStyle("B", parent=styles["Normal"], fontName=PDF_FONT, fontSize=9, leading=12)
    label = ParagraphStyle("L", parent=styles["Normal"], fontName=PDF_FONT_BOLD, fontSize=9, leading=12)

    elements = []
    cover = Table(
        [
            [Paragraph(f"{tip.upper()} RAPORU", title)],
            [
                Paragraph(
                    company_name or "—",
                    ParagraphStyle("C", parent=body, textColor=colors.white, alignment=TA_CENTER),
                )
            ],
            [
                Paragraph(
                    f"Form No: {incident.form_no}",
                    ParagraphStyle("C2", parent=body, textColor=colors.white, alignment=TA_CENTER),
                )
            ],
        ],
        colWidths=[180 * mm],
    )
    cover.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0d6efd")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(cover)
    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph(f"Düzenleme: {datetime.now().strftime('%d.%m.%Y %H:%M')} · İSG Suite OSGB", body))

    if gaps:
        warn = "<b>UYARI:</b> Bu raporda eksik alanlar vardır:<br/>" + "<br/>".join(f"• {g}" for g in gaps)
        elements.append(Spacer(1, 3 * mm))
        elements.append(
            Paragraph(warn, ParagraphStyle("W", parent=body, backColor=colors.HexColor("#fff3cd"), borderPadding=6))
        )

    def row(etiket: str, deger: str):
        elements.append(Paragraph(f"<b>{etiket}</b> {deger or '—'}", body))

    elements.append(Paragraph("1. İŞYERİ / OLAY BİLGİLERİ", h))
    row("Olay tipi:", tip)
    row("Tarih / Saat:", f"{_fmt(incident.event_date)} {incident.event_time or ''}".strip())
    if company is not None:
        row("İşyeri:", getattr(company, "name", None) or company_name)
        row("Adres:", getattr(company, "address", None) or "—")
        row("SGK sicil no:", getattr(company, "sgk_registry_no", None) or "—")
        row("Tehlike sınıfı:", getattr(company, "hazard_class", None) or "—")
    row("Yer:", incident.location or "—")
    row("Bölüm / Alan:", f"{incident.department or '—'} / {incident.area or '—'}")
    row("Yapılan iş:", incident.work_being_done or "—")
    row("İlgili kişiler:", incident.related_people or "—")
    row("Tanık var mı:", _yes(incident.has_witness))
    if incident.has_witness or incident.witness_names:
        row("Tanıklar:", incident.witness_names or "—")
    row("Kullanılan ekipman:", incident.equipment_used or "—")
    row("Kimyasal madde:", incident.chemical_used or "—")
    row("Sınıflandırma:", incident.classification or "—")
    row("Oluşturan:", incident.recorded_by_name or "—")
    row("İSG uzmanı:", incident.safety_specialist or "—")
    row("İşyeri hekimi:", incident.workplace_physician or "—")
    row("İşveren vekili:", incident.employer_representative or "—")

    elements.append(Paragraph("2. OLAY AÇIKLAMASI", h))
    row("Özet:", incident.short_summary or "—")
    row("Detay:", incident.detail or "—")
    row("Yaralanma:", _yes(incident.injury_occurred))
    row("Sağlık şikayeti:", _yes(incident.health_complaint))
    row("Tıbbi müdahale:", _yes(incident.medical_intervention))
    row("İş göremezlik:", _yes(incident.work_incapacity_report))
    row("Ekipman hasarı:", _yes(incident.equipment_damage))
    if incident.event_type == "is_kazasi":
        row("SGK bildirimi:", f"{_yes(incident.sgk_reported)} ({_fmt(incident.sgk_report_date)})")
        row("Kolluk bildirimi:", _yes(incident.police_reported))
        row("Kaza türü:", incident.accident_type or "—")
        row("Yaralanma türü:", incident.injury_type or "—")
        row("Müdahale detayı:", incident.intervention_detail or "—")
        row("Rapor günü:", str(incident.report_days or 0))

    if incident.auto_warning:
        elements.append(Paragraph("3. OTOMATİK UYARI", h))
        elements.append(Paragraph((incident.auto_warning or "").replace("\n", "<br/>"), body))

    elements.append(Paragraph("4. KÖK NEDEN ANALİZİ", h))
    if root_cause:
        for i, key in enumerate(["why_1", "why_2", "why_3", "why_4", "why_5"], 1):
            row(f"{i}. Neden:", getattr(root_cause, key, None) or "—")
        row("Kök neden:", root_cause.root_cause or "—")
        row("Kategori:", root_cause.root_cause_category or "—")
        row("Sistemsel eksiklik:", root_cause.systemic_gap or "—")
    else:
        elements.append(Paragraph("Kök neden analizi yapılmamış.", body))

    elements.append(Paragraph("5. RİSK DEĞERLENDİRME", h))
    row(
        "Risk analizinde:",
        risk_opts.get(incident.risk_analysis_status or "", incident.risk_analysis_status or "—"),
    )
    row("Not:", incident.risk_analysis_note or "—")
    row(
        "Olasılık / Şiddet / Skor:",
        f"{incident.probability} / {incident.severity} / {incident.risk_score} ({incident.risk_level or '—'})",
    )
    row("Acil durum ilgisi:", incident.emergency_relation or "—")
    row("Acil durum notu:", incident.emergency_note or "—")

    elements.append(Paragraph("6. DÜZELTİCİ / ÖNLEYİCİ FAALİYETLER", h))
    if dofs:
        for d in dofs:
            elements.append(Paragraph(f"<b>{d.dof_no}</b> — {d.status}", label))
            row("Tespit:", d.finding)
            row("Kök neden:", d.root_cause or "—")
            row("Düzeltici:", d.corrective_action or "—")
            row("Önleyici:", d.preventive_action or "—")
            row("Sorumlu / Termin:", f"{d.responsible_person or '—'} / {_fmt(d.term_date)}")
            elements.append(Spacer(1, 2 * mm))
    else:
        elements.append(Paragraph("DÖF kaydı oluşturulmamış.", body))

    elements.append(Paragraph("7. GENEL DEĞERLENDİRME", h))
    elements.append(Paragraph(incident.evaluation_text or "Genel değerlendirme metni oluşturulmamış.", body))

    elements.append(Paragraph("8. İMZA ALANLARI", h))
    sign = Table(
        [
            ["Unvan", "Ad Soyad", "Kaşe / İmza"],
            ["Olayı Bildiren", incident.recorded_by_name or " ", " "],
            ["İş Güvenliği Uzmanı", incident.safety_specialist or " ", " "],
            ["İşyeri Hekimi", incident.workplace_physician or " ", " "],
            ["İşveren Vekili", incident.employer_representative or " ", " "],
        ],
        colWidths=[55 * mm, 60 * mm, 55 * mm],
    )
    sign.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#212529")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("TOPPADDING", (0, 1), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 12),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(sign)

    doc.build(elements)
    buf.seek(0)
    return buf.read()
