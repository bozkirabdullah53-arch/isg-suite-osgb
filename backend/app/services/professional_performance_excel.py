"""Formatted XLSX exports for OSGB professional performance reports.

The legacy CSV endpoints remain available for backward compatibility. These
workbooks are intended for people opening reports directly in Microsoft Excel:
Turkish headers, readable labels, frozen panes, filters, widths and separate
sheets for gaps/checks are provided.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session

from app.models.entities import OsgbOrganization
from app.services.osgb_oversight import build_oversight, build_professional_performance

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXPORT_VERSION = "xlsx-v1"

_TYPE_LABELS = {
    "safety_specialist": "İş Güvenliği Uzmanı",
    "workplace_physician": "İşyeri Hekimi",
    "other_health_personnel": "Diğer Sağlık Personeli",
}
_STATUS_LABELS = {
    "ok": "Uygun",
    "warning": "İzlem",
    "critical": "Kritik",
    "unknown": "Belirsiz",
}
_MONTHS_TR = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
    7: "Temmuz",
    8: "Ağustos",
    9: "Eylül",
    10: "Ekim",
    11: "Kasım",
    12: "Aralık",
}

_TEAL = "0F766E"
_TEAL_DARK = "115E59"
_TEAL_LIGHT = "CCFBF1"
_BORDER = "CBD5E1"
_MUTED = "64748B"
_GREEN = "DCFCE7"
_GREEN_TEXT = "166534"
_YELLOW = "FEF3C7"
_YELLOW_TEXT = "92400E"
_RED = "FEE2E2"
_RED_TEXT = "991B1B"
_GREY = "F1F5F9"
_GREY_TEXT = "475569"
_WHITE = "FFFFFF"
_THIN = Side(style="thin", color=_BORDER)


def _safe(value: Any) -> Any:
    """Prevent formula injection while preserving numeric/date cell types."""
    if value is None:
        return ""
    if isinstance(value, (int, float, bool, date, datetime)):
        return value
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _period_label(period: dict[str, Any] | None) -> str:
    data = period or {}
    try:
        month = int(data.get("month") or 0)
        year = int(data.get("year") or 0)
    except (TypeError, ValueError):
        return "—"
    if month in _MONTHS_TR and year:
        return f"{_MONTHS_TR[month]} {year}"
    if month or year:
        return f"{month or '—'}/{year or '—'}"
    return "—"


def _status_label(status: Any) -> str:
    key = str(status or "unknown").strip().lower()
    return _STATUS_LABELS.get(key, str(status or "Belirsiz"))


def _role_label(role: Any) -> str:
    key = str(role or "").strip()
    return _TYPE_LABELS.get(key, key or "—")


def _status_fill(status: Any) -> tuple[PatternFill, Font]:
    key = str(status or "unknown").strip().lower()
    if key == "ok":
        return PatternFill("solid", fgColor=_GREEN), Font(color=_GREEN_TEXT, bold=True)
    if key == "warning":
        return PatternFill("solid", fgColor=_YELLOW), Font(color=_YELLOW_TEXT, bold=True)
    if key == "critical":
        return PatternFill("solid", fgColor=_RED), Font(color=_RED_TEXT, bold=True)
    return PatternFill("solid", fgColor=_GREY), Font(color=_GREY_TEXT, bold=True)


def _set_title(ws, title: str, subtitle: str, last_col: int) -> None:
    end = get_column_letter(last_col)
    ws.merge_cells(f"A1:{end}1")
    ws["A1"] = _safe(title)
    ws["A1"].font = Font(size=18, bold=True, color=_WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=_TEAL)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{end}2")
    ws["A2"] = _safe(subtitle)
    ws["A2"].font = Font(size=10, color=_TEAL_DARK)
    ws["A2"].fill = PatternFill("solid", fgColor=_TEAL_LIGHT)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28


def _write_kpis(ws, start_row: int, items: Iterable[tuple[str, Any]], last_col: int) -> int:
    values = list(items)
    if not values:
        return start_row
    width = max(2, last_col // len(values))
    col = 1
    for index, (label, value) in enumerate(values):
        start_col = col
        end_col = last_col if index == len(values) - 1 else min(last_col, col + width - 1)
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 1, end_column=end_col)
        label_cell = ws.cell(start_row, start_col, _safe(label))
        value_cell = ws.cell(start_row + 1, start_col, _safe(value))
        for row_cell in (label_cell, value_cell):
            row_cell.fill = PatternFill("solid", fgColor="F8FAFC")
            row_cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
            row_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.font = Font(size=9, color=_MUTED, bold=True)
        value_cell.font = Font(size=16, color=_TEAL_DARK, bold=True)
        col = end_col + 1
    ws.row_dimensions[start_row].height = 22
    ws.row_dimensions[start_row + 1].height = 27
    return start_row + 3


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row, col, _safe(text))
        cell.fill = PatternFill("solid", fgColor=_TEAL_DARK)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    ws.row_dimensions[row].height = 30


def _style_rows(ws, start_row: int, end_row: int, cols: int, wrap_cols: set[int] | None = None) -> None:
    wraps = wrap_cols or set()
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
            for col in range(1, cols + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor="F8FAFC")
        for col in range(1, cols + 1):
            cell = ws.cell(row, col)
            cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if col in {1, 4, 6, 7, 8, 9, 10, 11, 12} else "left",
                wrap_text=col in wraps,
            )


def _add_table(ws, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _apply_score_rules(ws, cell_range: str) -> None:
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["55"], fill=PatternFill("solid", fgColor=_RED)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["55", "84"], fill=PatternFill("solid", fgColor=_YELLOW)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["85"], fill=PatternFill("solid", fgColor=_GREEN)),
    )


def _save(wb: Workbook) -> bytes:
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def build_professional_performance_roster_xlsx(
    db: Session,
    osgb_id: int | None = None,
) -> tuple[bytes, str]:
    overview = build_oversight(db, osgb_id=osgb_id)
    oid = int(osgb_id or 0)
    osgb = db.get(OsgbOrganization, oid) if oid else None
    osgb_name = osgb.name if osgb else "OSGB"
    period = overview.get("period") or {}
    period_label = _period_label(period)
    generated = datetime.now().strftime("%d.%m.%Y %H:%M")
    professionals = list(overview.get("professionals") or [])
    summary = overview.get("summary") or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Performans Özeti"
    _set_title(
        ws,
        "ÇSGB Profesyonel Performans Raporu",
        f"OSGB: {osgb_name} · Dönem: {period_label} · Oluşturulma: {generated} · Sürüm: {EXPORT_VERSION}",
        12,
    )
    header_row = _write_kpis(
        ws,
        4,
        [
            ("Profesyonel", summary.get("professionals", len(professionals))),
            ("Uygun", summary.get("ok", 0)),
            ("İzlem", summary.get("warning", 0)),
            ("Kritik", summary.get("critical", 0)),
            ("Görevlendirmesiz", summary.get("unassigned", 0)),
        ],
        12,
    )
    ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=12)
    ws.cell(header_row, 1, "Puan ve durumlar seçilen dönemde sistemde bulunan kayıtlar üzerinden hesaplanır; resmî API onayı anlamına gelmez.")
    ws.cell(header_row, 1).font = Font(size=9, italic=True, color=_MUTED)
    ws.cell(header_row, 1).alignment = Alignment(wrap_text=True)
    header_row += 2

    headers = [
        "Sıra",
        "Ad Soyad",
        "Unvan",
        "Belge Sınıfı",
        "Belge Numarası",
        "Performans Puanı (%)",
        "Durum",
        "Sorumlu İşyeri",
        "Eksik Kontrol",
        "Görevlendirme",
        "Kayıt Durumu",
        "Dönem",
    ]
    _write_header(ws, header_row, headers)
    first_data_row = header_row + 1
    for index, pro in enumerate(professionals, start=1):
        row = first_data_row + index - 1
        values = [
            index,
            pro.get("full_name"),
            _role_label(pro.get("professional_type")),
            pro.get("certificate_class") or "—",
            pro.get("certificate_number") or "—",
            int(pro.get("score") or 0),
            _status_label(pro.get("status")),
            int(pro.get("firm_count") or 0),
            int(pro.get("gap_count") if pro.get("gap_count") is not None else len(pro.get("gaps") or [])),
            "Görevlendirme Yok" if pro.get("unassigned") else "Görevlendirme Var",
            "Aktif" if pro.get("is_active", True) else "Pasif",
            period_label,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row, col, _safe(value))
        fill, font = _status_fill(pro.get("status"))
        ws.cell(row, 7).fill = fill
        ws.cell(row, 7).font = font
        if pro.get("unassigned"):
            ws.cell(row, 10).fill = PatternFill("solid", fgColor=_RED)
            ws.cell(row, 10).font = Font(color=_RED_TEXT, bold=True)

    last_data_row = first_data_row + len(professionals) - 1
    if professionals:
        _style_rows(ws, first_data_row, last_data_row, 12, wrap_cols={2, 3, 5, 10})
        _add_table(ws, f"A{header_row}:L{last_data_row}", "ProfessionalPerformanceSummary")
        _apply_score_rules(ws, f"F{first_data_row}:F{last_data_row}")
    else:
        ws.cell(first_data_row, 1, "Bu OSGB için profesyonel kaydı bulunamadı.")
        ws.merge_cells(start_row=first_data_row, start_column=1, end_row=first_data_row, end_column=12)

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:L{max(header_row, last_data_row)}"
    ws.sheet_view.showGridLines = False
    _set_widths(
        ws,
        {
            "A": 7,
            "B": 28,
            "C": 25,
            "D": 14,
            "E": 20,
            "F": 19,
            "G": 13,
            "H": 16,
            "I": 15,
            "J": 22,
            "K": 14,
            "L": 16,
        },
    )

    gaps_ws = wb.create_sheet("Eksik Kontroller")
    _set_title(
        gaps_ws,
        "Eksik / Tamamlanmayan Kontroller",
        f"OSGB: {osgb_name} · Dönem: {period_label} · Her satır bir profesyonel–işyeri kontrol eksiğini gösterir.",
        7,
    )
    gaps_header = 4
    gap_headers = ["Sıra", "Profesyonel", "Unvan", "İşyeri", "Eksik Kontrol", "Açıklama", "Mevzuat Dayanağı"]
    _write_header(gaps_ws, gaps_header, gap_headers)
    gap_row = gaps_header + 1
    gap_index = 0
    for pro in professionals:
        for gap in pro.get("gaps") or []:
            gap_index += 1
            values = [
                gap_index,
                pro.get("full_name"),
                _role_label(pro.get("professional_type")),
                gap.get("company_name") or "—",
                gap.get("check_title") or gap.get("check_code") or "—",
                gap.get("detail") or "—",
                gap.get("legal") or "—",
            ]
            for col, value in enumerate(values, start=1):
                gaps_ws.cell(gap_row, col, _safe(value))
            gap_row += 1
    if gap_index:
        _style_rows(gaps_ws, gaps_header + 1, gap_row - 1, 7, wrap_cols={2, 3, 4, 5, 6, 7})
        _add_table(gaps_ws, f"A{gaps_header}:G{gap_row - 1}", "ProfessionalPerformanceGaps")
    else:
        gaps_ws.cell(gaps_header + 1, 1, "Eksik kontrol bulunamadı.")
        gaps_ws.merge_cells(start_row=gaps_header + 1, start_column=1, end_row=gaps_header + 1, end_column=7)
    gaps_ws.freeze_panes = f"A{gaps_header + 1}"
    gaps_ws.sheet_view.showGridLines = False
    _set_widths(gaps_ws, {"A": 7, "B": 28, "C": 24, "D": 28, "E": 26, "F": 55, "G": 55})

    checks_ws = wb.create_sheet("Kontrol Özeti")
    _set_title(
        checks_ws,
        "Kontrol Alanı Başarı Özeti",
        f"OSGB: {osgb_name} · Dönem: {period_label} · Başarı yüzdesi profesyonelin sorumlu işyerleri üzerinden hesaplanır.",
        8,
    )
    checks_header = 4
    check_headers = ["Sıra", "Profesyonel", "Unvan", "Kontrol Alanı", "Tamamlanan", "Toplam", "Başarı (%)", "Mevzuat Dayanağı"]
    _write_header(checks_ws, checks_header, check_headers)
    check_row = checks_header + 1
    check_index = 0
    for pro in professionals:
        for check in pro.get("check_columns") or []:
            check_index += 1
            values = [
                check_index,
                pro.get("full_name"),
                _role_label(pro.get("professional_type")),
                check.get("title") or check.get("code") or "—",
                int(check.get("passed") or 0),
                int(check.get("total") or 0),
                int(check.get("pct") or 0),
                check.get("legal") or "—",
            ]
            for col, value in enumerate(values, start=1):
                checks_ws.cell(check_row, col, _safe(value))
            fill, font = _status_fill(check.get("status"))
            checks_ws.cell(check_row, 7).fill = fill
            checks_ws.cell(check_row, 7).font = font
            check_row += 1
    if check_index:
        _style_rows(checks_ws, checks_header + 1, check_row - 1, 8, wrap_cols={2, 3, 4, 8})
        _add_table(checks_ws, f"A{checks_header}:H{check_row - 1}", "ProfessionalPerformanceChecks")
        _apply_score_rules(checks_ws, f"G{checks_header + 1}:G{check_row - 1}")
    else:
        checks_ws.cell(checks_header + 1, 1, "Kontrol verisi bulunamadı.")
        checks_ws.merge_cells(start_row=checks_header + 1, start_column=1, end_row=checks_header + 1, end_column=8)
    checks_ws.freeze_panes = f"A{checks_header + 1}"
    checks_ws.sheet_view.showGridLines = False
    _set_widths(checks_ws, {"A": 7, "B": 28, "C": 24, "D": 28, "E": 14, "F": 12, "G": 14, "H": 58})

    info_ws = wb.create_sheet("Açıklamalar")
    _set_title(info_ws, "Rapor Açıklamaları", "Bu sayfa rapordaki alanların ne anlama geldiğini açıklar.", 4)
    info_rows = [
        ("Durum", "Uygun", "Başarı oranı %85 ve üzeri.", "Yeşil"),
        ("Durum", "İzlem", "Başarı oranı %55–84 aralığı.", "Sarı"),
        ("Durum", "Kritik", "Başarı oranı %55'in altında veya zorunlu görevlendirme yok.", "Kırmızı"),
        ("Durum", "Belirsiz", "Değerlendirme için yeterli dönem verisi bulunmuyor.", "Gri"),
        ("Puan", "Performans Puanı", "Ağırlıklı kontrol sonuçlarından hesaplanır.", "0–100"),
        ("Kapsam", "Eksik Kontrol", "Profesyonelin sorumlu işyerlerindeki tamamlanmayan kontrol sayısıdır.", "Adet"),
        ("Not", "Resmî entegrasyon", "Rapor sistem kayıtlarından üretilir; İBYS/KATİP veya ÇSGB tarafından verilmiş resmî onay değildir.", "Bilgilendirme"),
    ]
    _write_header(info_ws, 4, ["Alan", "Değer", "Açıklama", "Gösterim"])
    for row_index, values in enumerate(info_rows, start=5):
        for col, value in enumerate(values, start=1):
            info_ws.cell(row_index, col, _safe(value))
    _style_rows(info_ws, 5, 4 + len(info_rows), 4, wrap_cols={2, 3, 4})
    _add_table(info_ws, f"A4:D{4 + len(info_rows)}", "ProfessionalPerformanceNotes")
    info_ws.sheet_view.showGridLines = False
    _set_widths(info_ws, {"A": 16, "B": 24, "C": 72, "D": 20})

    stamp = date.today().isoformat()
    filename = f"csgb-profesyonel-performans-{oid or 'osgb'}-{stamp}.xlsx"
    return _save(wb), filename


def build_professional_performance_detail_xlsx(
    db: Session,
    professional_id: int,
) -> tuple[bytes, str]:
    report = build_professional_performance(db, professional_id)
    pro = report.get("professional") or {}
    perf = report.get("performance") or {}
    period_label = _period_label(report.get("period") or {})
    generated = datetime.now().strftime("%d.%m.%Y %H:%M")
    name = str(pro.get("full_name") or f"Profesyonel {professional_id}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Profesyonel Özeti"
    _set_title(
        ws,
        f"Profesyonel Performans Raporu — {name}",
        f"Dönem: {period_label} · Oluşturulma: {generated} · Sürüm: {EXPORT_VERSION}",
        8,
    )
    row = _write_kpis(
        ws,
        4,
        [
            ("Performans Puanı", f"%{int(perf.get('score') or 0)}"),
            ("Durum", _status_label(perf.get("status"))),
            ("İşyeri", int(perf.get("firm_count") or 0)),
            ("Tamamlanan", int(perf.get("completed_checks") or 0)),
            ("Eksik", int(perf.get("gap_count") or 0)),
        ],
        8,
    )
    identity = [
        ("Ad Soyad", name),
        ("Unvan", pro.get("role_label") or _role_label(pro.get("professional_type"))),
        ("Belge Sınıfı", pro.get("certificate_class") or "—"),
        ("Belge Numarası", pro.get("certificate_number") or "—"),
        ("Belge Tarihi", pro.get("certificate_date") or "—"),
        ("E-posta", pro.get("email") or "—"),
        ("Telefon", pro.get("phone") or "—"),
        ("Kayıt Durumu", "Aktif" if pro.get("is_active", True) else "Pasif"),
    ]
    _write_header(ws, row, ["Alan", "Bilgi", "Alan", "Bilgi", "Alan", "Bilgi", "Alan", "Bilgi"])
    info_row = row + 1
    for index in range(0, len(identity), 4):
        chunk = identity[index:index + 4]
        col = 1
        for label, value in chunk:
            ws.cell(info_row, col, _safe(label))
            ws.cell(info_row, col + 1, _safe(value))
            ws.cell(info_row, col).font = Font(bold=True, color=_TEAL_DARK)
            ws.cell(info_row, col).fill = PatternFill("solid", fgColor=_TEAL_LIGHT)
            for target in (ws.cell(info_row, col), ws.cell(info_row, col + 1)):
                target.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
                target.alignment = Alignment(vertical="top", wrap_text=True)
            col += 2
        info_row += 1
    ws.sheet_view.showGridLines = False
    _set_widths(ws, {"A": 18, "B": 28, "C": 18, "D": 24, "E": 18, "F": 26, "G": 18, "H": 28})

    def _control_sheet(title: str, sheet_name: str, rows: list[dict[str, Any]], table_name: str) -> None:
        sheet = wb.create_sheet(sheet_name)
        _set_title(sheet, title, f"Profesyonel: {name} · Dönem: {period_label}", 6)
        headers = ["Sıra", "İşyeri", "Kontrol", "Açıklama", "Mevzuat Dayanağı", "Sonuç"]
        _write_header(sheet, 4, headers)
        for index, item in enumerate(rows, start=1):
            values = [
                index,
                item.get("company_name") or "—",
                item.get("check_title") or item.get("check_code") or "—",
                item.get("detail") or "—",
                item.get("legal") or "—",
                "Eksik" if sheet_name == "Eksik Kontroller" else "Tamamlandı",
            ]
            for col, value in enumerate(values, start=1):
                sheet.cell(4 + index, col, _safe(value))
            if sheet_name == "Eksik Kontroller":
                sheet.cell(4 + index, 6).fill = PatternFill("solid", fgColor=_RED)
                sheet.cell(4 + index, 6).font = Font(color=_RED_TEXT, bold=True)
            else:
                sheet.cell(4 + index, 6).fill = PatternFill("solid", fgColor=_GREEN)
                sheet.cell(4 + index, 6).font = Font(color=_GREEN_TEXT, bold=True)
        if rows:
            _style_rows(sheet, 5, 4 + len(rows), 6, wrap_cols={2, 3, 4, 5})
            _add_table(sheet, f"A4:F{4 + len(rows)}", table_name)
        else:
            sheet.cell(5, 1, "Kayıt bulunamadı.")
            sheet.merge_cells("A5:F5")
        sheet.freeze_panes = "A5"
        sheet.sheet_view.showGridLines = False
        _set_widths(sheet, {"A": 7, "B": 30, "C": 28, "D": 58, "E": 58, "F": 16})

    _control_sheet(
        "Eksik / Tamamlanmayan Kontroller",
        "Eksik Kontroller",
        list(report.get("incomplete") or []),
        "ProfessionalDetailGaps",
    )
    _control_sheet(
        "Tamamlanan Kontroller",
        "Tamamlanan",
        list(report.get("completed") or []),
        "ProfessionalDetailCompleted",
    )

    firms_ws = wb.create_sheet("Firma Checklist")
    _set_title(firms_ws, "Firma Bazlı Checklist", f"Profesyonel: {name} · Dönem: {period_label}", 8)
    firm_headers = ["Sıra", "İşyeri", "Tehlike Sınıfı", "Firma Puanı (%)", "Kontrol", "Sonuç", "Açıklama", "Mevzuat Dayanağı"]
    _write_header(firms_ws, 4, firm_headers)
    firm_row = 5
    firm_index = 0
    for firm in report.get("firms") or []:
        for check in firm.get("checks") or []:
            firm_index += 1
            values = [
                firm_index,
                firm.get("company_name") or "—",
                firm.get("hazard_class") or "—",
                int(firm.get("score") or 0),
                check.get("title") or check.get("code") or "—",
                "Tamamlandı" if check.get("passed") else "Eksik",
                check.get("detail") or "—",
                check.get("legal") or "—",
            ]
            for col, value in enumerate(values, start=1):
                firms_ws.cell(firm_row, col, _safe(value))
            if check.get("passed"):
                firms_ws.cell(firm_row, 6).fill = PatternFill("solid", fgColor=_GREEN)
                firms_ws.cell(firm_row, 6).font = Font(color=_GREEN_TEXT, bold=True)
            else:
                firms_ws.cell(firm_row, 6).fill = PatternFill("solid", fgColor=_RED)
                firms_ws.cell(firm_row, 6).font = Font(color=_RED_TEXT, bold=True)
            firm_row += 1
    if firm_index:
        _style_rows(firms_ws, 5, firm_row - 1, 8, wrap_cols={2, 3, 5, 7, 8})
        _add_table(firms_ws, f"A4:H{firm_row - 1}", "ProfessionalFirmChecklist")
        _apply_score_rules(firms_ws, f"D5:D{firm_row - 1}")
    else:
        firms_ws.cell(5, 1, "Atanmış işyeri veya firma kontrol verisi bulunamadı.")
        firms_ws.merge_cells("A5:H5")
    firms_ws.freeze_panes = "A5"
    firms_ws.sheet_view.showGridLines = False
    _set_widths(firms_ws, {"A": 7, "B": 30, "C": 18, "D": 18, "E": 28, "F": 16, "G": 58, "H": 58})

    stamp = date.today().isoformat()
    filename = f"csgb-profesyonel-performans-detay-{professional_id}-{stamp}.xlsx"
    return _save(wb), filename
